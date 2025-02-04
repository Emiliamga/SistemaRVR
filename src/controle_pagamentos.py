import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
from openpyxl import load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
import os
import sys
from pathlib import Path

# Ajusta o path para incluir o diretório raiz
sys.path.append(str(Path(__file__).parent.parent))

from config.utils import (
    PASTA_CLIENTES,
    formatar_moeda,
    validar_data
)

class ControlePagamentos:
    def __init__(self, parent=None):
        self.parent = parent
        self.root = tk.Toplevel(parent) if parent else tk.Tk()
        self.root.title("Controle de Pagamentos de Taxas")
        self.root.geometry("1200x700")
        
        # Variáveis de controle
        self.cliente_selecionado = None
        self.parcelas_selecionadas = []
        self.scrollbar_y = None
        self.scrollbar_x = None
        
        self.setup_gui()
    
    def setup_gui(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill='both', expand=True)
        
        # Frame seleção de cliente
        frame_cliente = ttk.LabelFrame(main_frame, text="Selecione o Cliente")
        frame_cliente.pack(fill='x', pady=5)
        
        self.cliente_combo = ttk.Combobox(frame_cliente, state='readonly')
        self.cliente_combo.pack(side='left', padx=5)
        self.cliente_combo.bind('<<ComboboxSelected>>', self.carregar_parcelas)
        
        # Frame lista de parcelas com scrollbars
        self.frame_parcelas = ttk.LabelFrame(main_frame, text="Parcelas Pendentes")
        self.frame_parcelas.pack(fill='both', expand=True, pady=5)
        
        # Container para treeview e scrollbars
        self.tree_container = ttk.Frame(self.frame_parcelas)
        self.tree_container.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Treeview
        colunas = ('Nº Contrato', 'Nº Parcela', 'CNPJ', 'Adm', 'Vencimento Original', 
                   'Valor', 'Status', 'Data Pagamento')
        self.tree_parcelas = ttk.Treeview(self.tree_container, columns=colunas, show='headings')
        
        # Configurar colunas
        for col in colunas:
            self.tree_parcelas.heading(col, text=col)
            self.tree_parcelas.column(col, width=100)
        
        # Scrollbars
        self.scrollbar_y = ttk.Scrollbar(self.tree_container, orient='vertical',
                                       command=self.tree_parcelas.yview)
        self.scrollbar_x = ttk.Scrollbar(self.tree_container, orient='horizontal',
                                       command=self.tree_parcelas.xview)
        
        # Configurar treeview
        self.tree_parcelas.configure(yscrollcommand=self.scrollbar_y.set,
                                   xscrollcommand=self.scrollbar_x.set)
        
        # Grid layout para treeview e scrollbars
        self.tree_parcelas.grid(row=0, column=0, sticky='nsew')
        self.scrollbar_y.grid(row=0, column=1, sticky='ns')
        self.scrollbar_x.grid(row=1, column=0, sticky='ew')
        
        # Configurar grid weights
        self.tree_container.grid_rowconfigure(0, weight=1)
        self.tree_container.grid_columnconfigure(0, weight=1)
        
        # Frame para registrar pagamento
        frame_pagamento = ttk.LabelFrame(main_frame, text="Registrar Pagamento")
        frame_pagamento.pack(fill='x', pady=5)
        
        ttk.Label(frame_pagamento, text="Data do Pagamento:").pack(side='left', padx=5)
        self.data_pagamento = DateEntry(frame_pagamento, width=12,
                                      background='darkblue', foreground='white',
                                      borderwidth=2, date_pattern='dd/mm/yyyy')
        self.data_pagamento.pack(side='left', padx=5)
        
        ttk.Button(frame_pagamento, text="Registrar Pagamento",
                  command=self.registrar_pagamento).pack(side='left', padx=5)
        
        # Frame de botões
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x', pady=5)
        
        ttk.Button(frame_botoes, text="Voltar ao Menu",
                  command=self.voltar_menu).pack(side='right', padx=5)
        
        # Carregar lista de clientes
        self.carregar_clientes()
        
    def tem_taxa_fixa(self, arquivo_cliente):
        """
        Verifica se o cliente possui contratos com taxa fixa.
        Considera a estrutura:
        - Linha N: Contrato ATIVO
        - Linha N+1: Informações do administrador e tipo da taxa
        - Linha N+2: Parcelas
        
        Args:
            arquivo_cliente: Path do arquivo do cliente
            
        Returns:
            bool: True se o cliente tem pelo menos um contrato com taxa fixa ativo
        """
        try:
            print(f"\nVerificando taxa fixa para: {arquivo_cliente}")
            wb = load_workbook(arquivo_cliente)
            if 'Contratos_ADM' not in wb.sheetnames:
                print("Aba Contratos_ADM não encontrada")
                wb.close()
                return False
                
            ws = wb['Contratos_ADM']
            
            # Converter todas as linhas em lista para facilitar a navegação
            rows = list(ws.iter_rows(min_row=3, values_only=True))
            
            for i in range(len(rows) - 1):  # -1 para evitar IndexError na última linha
                row = rows[i]
                
                # Se encontrou um contrato com status ATIVO
                if row[0] and row[3] == 'ATIVO':
                    num_contrato = row[0]
                    print(f"Contrato ativo encontrado: {num_contrato}")
                    
                    # Verificar a próxima linha para o tipo de taxa
                    if i + 1 < len(rows):
                        next_row = rows[i + 1]
                        # Verifica se é a linha de administrador (tem o mesmo número de contrato na coluna G)
                        if next_row[6] == num_contrato and next_row[9] == 'Fixo':
                            print(f"Taxa fixa encontrada para contrato {num_contrato}")
                            wb.close()
                            return True
            
            print("Nenhuma taxa fixa encontrada")
            wb.close()
            return False
            
        except Exception as e:
            print(f"Erro ao verificar taxa fixa: {str(e)}")
            if 'wb' in locals():
                wb.close()
            return False

    def carregar_clientes(self):
        """Carrega a lista de clientes que possuem contratos com taxa fixa ativa"""
        clientes = []
        print("\nCarregando clientes com taxa fixa...")
        
        for arquivo in os.listdir(PASTA_CLIENTES):
            if arquivo.endswith('.xlsx'):
                try:
                    arquivo_path = PASTA_CLIENTES / arquivo
                    if self.tem_taxa_fixa(arquivo_path):
                        nome_cliente = arquivo.replace('.xlsx', '')
                        clientes.append(nome_cliente)
                        print(f"Cliente com taxa fixa encontrado: {nome_cliente}")
                        
                except Exception as e:
                    print(f"Erro ao verificar arquivo {arquivo}: {str(e)}")
        
        print(f"\nTotal de clientes com taxa fixa: {len(clientes)}")
        self.cliente_combo['values'] = sorted(clientes)


    def carregar_parcelas(self, event=None):
        """Carrega as parcelas do cliente selecionado"""
        cliente = self.cliente_combo.get()
        if not cliente:
            return
            
        try:
            print(f"\nCarregando parcelas para {cliente}")
            
            # Limpar lista atual
            for item in self.tree_parcelas.get_children():
                self.tree_parcelas.delete(item)
                
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            print(f"Arquivo: {arquivo_cliente}")
            
            wb = load_workbook(arquivo_cliente)
            ws = wb['Contratos_ADM']
            
            print("\nBuscando parcelas...")
            parcelas_encontradas = 0
            
            # Buscar parcelas
            for row in ws.iter_rows(min_row=3, values_only=True):
                num_contrato = row[24]  # Coluna Y - Número do contrato
                num_parcela = row[25]   # Coluna Z - Número da parcela
                cnpj_cpf = row[26]      # Coluna AA - CNPJ/CPF
                nome = row[27]          # Coluna AB - Nome
                dt_vencto = row[28]     # Coluna AC - Data vencimento
                valor = row[29]         # Coluna AD - Valor
                status = row[30] if len(row) > 30 else "PENDENTE"  # Coluna AE - Status
                dt_pagto = row[31] if len(row) > 31 else None      # Coluna AF - Data pagamento
                
                # Se tiver número do contrato, é uma linha válida
                if num_contrato:
                    print("\nParcela encontrada!")
                    print(f"Contrato: {num_contrato}")
                    print(f"Parcela: {num_parcela}")
                    print(f"CNPJ/CPF: {cnpj_cpf}")
                    print(f"Nome: {nome}")
                    print(f"Vencimento: {dt_vencto}")
                    print(f"Valor: {valor}")
                    print(f"Status: {status}")
                    print(f"Data Pagamento: {dt_pagto}")
                    
                    # Formatar datas
                    dt_vencto_str = dt_vencto.strftime('%d/%m/%Y') if isinstance(dt_vencto, datetime) else str(dt_vencto)
                    dt_pagto_str = dt_pagto.strftime('%d/%m/%Y') if isinstance(dt_pagto, datetime) else ""
                    
                    # Formatar valor usando formato brasileiro
                    valor_str = f"{float(valor):,.2f}".replace(',', '_').replace('.', ',').replace('_', '.') if valor else ""
                    
                    # Inserir na treeview
                    self.tree_parcelas.insert('', 'end', values=(
                        num_contrato,
                        num_parcela,
                        cnpj_cpf,
                        nome,           
                        dt_vencto_str,
                        valor_str,
                        status or "PENDENTE",
                        dt_pagto_str
                    ))
                    
                    parcelas_encontradas += 1
            
            print(f"\nTotal de parcelas encontradas: {parcelas_encontradas}")
            wb.close()
            
        except Exception as e:
            print(f"\nERRO ao carregar parcelas: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao carregar parcelas: {str(e)}")
            if 'wb' in locals():
                wb.close()

    def registrar_pagamento(self):
        """Registra o pagamento das parcelas selecionadas"""
        print("\nIniciando registro de pagamento...")
        
        selecionados = self.tree_parcelas.selection()
        if not selecionados:
            print("Nenhuma parcela selecionada")
            messagebox.showwarning("Aviso", "Selecione as parcelas para pagamento!")
            return
            
        print(f"Parcelas selecionadas: {len(selecionados)}")
        
        data_pagto = self.data_pagamento.get_date()
        print(f"Data de pagamento: {data_pagto}")
        
        if not validar_data(data_pagto.strftime('%d/%m/%Y')):
            print("Data inválida")
            messagebox.showerror("Erro", "Data de pagamento inválida!")
            return
            
        cliente = self.cliente_combo.get()
        if not cliente:
            print("Cliente não selecionado")
            return
            
        print(f"Cliente: {cliente}")
        
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            print(f"Arquivo: {arquivo_cliente}")
            
            wb = load_workbook(arquivo_cliente)
            ws_contratos = wb['Contratos_ADM']
            ws_dados = wb['Dados']
            
            parcelas_processadas = []
            
            print("\nProcessando pagamentos...")
            
            # Para cada parcela selecionada
            for item in selecionados:
                valores = self.tree_parcelas.item(item)['values']
                num_contrato = str(valores[0])
                num_parcela = int(valores[1])
                cnpj_cpf = str(valores[2])
                nome = str(valores[3])
                
                # Buscar total de parcelas para este contrato
                total_parcelas = 0
                for row in ws_contratos.iter_rows(min_row=3, values_only=True):
                    if str(row[24]) == num_contrato:  # Mesmo número de contrato
                        total_parcelas += 1
                
                print(f"\nProcessando parcela {num_parcela}/{total_parcelas} do contrato {num_contrato}")
                
                # Atualizar na aba Contratos_ADM
                for row_idx, row in enumerate(ws_contratos.iter_rows(min_row=3), start=3):
                    if (str(row[24].value) == num_contrato and
                        int(row[25].value) == num_parcela and
                        str(row[26].value) == cnpj_cpf):
                        
                        # Pegar o valor original da planilha
                        valor = float(row[29].value) if row[29].value else 0
                        
                        # Atualizar status e data de pagamento
                        ws_contratos.cell(row=row_idx, column=31, value='PAGO')
                        ws_contratos.cell(row=row_idx, column=32, value=data_pagto)
                        
                        # Registrar na aba Dados
                        proxima_linha = ws_dados.max_row + 1
                        
                        # Data do relatório (manter data original)
                        data_ref = row[28].value  # Data de vencimento original
                        ws_dados.cell(row=proxima_linha, column=1, value=data_ref)
                        ws_dados.cell(row=proxima_linha, column=1).number_format = 'DD/MM/YYYY'
                        
                        # Tipo e dados
                        ws_dados.cell(row=proxima_linha, column=2, value=2)
                        ws_dados.cell(row=proxima_linha, column=3, value=cnpj_cpf)
                        ws_dados.cell(row=proxima_linha, column=4, value=nome)
                        ws_dados.cell(row=proxima_linha, column=5, value=f"ADM OBRA - PARC. {num_parcela}/{total_parcelas}")
                        
                        # Valores com formato brasileiro
                        ws_dados.cell(row=proxima_linha, column=7, value=valor)
                        ws_dados.cell(row=proxima_linha, column=7).number_format = '#.##0,00'
                        
                        ws_dados.cell(row=proxima_linha, column=8, value=1)
                        
                        ws_dados.cell(row=proxima_linha, column=9, value=valor)
                        ws_dados.cell(row=proxima_linha, column=9).number_format = '#.##0,00'
                        
                        # Data de pagamento
                        ws_dados.cell(row=proxima_linha, column=10, value=data_pagto)
                        ws_dados.cell(row=proxima_linha, column=10).number_format = 'DD/MM/YYYY'
                        
                        ws_dados.cell(row=proxima_linha, column=11, value='ADM')
                        ws_dados.cell(row=proxima_linha, column=12, value='')
                        ws_dados.cell(row=proxima_linha, column=13, value='LANÇAMENTO AUTOMÁTICO')
                        
                        parcelas_processadas.append(f"Contrato {num_contrato} - Parcela {num_parcela}/{total_parcelas}")
                        break
            
            print("\nSalvando alterações...")
            wb.save(arquivo_cliente)
            
            print("\nAtualizando visualização...")
            self.carregar_parcelas()
            
            if parcelas_processadas:
                mensagem = "Pagamentos registrados:\n\n" + "\n".join(parcelas_processadas)
                print(f"\nMensagem de sucesso: {mensagem}")
                messagebox.showinfo("Sucesso", mensagem)
            else:
                print("\nNenhuma parcela foi processada!")
                messagebox.showwarning("Aviso", "Nenhuma parcela foi processada!")
            
        except Exception as e:
            print(f"\nERRO: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao registrar pagamento: {str(e)}")
            if 'wb' in locals():
                wb.close()

    def voltar_menu(self):
        """Fecha a janela e retorna ao menu principal"""
        self.root.destroy()
        if self.parent:
            self.parent.deiconify()

    def run(self):
        """Inicia a execução do sistema"""
        self.root.mainloop()


if __name__ == "__main__":
    app = ControlePagamentos()
    app.run()
