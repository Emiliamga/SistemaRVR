import sys
import os
import pandas as pd
import xlwings as xw
import openpyxl
import warnings
import platform
import subprocess
import tkinter as tk
from tkinter import Tk
from openpyxl import load_workbook
from tkinter import ttk, messagebox, filedialog, StringVar, Toplevel, BooleanVar
from tkcalendar import Calendar
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from reportlab.lib.pagesizes import landscape, A4
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, 
    PageTemplate, Frame, Spacer, PageBreak, Image
)
from reportlab.lib.enums import TA_LEFT
from reportlab.lib import colors
from reportlab.platypus import KeepTogether

# Configuração inicial
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


# Variáveis globais
arquivo_path = None
arquivo_selecionado = None
data_selecionada = None
incluir_futuros = None
status_label = None
root = None
handler = None



class RelatorioUI:
    def __init__(self, parent):
        print(f"Iniciando __init__ com parent: {parent}")
        if parent is None:
            self.root = tk.Tk()
        else:
            self.root = parent
            
        print("Criando StringVars...")
        self.arquivo_selecionado = StringVar(self.root, value="Nenhum arquivo selecionado")
        self.data_selecionada = StringVar(self.root, value=datetime.now().strftime('%d/%m/%Y'))
        print(f"StringVars criados. Data: {self.data_selecionada.get()}")
        
        self.incluir_futuros = BooleanVar(value=True)
        self.status_label = None
        self.handler = RelatorioHandler()
        self.arquivos_lote = []
        self.menu_principal = None  # Adicionado aqui, antes do setup_ui
        self.setup_ui()

    def setup_ui(self):
        print(f"Iniciando setup_ui, data_selecionada: {self.data_selecionada.get()}")
        self.root.title("Gerador de Relatório de Despesas")
        self.root.geometry("500x400")
        self.root.update_idletasks()


        # Container principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill='both', expand=True)

        # Data
        frame_data = ttk.Frame(main_frame)
        frame_data.pack(pady=10, padx=20, fill='x')
        
        self.data_selecionada.set(datetime.now().strftime('%d/%m/%Y'))
        self.arquivo_selecionado.set("Nenhum arquivo selecionado")
        ttk.Label(frame_data, text="Data do relatório:").pack(side='left', padx=(0, 10))
        ttk.Label(frame_data, textvariable=self.data_selecionada, width=10).pack(side='left')
        ttk.Button(frame_data, text="Escolher Data", command=self.escolher_data).pack(side='left', padx=5)

        # Relatório Individual
        frame_arquivo = ttk.LabelFrame(main_frame, text="Relatório Individual")
        frame_arquivo.pack(pady=10, padx=20, fill='x')

        self.arquivo_selecionado.set("Nenhum arquivo selecionado")
        ttk.Button(frame_arquivo, text="Escolher arquivo", 
                  command=self.selecionar_arquivo_local).pack(pady=5, fill='x')
        ttk.Label(frame_arquivo, textvariable=self.arquivo_selecionado).pack(pady=5)
        ttk.Button(frame_arquivo, text="Gerar Relatório Individual",
                  command=self.gerar_relatorio).pack(pady=5, fill='x')

        # Relatório em Lote
        frame_lote = ttk.LabelFrame(main_frame, text="Relatório em Lote")
        frame_lote.pack(pady=10, padx=20, fill='x')
        ttk.Button(frame_lote, text="Selecionar Arquivos para Lote", 
                  command=self.selecionar_arquivos_lote).pack(pady=5, fill='x')

        # Checkbox para lançamentos futuros
        ttk.Checkbutton(main_frame, text="Incluir lançamentos futuros",
                       variable=self.incluir_futuros).pack(pady=10, anchor='w')

        # Status label
        self.status_label = ttk.Label(main_frame, text="", wraplength=350)
        self.status_label.pack(pady=10)

    def escolher_data(self):
        top = Toplevel(self.root)
        top.title("Selecione a Data")
        
        x = self.root.winfo_x() + 50
        y = self.root.winfo_y() + 50
        top.geometry(f"+{x}+{y}")
        
        cal = Calendar(top,
                      selectmode='day',
                      year=datetime.now().year,
                      month=datetime.now().month,
                      day=datetime.now().day,
                      locale='pt_BR',
                      date_pattern='dd/mm/yyyy')
        cal.pack(padx=10, pady=10)
        
        def definir_data():
            data = cal.get_date()
            self.data_selecionada.set(data)
            top.destroy()
            
        ttk.Button(top, text="Confirmar", command=definir_data).pack(pady=5)
        top.transient(self.root)
        top.grab_set()
        self.root.wait_window(top)

        

    def selecionar_arquivo_local(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o arquivo Excel",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
        )
        if arquivo:
            self.arquivo_path = arquivo
            nome_arquivo = os.path.basename(arquivo)
            self.arquivo_selecionado.set(nome_arquivo)
            self.root.update_idletasks()

    def selecionar_arquivos_lote(self):
        files = filedialog.askopenfilenames(
            title="Selecione os arquivos Excel",
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )
        if files:
            self.arquivos_lote = files
            self.processar_lote(files)

    def gerar_relatorio(self):
        try:
            if not self.arquivo_path:
                self.status_label.config(text="Selecione um arquivo Excel!")
                return

            data_rel = datetime.strptime(self.data_selecionada.get(), '%d/%m/%Y')
            print(f"\nGerando relatório para data: {data_rel}")
                
            # Carregar e processar dados
            df = self.handler.carregar_dados_excel(self.arquivo_path)
            df_filtrado, df_diaria, df_tp_desp_1 = self.handler.processar_dados(df, data_rel)
                
            # Processar lançamentos futuros
            df_futuro = None
            if self.incluir_futuros.get():
                df_futuro = self.handler.processar_lancamentos_futuros(df, data_rel)
                    
            # Processar workbook
            workbook = load_workbook(self.arquivo_path, data_only=True)
            ws_resumo = workbook['RESUMO']
            nome_cliente = ws_resumo['A3'].value
                
            # Obter número do relatório e acumulado
            numero_relatorio = self.handler.obter_numero_relatorio(ws_resumo, data_rel)
            acumulado = self.handler.obter_acumulado(ws_resumo, data_rel)
                
            dados_completos = {
                'df_filtrado': df_filtrado,
                'df_diaria': df_diaria,
                'df_tp_desp_1': df_tp_desp_1,
                'df_futuro': df_futuro,
                'incluir_futuros': self.incluir_futuros.get(),
                'data_relatorio': data_rel,
                'nome_cliente': nome_cliente,
                'endereco_cliente': ws_resumo['A4'].value,
                'numero_relatorio': numero_relatorio,  # Adicionado
                'acumulado': acumulado  # Adicionado
            }
            
            # Gerar nome do arquivo
            data_formatada = data_rel.strftime('%d-%m-%Y')
            nome_arquivo = f"REL - {nome_cliente} - {data_formatada}.pdf"
            caminho_output = os.path.join(os.path.dirname(self.arquivo_path), nome_arquivo)
            
            self.handler.gerar_relatorio_pdf(dados_completos, caminho_output, self.arquivo_path)
            self.status_label.config(text=f"Relatório gerado com sucesso para {nome_cliente}")
            self.criar_dialog_relatorio_gerado(nome_cliente, data_formatada)
            
        except Exception as e:
            self.status_label.config(text=f"Erro: {str(e)}")


    def processar_lote(self, arquivos):
        # Implementar lógica de processamento em lote
        progress_window = Toplevel(self.root)
        progress_window.title("Gerando Relatórios em Lote")
        progress_window.geometry("400x400")
        progress_window.transient(root)

        # Label para mostrar progresso
        progress_label = ttk.Label(progress_window, text="Processando...", font=('Helvetica', 10))
        progress_label.pack(pady=10)

        # Barra de progresso
        progress_bar = ttk.Progressbar(progress_window, length=300, mode='determinate')
        progress_bar.pack(pady=10)

        # Lista para mostrar arquivos processados
        lista_processados = tk.Listbox(progress_window, width=50, height=10)
        lista_processados.pack(pady=10, padx=10)

        # Configurar barra de progresso
        total_arquivos = len(arquivos)
        progress_bar['maximum'] = total_arquivos

        # Processar cada arquivo
        
        for i, arquivo in enumerate(arquivos, 1):
            try:
                arquivo_nome = os.path.basename(arquivo)
                progress_label.config(text=f"Processando: {arquivo_nome}")
                progress_bar['value'] = i
                
                wb = load_workbook(arquivo, data_only=True)
                try:
                    ws_resumo = wb['RESUMO']
                    nome_cliente = ws_resumo['A3'].value
                    
                    data_rel = datetime.strptime(self.data_selecionada.get(), '%d/%m/%Y')
                    
                    df = self.handler.carregar_dados_excel(arquivo)  # Fixed: Use self.handler
                    df_filtrado, df_diaria, df_tp_desp_1 = self.handler.processar_dados(df, data_rel)

                    df_futuro = None
                    if self.incluir_futuros.get():  # Fixed: Use self.incluir_futuros
                        df_futuro = self.handler.processar_lancamentos_futuros(df, data_rel)
                        
                    dados_completos = {
                        'df_filtrado': df_filtrado,
                        'df_diaria': df_diaria,
                        'df_tp_desp_1': df_tp_desp_1,
                        'df_futuro': df_futuro,
                        'incluir_futuros': self.incluir_futuros.get(),
                        'data_relatorio': data_rel,
                        'nome_cliente': nome_cliente,
                        'endereco_cliente': ws_resumo['A4'].value,
                    }
                    
                    # Gerar relatório
                    data_formatada = data_rel.strftime('%d-%m-%Y')
                    nome_arquivo = f"REL - {nome_cliente} - {data_formatada}.pdf"
                    caminho_output = os.path.join(os.path.dirname(arquivo), nome_arquivo)
                    
                    self.handler.gerar_relatorio_pdf(dados_completos, caminho_output, arquivo)
                    
                    lista_processados.insert(tk.END, f"✓ {arquivo_nome} - Concluído")
                    lista_processados.see(tk.END)

                finally:
                    wb.close()  # Garantir que o arquivo seja fechado


            except Exception as e:
                lista_processados.insert(tk.END, f"✗ {arquivo_nome} - Erro: {str(e)}")

            # Atualizar interface
            progress_window.update()

         # Finalização
        progress_label.config(text="Processamento concluído!")
        ttk.Button(progress_window, 
                   text="Fechar", 
                   command=lambda: self.criar_dialog_relatorio_gerado(None, None) or progress_window.destroy()).pack(pady=10)


            
    def gerar_relatorio_lote():
        try:
            # Verificar se há arquivos selecionados
            if not self.arquivo_path:  # Usar self em vez de variável global
                self.status_label.config(text="Selecione um arquivo Excel!")
                return
            
            processar_lote(arquivos_selecionados)


            status_label.config(text="Relatórios em lote gerados com sucesso!")

            # Criar diálogo após gerar os relatórios em lote
            # criar_dialog_relatorio_gerado(None, None)

        except Exception as e:
            erro = str(e)
            print(f"Erro ao gerar relatórios em lote: {erro}")
            status_label.config(text=f"Erro: {erro}")


    def criar_dialog_relatorio_gerado(self, nome_cliente, data_formatada):
        dialog = Toplevel(self.root)
        dialog.title("Relatório Gerado")
        dialog.geometry("300x150")
        dialog.transient(self.root)
        dialog.grab_set()
        
        msg = f"Relatório individual gerado com sucesso para:\n{nome_cliente}\nData: {data_formatada}" if nome_cliente else "Relatórios em lote gerados com sucesso!"
        
        ttk.Label(dialog, text=msg, font=('Helvetica', 10, 'bold')).pack(pady=10)
        
        def continuar():
            dialog.destroy()
            
        def voltar_menu():
            dialog.destroy()
            self.root.destroy()
            if self.menu_principal:
                self.menu_principal.deiconify()
                self.menu_principal.lift()
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill='x', pady=10)
        
        ttk.Button(btn_frame, text="Gerar Outro Relatório", 
                  command=continuar).pack(pady=5, padx=10, fill='x')
        ttk.Button(btn_frame, text="Voltar ao Menu Principal", 
                  command=voltar_menu).pack(pady=5, padx=10, fill='x')          


        

class RelatorioConfig:
    """Classe para gerenciar configurações e estilos do relatório"""
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.setup_custom_styles()
        
    def setup_custom_styles(self):
        """Configura os estilos personalizados para o relatório"""
        self.style_heading = ParagraphStyle(
            'HeadingStyle',
            parent=self.styles['Heading1'],
            fontSize=12,
            leading=14,
            alignment=TA_LEFT,
            leftIndent=0,
            textColor=colors.black,
            spaceBefore=20,
            spaceAfter=12
        )
        
        self.style_normal = ParagraphStyle(
            'NormalStyle',
            parent=self.styles['Normal'],
            fontSize=10,
            leading=12,
            textColor=colors.black,
            spaceBefore=12,
            spaceAfter=6
        )
        
        self.style_despesa = ParagraphStyle(
            name='TipoDespesa',
            parent=self.styles['Normal'],
            fontSize=12,
            leading=14,
            alignment=TA_LEFT,
            leftIndent=0,
            firstLineIndent=0,
            rightIndent=0,
            spaceBefore=12,
            spaceAfter=6,
            keepWithNext=True
        )

def resource_path(relative_path):
    """Obtém o caminho absoluto para recursos empacotados"""
    try:
        # PyInstaller cria um temp folder e armazena o caminho em _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)



class RelatorioHandler:
    def __init__(self):
        self.config = RelatorioConfig()
        self.tipos_despesas = {
            1: "1) DESPESAS COM COLABORADORES",
            2: "2) TRANSF. PROGR. - MATERIAIS, LOCAÇÕES E PREST.SERVIÇOS",
            3: "3) BOLETOS - MATERIAIS, PREST. SERVIÇOS, IMPOSTOS, ETC.",
            4: "4) RESSARCIMENTOS E RESTITUIÇÕES",
            5: "5) DESPESAS PAGAS PELO CLIENTE",
            6: "6) PAGAMENTOS CAIXA DE OBRA",
            7: "7) ADMINISTRAÇÃO DA OBRA"
        }

        # Verificar se a logomarca existe na mesma pasta do script
        script_dir = os.path.dirname(os.path.abspath(__file__))
        self.logo_path = os.path.join(script_dir, "logo1.png")
        if not os.path.exists(self.logo_path):
            self.logo_path = None
            print("Aviso: Logomarca não encontrada na pasta do script.")
        
        self.tipos_despesas_futuras = {
            "Próximos 30 dias": lambda x: x <= self.data_ref + pd.Timedelta(days=30),
            "31 a 60 dias": lambda x: (x > self.data_ref + pd.Timedelta(days=30)) & 
                                     (x <= self.data_ref + pd.Timedelta(days=60)),
            "Após 60 dias": lambda x: x > self.data_ref + pd.Timedelta(days=60)
        }
        self.data_ref = None


       
        
    def selecionar_arquivo(self):
        """Interface para seleção do arquivo Excel"""
        root = Tk()
        root.withdraw()
        arquivo = filedialog.askopenfilename(
            title="Selecione o arquivo Excel",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
        )
        return arquivo

    def parse_data_excel(self, valor_celula):
        """Função auxiliar para converter diferentes formatos de data do Excel"""
        print(f"\nTentando converter valor: {valor_celula}")
        print(f"Tipo do valor: {type(valor_celula)}")

        if valor_celula is None:
            print("Valor é None")
            return None

        # Se já for datetime, converter para date
        if isinstance(valor_celula, datetime):
            data = valor_celula.date()
            print(f"Valor é datetime, convertido para date: {data}")
            return data

        # Se for string, tentar diferentes formatos
        if isinstance(valor_celula, str):
            valor_celula = valor_celula.strip()
            formatos = [
                '%Y-%m-%d',     # 2024-12-31
                '%d/%m/%Y',     # 31/12/2024
                '%d-%m-%Y',     # 31-12-2024
                '%d.%m.%Y',     # 31.12.2024
                '%Y/%m/%d'      # 2024/12/31
            ]
            
            for formato in formatos:
                try:
                    data = datetime.strptime(valor_celula, formato).date()
                    print(f"Valor string convertido usando formato {formato}: {data}")
                    return data
                except ValueError:
                    continue
            print("Não foi possível converter a string para data")

        # Se for número (Excel armazena datas como números), tentar converter
        if isinstance(valor_celula, (int, float)):
            try:
                # Excel usa sistema onde 1 = 1/1/1900
                timestamp = pd.Timestamp.fromordinal(int(valor_celula) + 693594)
                data = timestamp.date()
                print(f"Valor numérico convertido para data: {data}")
                return data
            except (ValueError, OverflowError) as e:
                print(f"Erro ao converter número para data: {e}")
                pass

        print("Não foi possível converter o valor para data")
        return None

    def obter_numero_relatorio(self, ws_resumo, data_relatorio):
        """Obtém o número do relatório com base na data"""
        try:
            from datetime import date
            
            # Converter a data de referência para date
            data_ref = pd.to_datetime(data_relatorio).date()
##            print(f"\nBuscando relatório para data: {data_ref}")

            # Obter a data inicial da célula L3
            data_inicial_cell = ws_resumo['L3'].value
            if not data_inicial_cell or not isinstance(data_inicial_cell, datetime):
                print("Data inicial não encontrada na célula L3 ou formato inválido")
                return 1
                
            data_inicial = data_inicial_cell.date()
##            print(f"Data inicial do arquivo: {data_inicial}")
            
            # Se a data de referência é anterior à data inicial
            if data_ref < data_inicial:
                print(f"Data {data_ref} é anterior ao início dos relatórios")
                return 1

            # Calcular número do relatório
            data_atual = data_inicial
            numero = 1
            
            while data_atual <= data_ref:
                if data_atual == data_ref:
##                    print(f"Encontrado relatório número {numero} para data {data_ref}")
                    return numero
                    
                # Incrementar para próxima data
                if data_atual.day == 5:
                    data_atual = data_atual.replace(day=20)
                else:  # day == 20
                    if data_atual.month == 12:
                        data_atual = data_atual.replace(year=data_atual.year + 1, month=1, day=5)
                    else:
                        data_atual = data_atual.replace(month=data_atual.month + 1, day=5)
                numero += 1

            print(f"Data {data_ref} não é uma data válida de relatório")
            return 1

        except Exception as e:
            print(f"Erro ao obter número do relatório: {str(e)}")
            return 1

    def obter_acumulado(self, ws_resumo, data_relatorio):
        """Obtém o valor acumulado do relatório anterior"""
        try:
            numero_relatorio = self.obter_numero_relatorio(ws_resumo, data_relatorio)
##            print(f"\nBuscando acumulado para relatório número: {numero_relatorio}")

            if numero_relatorio == 1:
                print("Primeiro relatório, retornando acumulado zero")
                return 0.0
            
            mapeamento = []
            for row in range(9, 81):  # Até linha 80 para evitar o TOTAL
                cell = ws_resumo.cell(row=row, column=1)
                if cell.value:
                    data_convertida = None
                    valor = cell.value
                    
                    # Se já for datetime
                    if isinstance(valor, datetime):
                        data_convertida = valor.date()
                    elif isinstance(valor, str):
                        try:
                            # Tentar diferentes formatos
                            formatos = ['%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d']
                            for formato in formatos:
                                try:
                                    data_convertida = datetime.strptime(valor, formato).date()
                                    break
                                except ValueError:
                                    continue
                        except Exception:
                            pass
                    elif isinstance(valor, (int, float)):
                        try:
                            data_convertida = pd.Timestamp.fromordinal(
                                int(valor) + 693594
                            ).date()
                        except Exception:
                            pass
                    
                    if data_convertida:
                        numero = ws_resumo.cell(row=row, column=2).value
                        valor_acumulado = ws_resumo.cell(row=row, column=12).value
                        
                        if numero is not None and valor_acumulado is not None:
                            try:
                                mapeamento.append({
                                    'data': data_convertida,
                                    'numero': int(numero),
                                    'valor': float(valor_acumulado)
                                })
##                                print(f"Dados encontrados: Data={data_convertida}, Número={numero}, Valor={valor_acumulado}")
                            except (ValueError, TypeError) as e:
                                print(f"Erro ao converter valores na linha {row}: {e}")

            # Ordenar por data
            mapeamento.sort(key=lambda x: x['data'])

            # Buscar o relatório anterior
            for i, item in enumerate(mapeamento):
                if item['numero'] == numero_relatorio and i > 0:
                    valor_anterior = mapeamento[i-1]['valor']
##                    print(f"Valor acumulado encontrado: {valor_anterior}")
                    return valor_anterior

            print("Nenhum valor acumulado encontrado")
            return 0.0

        except Exception as e:
            print(f"Erro ao obter acumulado: {str(e)}")
            return 0.0

    def carregar_dados_excel(self, arquivo_excel):
        """Carrega e processa os dados do arquivo Excel"""
        try:
            df = pd.read_excel(arquivo_excel, sheet_name='Dados')
            df = df.fillna("")
            
            # Verificar colunas necessárias
            colunas_necessarias = {'DATA_REL', 'TP_DESP', 'REFERÊNCIA', 'DT_VENCTO', 'VALOR'}
            if not colunas_necessarias.issubset(df.columns):
                raise ValueError(f"Colunas necessárias ausentes: {colunas_necessarias - set(df.columns)}")
                
            return df
            
        except Exception as e:
            raise Exception(f"Erro ao carregar arquivo Excel: {str(e)}")
            
    def processar_dados(self, df, data_relatorio):
        """Processa os dados conforme os critérios especificados"""
        # Converter data para datetime usando formato explícito
        try:
            data_rel = pd.to_datetime(data_relatorio)
        except:
            # Se falhar, tenta converter assumindo formato brasileiro
            data_rel = pd.to_datetime(data_relatorio, format='%d/%m/%Y')
        
        # Criar cópia do DataFrame para não modificar o original
        df = df.copy()
        
        # Formatar as datas no DataFrame usando formato brasileiro
        if 'DT_VENCTO' in df.columns:
            df['DT_VENCTO'] = pd.to_datetime(df['DT_VENCTO'], dayfirst=True)  # Forçar interpretação dia/mês
            df['DT_VENCTO'] = df['DT_VENCTO'].dt.strftime('%d/%m/%Y')
        
        # Filtrar dados
        df_filtrado = df[
            (df['DATA_REL'] == data_rel) & 
            (df['TP_DESP'] != 1)
        ].sort_values(
            by=['TP_DESP', 'DT_VENCTO', 'VALOR'], 
            ascending=[True, True, False]  # True para ordenar vencimento do mais antigo
        )
        
        df_diaria = df[
            (df['DATA_REL'] == data_rel) & 
            (df['TP_DESP'] == 1) & 
            (df['REFERÊNCIA'] == 'DIÁRIA')
        ].sort_values(
            by=['TP_DESP', 'DT_VENCTO', 'VALOR'], 
            ascending=[True, False, False]
        )
        
        df_tp_desp_1 = df[
            (df['DATA_REL'] == data_rel) & 
            (df['TP_DESP'] == 1) & 
            (df['REFERÊNCIA'] != "DIÁRIA")
        ]
        
        return df_filtrado, df_diaria, df_tp_desp_1

    def processar_lancamentos_futuros(self, df, data_relatorio):
        """Processa os lançamentos futuros do DataFrame usando DATA_REL"""
        # Converter a data do relatório para datetime usando formato explícito
        try:
            self.data_ref = pd.to_datetime(data_relatorio)
        except:
            # Se falhar, tenta converter assumindo formato brasileiro
            self.data_ref = pd.to_datetime(data_relatorio, format='%d/%m/%Y')

        # Converter a coluna DATA_REL para datetime
        df = df.copy()
        df['DATA_REL'] = pd.to_datetime(df['DATA_REL'])
        df['DT_VENCTO'] = pd.to_datetime(df['DT_VENCTO'], format='%d/%m/%Y', errors='coerce')
        
        # Formatar a data de vencimento para DD/MM/AAAA
        df['DT_VENCTO'] = df['DT_VENCTO'].dt.strftime('%d/%m/%Y')

        # Filtrar apenas lançamentos futuros baseado em DATA_REL
        df_futuro = df[(df['DATA_REL'] > self.data_ref) & (df['TP_DESP'] != 1)].copy()

        # Ordenar por data de vencimento
        df_futuro = df_futuro.sort_values('DT_VENCTO')

        # Agrupar por período baseado na DATA_REL
        df_futuro['periodo'] = df_futuro['DATA_REL'].apply(
            lambda x: next(
                (nome for nome, func in self.tipos_despesas_futuras.items() 
                 if func(x)),
                "Após 60 dias"
            )
        )

        return df_futuro
    
    def adicionar_lancamentos_futuros(self, elementos, dados):
        """Adiciona a seção de lançamentos futuros ao relatório"""
        if not dados['df_futuro'].empty:
            elementos.append(PageBreak())
            elementos.append(Paragraph("LANÇAMENTOS FUTUROS", self.config.style_heading))
            
            total_geral_futuro = 0
            
            # Agrupar por período e tipo de despesa
            for periodo in ["Próximos 30 dias", "31 a 60 dias", "Após 60 dias"]:
                df_periodo = dados['df_futuro'][dados['df_futuro']['periodo'] == periodo]
                
                if not df_periodo.empty:
                    # Adicionar título do período com estilo destacado
                    elementos.append(Paragraph(
                        f"\n{periodo}",
                        ParagraphStyle(
                            'PeriodoStyle',
                            parent=self.config.style_heading,
                            fontSize=14,
                            leading=16,
                            spaceBefore=12,
                            spaceAfter=6,
                            textColor=colors.HexColor('#2F4F4F')  # Cor mais escura para destaque
                        )
                    ))
                    
                    total_periodo = 0
                    
                    # Agrupar por tipo de despesa dentro do período
                    for tipo in sorted(df_periodo['TP_DESP'].unique()):
                        df_tipo = df_periodo[df_periodo['TP_DESP'] == tipo]
                        if not df_tipo.empty:
                            elementos.append(Paragraph(
                                self.tipos_despesas.get(tipo, f"Tipo {tipo}"),
                                self.config.style_normal
                            ))
                            
                            # Renomear colunas para corresponder ao formato esperado
                            df_tipo = df_tipo.rename(columns={
                                'DT_VENCTO': 'VENCIMENTO',
                                'DADOS_BANCARIOS': 'DADOS BANCÁRIOS'
                            })
                            
                            tabela = self.criar_tabela_despesas(
                                df_tipo,
                                ['NOME', 'VENCIMENTO', 'REFERÊNCIA', 'VALOR', 'DADOS BANCÁRIOS'],
                                [240, 70, 220, 80, 170]
                            )
                            elementos.append(tabela)
                            elementos.append(Spacer(1, 12))
                            
                            total_periodo += df_tipo['VALOR'].sum()
                    
                    # Adicionar subtotal do período
                    elementos.append(Paragraph(
                        f"Subtotal {periodo}: {self.formatar_numero(total_periodo)}",
                        ParagraphStyle(
                            'SubtotalStyle',
                            parent=self.config.style_normal,
                            fontSize=10,
                            leading=12,
                            spaceBefore=6,
                            spaceAfter=12,
                            textColor=colors.HexColor('#4A4A4A')
                        )
                    ))
                    
                    total_geral_futuro += total_periodo
            
            # Adicionar total geral dos lançamentos futuros
            elementos.append(Paragraph(
                f"\nTotal Geral de Lançamentos Futuros: {self.formatar_numero(total_geral_futuro)}",
                self.config.style_heading
            ))
    
    def formatar_numero(self, valor):
        """Formata valor numérico, tratando possíveis strings e NaN"""
        if pd.isna(valor) or valor == "":
            return "0,00"
        try:
            if isinstance(valor, str):
                valor = float(valor.replace('.', '').replace(',', '.'))
            return f"{float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        except:
            return "0,00"  # Retorna zero formatado em caso de erro

    def formatar_data(self, data):
        """Formata data para o padrão brasileiro"""
        if pd.isna(data):
            return ''
        try:
            return pd.to_datetime(data).strftime('%d/%m/%Y')
        except:
            return str(data)

    def consolidar_despesas_colaboradores(self, df):
        """Consolida as despesas dos colaboradores"""
        # Criar cópia e preencher NaN
        df = df.copy()
        df = df.infer_objects()  # Adicionar essa linha
        df = df.fillna("")
    
        agregacoes = {
            'SALÁRIO/FÉRIAS': ['SALÁRIO', 'FÉRIAS'],
            'RESCISÃO/13º SALÁRIO': ['RESCISÃO', '13º SALÁRIO'],
            'TRANSPORTE/CAFÉ': ['TRANSPORTE', 'CAFÉ']
        }
    
        if 'DADOS_BANCARIOS' in df.columns:
            df = df.rename(columns={'DADOS_BANCARIOS': 'DADOS BANCÁRIOS'})
    
        resultados = []
        for nome, grupo in df.groupby('NOME'):
            linha = {'NOME': nome}
        
            for coluna, referencias in agregacoes.items():
                valor = grupo[grupo['REFERÊNCIA'].isin(referencias)]['VALOR'].sum()
                linha[coluna] = valor if not pd.isna(valor) else 0
                
            # Pegar DIAS do lançamento de TRANSPORTE
            transporte_row = grupo[grupo['REFERÊNCIA'] == 'TRANSPORTE']
            dias = transporte_row['DIAS'].iloc[0] if not transporte_row.empty else 0
            linha['DIAS'] = int(dias) if pd.notnull(dias) else 0
            
            linha['DADOS BANCÁRIOS'] = grupo['DADOS BANCÁRIOS'].iloc[0] if not grupo['DADOS BANCÁRIOS'].empty else ''
            linha['TOTAL'] = sum(linha.get(col, 0) for col in ['SALÁRIO/FÉRIAS', 'RESCISÃO/13º SALÁRIO', 'TRANSPORTE/CAFÉ'])
        
            resultados.append(linha)
    
        df_result = pd.DataFrame(resultados)
        df_result = df_result.fillna("")  # Garantir que não há NaN no resultado
    
        colunas_ordem = ['NOME', 'SALÁRIO/FÉRIAS', 'RESCISÃO/13º SALÁRIO', 'DIAS', 
                     'TRANSPORTE/CAFÉ', 'TOTAL', 'DADOS BANCÁRIOS']
        df_result = df_result.reindex(columns=colunas_ordem)
    
        return df_result

    def criar_tabela_despesas(self, dados, colunas, larguras, incluir_total=True):
        """Cria uma tabela formatada para o relatório"""
        dados_formatados = dados.copy()
        dados_formatados = dados_formatados.fillna("")
        dados_formatados = dados_formatados.infer_objects()

        # Estilo para o cabeçalho com quebra de linha
        estilo_cabecalho = ParagraphStyle(
            'CabecalhoTabela',
            parent=self.config.style_normal,
            fontSize=8,
            leading=10,
            alignment=1,
            textColor=colors.whitesmoke
        )

        # Estilo para células com quebra de texto
        estilo_celula = ParagraphStyle(
            'CelulaTabela',
            parent=self.config.style_normal,
            fontSize=8,
            leading=10,
            alignment=0  # Alinhamento à esquerda
        )

        # Converter cabeçalhos simples em Paragraphs com quebras de linha
        cabecalhos_formatados = []
        for coluna in colunas:
            if '/' in coluna:
                texto_formatado = Paragraph(coluna.replace('/', '<br/>'), estilo_cabecalho)
            elif ' - ' in coluna:
                texto_formatado = Paragraph(coluna.replace(' - ', '<br/>'), estilo_cabecalho)
            else:
                texto_formatado = Paragraph(coluna, estilo_cabecalho)
            cabecalhos_formatados.append(texto_formatado)

        colunas_numericas = ['VALOR', 'TOTAL', 'SALÁRIO/FÉRIAS', 'RESCISÃO/13º SALÁRIO', 
                            'TRANSPORTE/CAFÉ', 'DIÁRIA', 'DIAS']

        # Processar dados linha por linha
        dados_tabela = [cabecalhos_formatados]
        for _, linha in dados_formatados.iterrows():
            linha_formatada = []
            for i, coluna in enumerate(colunas):
                valor = linha[coluna]
                
                # Formatar números
                if coluna in colunas_numericas:
                    valor = pd.to_numeric(valor, errors='coerce')
                    valor = 0 if pd.isna(valor) else valor
                    if coluna == 'DIAS':
                        valor = str(int(valor))  # Converter para inteiro e depois string
                    else:
                        valor = self.formatar_numero(valor)
                    linha_formatada.append(valor)
                
                # Formatar datas
                elif coluna in ['DT_VENCTO', 'VENCIMENTO']:
                    try:
                        valor = pd.to_datetime(valor, dayfirst=True).strftime('%d/%m/%Y')
                    except:
                        valor = str(valor)
                    linha_formatada.append(valor)
                
                # Adicionar quebra de texto para a coluna Referência
                elif coluna == 'REFERÊNCIA':
                    valor = str(valor)
                    linha_formatada.append(Paragraph(valor, estilo_celula))
                
                # Outras colunas
                else:
                    linha_formatada.append(str(valor))
                    
            dados_tabela.append(linha_formatada)

        # Adicionar linha de total se necessário
        if incluir_total:
            coluna_valor = next((i for i, col in enumerate(colunas) 
                           if col in ['VALOR', 'TOTAL']), -1)
            if coluna_valor >= 0:
                total = dados[colunas[coluna_valor]].sum()
                linha_total = [''] * len(colunas)
                linha_total[coluna_valor-1] = 'Subtotal'
                linha_total[coluna_valor] = self.formatar_numero(total)
                dados_tabela.append(linha_total)

        # Criar tabela com os dados formatados
        tabela = Table(dados_tabela, colWidths=larguras, repeatRows=1)
        
        # Definir estilos da tabela
        estilo_tabela = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]

        # Alinhar colunas numéricas à direita
        for i, col in enumerate(colunas):
            if col in colunas_numericas:
                estilo_tabela.append(('ALIGN', (i, 1), (i, -1), 'RIGHT'))

        if incluir_total:
            estilo_tabela.extend([
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ])

        tabela.setStyle(TableStyle(estilo_tabela))
        return tabela

    def criar_resumo_despesas(self, dados):
        """Cria o resumo das despesas para o relatório"""
        subtotais = {}
    
        # Calcular subtotais por tipo de despesa
        for tipo, descricao in self.tipos_despesas.items():
            valor = 0
            if tipo == 1:
                # Somar despesas de colaboradores (incluindo diárias)
                valor = (dados['df_tp_desp_1']['VALOR'].sum() +
                    dados['df_diaria']['VALOR'].sum())
            else:
                # Somar outras despesas
                df_tipo = dados['df_filtrado'][dados['df_filtrado']['TP_DESP'] == tipo]
                valor = df_tipo['VALOR'].sum()
                
            subtotais[tipo] = valor
    
        # Calcular despesas agrupadas
        despesas_a_pagar = sum(subtotais.get(tp, 0) for tp in [1, 2, 3, 4, 6, 7])
        despesas_pagas_cliente = sum(subtotais.get(tp, 0) for tp in [ 5])
    
        total_quinzena = sum(subtotais.values())
        total_obra = total_quinzena + dados.get('acumulado', 0)
    
        # Criar tabelas de resumo com formatação consistente
        tabela_subtotais = []
        for tipo, descricao in self.tipos_despesas.items():
            if tipo in subtotais:
                valor_formatado = self.formatar_numero(subtotais[tipo])
                tabela_subtotais.append([descricao, valor_formatado])
    
        tabela_totais = [
            ['DESPESAS A PAGAR', self.formatar_numero(despesas_a_pagar)],
            ['DESPESAS PAGAS PELO CLIENTE', self.formatar_numero(despesas_pagas_cliente)],
            ['COMPLEMENTO DE CAIXA', self.formatar_numero(0)],
            [''],
            ['TOTAL DA QUINZENA', self.formatar_numero(total_quinzena)],
            [f'TOTAL ACUMULADO RELATÓRIO Nº {dados.get("numero_relatorio", 0) - 1}',
             self.formatar_numero(dados.get('acumulado', 0))],
            ['TOTAL DA OBRA', self.formatar_numero(total_obra)]
        ]
    
        return tabela_subtotais, tabela_totais

    def adicionar_cabecalho(self, elementos, dados):
##        print("\nIniciando adicionar_cabecalho")
##        print(f"Tipo de elementos: {type(elementos)}")
##        print(f"Tipo de dados: {type(dados)}")
        
        try:
            if not isinstance(elementos, list):
                print("ERRO: elementos não é uma lista!")
                elementos = []
                
            # Criar estilo customizado com espaçamento de 0
            style_cabecalho = ParagraphStyle(
                'CabecalhoStyle',
                parent=self.config.style_normal,
                alignment=2,
                spaceBefore=0,
                spaceAfter=0,
                leading=12
            )

            try:
##                print(f"Antes de verificar logo - self.logo_path: {self.logo_path}")
##                print(f"Caminho da logo existe? {os.path.exists(self.logo_path)}")
                
                if self.logo_path and os.path.exists(self.logo_path):
##                    print("Tentando criar Image")
                    logo = Image(self.logo_path, width=200, height=100)
##                    print("Image criada com sucesso")
                    
                    info_empresa = [
                        Paragraph("Rua Zodiaco, 87 Sala 07 – Santa Lúcia - Belo Horizonte - MG", style_cabecalho),
                        Paragraph("(31) 3654-6616 / (31) 99974-1241 / (31) 98711-1139", style_cabecalho),
                        Paragraph("rvr.engenharia@gmail.com", style_cabecalho)
                    ]
                    
##                    print("Criando tabela do cabeçalho")
                    cabecalho_table = Table(
                        [[logo, info_empresa]], 
                        colWidths=[80, 650],
                        rowHeights=[60]
                    )
                    
                    cabecalho_table.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                        ('VALIGN', (0, 0), (1, 0), 'TOP'),
                        ('RIGHTPADDING', (1, 0), (1, 0), 0),
                    ]))
                    
##                    print("Adicionando tabela aos elementos")
                    elementos.append(cabecalho_table)
##                    print("Tabela adicionada com sucesso")
                    
            except Exception as e:
                print(f"Erro ao processar logo: {str(e)}")

                
        except Exception as e:
            print(f"Aviso: Não foi possível adicionar a logo ao cabeçalho: {e}")
            # Continua sem a logo, apenas com as informações
            info_empresa = [
                Paragraph("Rua Zodiaco, 87 Sala 07 – Santa Lúcia - Belo Horizonte - MG", style_cabecalho),
                Paragraph("(31) 3654-6616 / (31) 99974-1241 / (31) 98711-1139", style_cabecalho),
                Paragraph("rvr.engenharia@gmail.com", style_cabecalho)
            ]
            elementos.extend(info_empresa)

        # Espaço após o cabeçalho
        elementos.append(Spacer(1, 40))
        
        # Criar tabela para nome/endereço do cliente e número/data do relatório
        data_formatada = pd.to_datetime(dados.get('data_relatorio')).strftime('%d/%m/%Y')
        info_cliente = [
            [Paragraph(dados.get('nome_cliente', ''), self.config.style_heading),
             Paragraph(f"Relatório nº: {dados.get('numero_relatorio', '')}", self.config.style_normal)],
            [Paragraph(dados.get('endereco_cliente', ''), self.config.style_normal),
             Paragraph(f"Data: {data_formatada}", self.config.style_normal)]
        ]

        cliente_table = Table(
            info_cliente,
            colWidths=[680, 100],  # Ajuste as larguras conforme necessário
            rowHeights=[20, 20]   
        )
        cliente_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),    # Alinhar informações do cliente à esquerda
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),   # Alinhar número e data à direita
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elementos.append(cliente_table)

    
    def adicionar_detalhes(self, elementos, dados):
        """Adiciona os detalhes das despesas ao relatório"""
        elementos.append(Paragraph("DETALHES DAS DESPESAS", self.config.style_heading))
    
        # 1. Despesas com Colaboradores - Funcionários
        if not dados['df_tp_desp_1'].empty:
            elementos.append(Paragraph("1) DESPESAS COM COLABORADORES - FUNCIONÁRIOS", 
                                self.config.style_despesa))
            df_consolidado = self.consolidar_despesas_colaboradores(dados['df_tp_desp_1'])
            tabela = self.criar_tabela_despesas(
                df_consolidado,
                ['NOME', 'SALÁRIO/FÉRIAS', 'RESCISÃO/13º SALÁRIO', 'DIAS', 
                 'TRANSPORTE/CAFÉ', 'TOTAL', 'DADOS BANCÁRIOS'],
                [210, 70, 70, 40, 70, 80, 230]
            )
            elementos.append(tabela)
            elementos.append(Spacer(1, 12))
    
        # 2. Despesas com Colaboradores - Diaristas
        if not dados['df_diaria'].empty:
            elementos.append(Paragraph("1) DESPESAS COM COLABORADORES - DIARISTAS", 
                                self.config.style_despesa))
            # Renomear colunas para corresponder ao formato esperado
            df_diaria_formatado = dados['df_diaria'].copy()
            df_diaria_formatado = df_diaria_formatado.rename(columns={
                'VR_UNIT': 'DIÁRIA',
                'VALOR': 'TOTAL',
                'DADOS_BANCARIOS': 'DADOS BANCÁRIOS'
            })
            tabela = self.criar_tabela_despesas(
                df_diaria_formatado,
                ['NOME', 'DIÁRIA', 'DIAS', 'TOTAL', 'DADOS BANCÁRIOS'],
                [254, 80, 50, 90, 300]
            )
            elementos.append(tabela)
            elementos.append(Spacer(1, 12))
    
        # 3. Outras despesas
        for tipo in range(2, 8):
            df_tipo = dados['df_filtrado'][dados['df_filtrado']['TP_DESP'] == tipo]
            if not df_tipo.empty:
                elementos.append(Paragraph(self.tipos_despesas[tipo], 
                                    self.config.style_despesa))
                # Renomear colunas para corresponder ao formato esperado
                df_tipo = df_tipo.rename(columns={
                    'DT_VENCTO': 'VENCIMENTO',
                    'DADOS_BANCARIOS': 'DADOS BANCÁRIOS'
                })
                tabela = self.criar_tabela_despesas(
                    df_tipo,
                    ['NOME', 'VENCIMENTO', 'REFERÊNCIA', 'VALOR', 'DADOS BANCÁRIOS'],
                    [240, 70, 220, 80, 170]
                )
                elementos.append(tabela)
                elementos.append(Spacer(1, 16))

            
    def gerar_relatorio_pdf(self, dados, caminho_output, arquivo_excel):
        """Gera o relatório PDF final"""
        try:
##            print("\nIniciando geração do relatório PDF")
##            print("Dados recebidos:", dados.keys())
##            print("Caminho da logo no handler:", self.logo_path)
##            print("Caminho do output:", caminho_output)
##            print("Arquivo Excel:", arquivo_excel)
            
            
            workbook = load_workbook(arquivo_excel, data_only=True)  # <-- AQUI PODE ESTAR O PROBLEMA!
##            print("Workbook carregado")
            
            ws_resumo = workbook['RESUMO']
            
            
            data_rel = pd.to_datetime(dados['data_relatorio'])
            

            # Get report number directly from worksheet
            relatorio_num = None
            acumulado = 0.0
            
            for row in range(9, 150):  # Scan reasonable range of rows
                data_cell = ws_resumo.cell(row=row, column=1).value
                if isinstance(data_cell, datetime):
                    if data_cell.date() == data_rel.date():
                        relatorio_num = ws_resumo.cell(row=row, column=2).value
                        # Get previous report's accumulated value
                        if row > 9:
                            acumulado = ws_resumo.cell(row=row-1, column=12).value or 0.0
                        break

            # Update dados with correct values
            dados.update({
                'numero_relatorio': relatorio_num or 1,
                'acumulado': float(acumulado)
            })

            # Continue with PDF generation
            doc = SimpleDocTemplate(
                    caminho_output, 
                    pagesize=landscape(A4),
                    rightMargin=30,
                    leftMargin=30,
                    topMargin=40,
                    bottomMargin=30
            )
                
            elementos = []
            
            # Adicionar cabeçalho
            self.adicionar_cabecalho(elementos, dados)
            
            # Adicionar resumo
            elementos.append(Paragraph("RESUMO DAS DESPESAS", self.config.style_heading))
            tabela_subtotais, tabela_totais = self.criar_resumo_despesas(dados)
            
             # Criar tabelas com estilos específicos
            estilo_subtotais = TableStyle([
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),    # Texto à esquerda
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),   # Valores à direita
                ('FONTSIZE', (0, 0), (-1, -1), 9),     # Tamanho da fonte
            ])

            estilo_totais = TableStyle([
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),    # Texto à esquerda
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),   # Valores à direita
                ('FONTSIZE', (0, 0), (-1, -1), 9),     # Tamanho da fonte
                # Destacar "DESPESAS A PAGAR"
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Negrito
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),  # Fundo cinza claro
                ('BOX', (0, 0), (-1, 0), 1, colors.grey),  # Borda ao redor
                # Negrito para "TOTAL DA OBRA" (última linha)
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
            ])

            tabela_esquerda = Table(tabela_subtotais, colWidths=[300, 70])
            tabela_esquerda.setStyle(estilo_subtotais)

            tabela_direita = Table(tabela_totais, colWidths=[180, 70])
            tabela_direita.setStyle(estilo_totais)

            # Criar tabela que combina as duas anteriores
            tabela_resumo = Table(
                [[tabela_esquerda, Spacer(1, 12), tabela_direita]],
                colWidths=[400, 60, 280]
            )
        
            elementos.append(tabela_resumo)
            
            # Adicionar quebra de página
            elementos.append(PageBreak())
            
            # Adicionar detalhes
            self.adicionar_detalhes(elementos, dados)

            if dados.get('incluir_futuros', True) and dados.get('df_futuro') is not None:
                self.adicionar_lancamentos_futuros(elementos, dados)
           
            
            # Gerar PDF
            doc.build(elementos)

        except Exception as e:
            print(f"Erro na geração do relatório: {e}")
            raise       
        



        

def main():
    try:
        app = RelatorioUI(None)
        app.root.mainloop()
    except Exception as e:
        print(f"Erro durante a execução: {str(e)}")

if __name__ == "__main__":
    main()
