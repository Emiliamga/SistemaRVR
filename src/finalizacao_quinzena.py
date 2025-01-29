from Sistema_Entrada_Dados import GestaoTaxasFixas

import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
from openpyxl import load_workbook, Workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
import calendar
import os
import openpyxl

from config.utils import (
    validar_data,
    validar_data_quinzena,
    formatar_moeda,
    ARQUIVO_CLIENTES,
    ARQUIVO_CONTROLE,
    PASTA_CLIENTES,
    BASE_PATH
)

class ControleLancamentosTaxaADM:
    def __init__(self):
        self.planilha_controle = ARQUIVO_CONTROLE
        self.verificar_planilha_controle()

    def verificar_planilha_controle(self):
        """Cria a planilha de controle se não existir"""
        if not os.path.exists(self.planilha_controle):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Controle'
            
            headers = ['Cliente', 'Data Referência', 'Data Lançamento', 'Valor', 'Status']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            wb.save(self.planilha_controle)

    def verificar_lancamento_existente(self, ws_contratos, cliente, data_ref):
        """Verifica se já existe lançamento na aba de controle"""
        data_str = data_ref.strftime("%d/%m/%Y")
        for row in ws_contratos.iter_rows(min_row=3, values_only=True):
            if (row[22] and  # Tem referência na coluna PARCELAS
                row[25] == data_str):  # Mesma data
                return True
        return False

    def registrar_lancamento(self, cliente, data_ref, valor):
        """Registra um novo lançamento de taxa"""
        try:
            print(f"\nRegistrando lançamento no controle...")
            print(f"Cliente: {cliente}")
            print(f"Data: {data_ref}")
            print(f"Valor: R$ {valor:.2f}")
            
            # Se a data for string, converter para datetime
            if isinstance(data_ref, str):
                data_ref = datetime.strptime(data_ref, '%d/%m/%Y')
            
            # Verificar se já existe
            existe, detalhes = self.verificar_lancamento_existente(cliente, data_ref)
            if existe:
                print("AVISO: Lançamento já existe no controle!")
                return False
                    
            
            wb = load_workbook(self.planilha_controle)
            ws = wb['Controle']
            
            # Adicionar novo registro
            proxima_linha = ws.max_row + 1
            
            
            ws.cell(row=proxima_linha, column=1, value=cliente)  # Cliente
            ws.cell(row=proxima_linha, column=2, value=data_ref)  # Data Referência
            data_atual = datetime.now()
            ws.cell(row=proxima_linha, column=3, value=data_atual)  # Data Lançamento
            ws.cell(row=proxima_linha, column=4, value=float(valor))  # Valor
            ws.cell(row=proxima_linha, column=5, value='LANÇADO')  # Status
            
            # Formatar células
            ws.cell(row=proxima_linha, column=2).number_format = 'DD/MM/YYYY'
            ws.cell(row=proxima_linha, column=3).number_format = 'DD/MM/YYYY'
            ws.cell(row=proxima_linha, column=4).number_format = '#,##0.00'
            
            
            wb.save(self.planilha_controle)
            print("Registro concluído com sucesso!")
            return True
                
        except Exception as e:
            print(f"Erro ao registrar lançamento no controle: {str(e)}")
            if 'wb' in locals():
                wb.close()
            raise Exception(f"Erro ao registrar lançamento no controle: {str(e)}")


class FinalizacaoQuinzena:
    def __init__(self, parent=None):
        self.parent = parent  # Guarda referência ao parent
        self.root = tk.Toplevel(parent) if parent else tk.Tk()
        self.root.title("Finalização de Quinzena")
        self.root.geometry("1000x600")
        self.gestao_taxas = GestaoTaxasFixas(self)
        
        # Variável de controle para logs
        self.calculando_final = tk.BooleanVar(value=False)
        
        # Inicializar controle de taxa ADM
        self.controle = ControleLancamentosTaxaADM()
        
        # Inicializar atributos que serão usados na interface
        self.data_ref_entry = None
        self.tree_clientes = None
        
        # Configurar a interface
        self.setup_gui()
        
    def run(self):
        """Inicia a execução do sistema"""
        # Configurar o comportamento ao fechar a janela
        self.root.protocol("WM_DELETE_WINDOW", self.voltar_menu)
        
        # Centralizar a janela
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        
        # Trazer janela para frente
        self.root.lift()
        self.root.focus_force()
        
        # Iniciar loop de eventos
        self.root.mainloop()

    def setup_gui(self):
        # Frame principal com padding
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill='both', expand=True)

        # Frame para seleção de data
        frame_data = ttk.LabelFrame(main_frame, text="Data de Referência")
        frame_data.pack(fill='x', pady=5)

        ttk.Label(frame_data, text="Data:").pack(side='left', padx=5)
        self.data_ref_entry = DateEntry(
            frame_data,
            format='dd/mm/yyyy',
            locale='pt_BR'
        )
        self.data_ref_entry.pack(side='left', padx=5)
        
        # Sempre adicionar o botão de carregar clientes
        ttk.Button(frame_data, 
                  text="Carregar Clientes",
                  command=self.carregar_clientes).pack(side='left', padx=5)

        # Frame para lista de clientes
        frame_lista = ttk.LabelFrame(main_frame, text="Clientes")
        frame_lista.pack(fill='both', expand=True, pady=5)

        # Treeview com todas as colunas necessárias
        self.tree_clientes = ttk.Treeview(frame_lista, 
                                        columns=('Cliente', 'Taxa ADM', 'Valor Taxa', 'Status'),
                                        show='headings')
        
        self.tree_clientes.heading('Cliente', text='Cliente')
        self.tree_clientes.heading('Taxa ADM', text='Taxa ADM (%)')
        self.tree_clientes.heading('Valor Taxa', text='Valor Taxa')
        self.tree_clientes.heading('Status', text='Status')
        
        self.tree_clientes.column('Cliente', width=300)
        self.tree_clientes.column('Taxa ADM', width=100)
        self.tree_clientes.column('Valor Taxa', width=150)
        self.tree_clientes.column('Status', width=150)
        
        self.tree_clientes.pack(fill='both', expand=True)

        # Frame para botões
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x', pady=5)

        ttk.Button(frame_botoes, 
                  text="Processar Selecionados",
                  command=self.processar_clientes).pack(side='left', padx=5)
        
        ttk.Button(frame_botoes,
                  text="Voltar ao Menu Principal",
                  command=self.voltar_menu).pack(side='right', padx=5)


    def validar_data(self, data_str):
        """Valida o formato da data"""
        return validar_data(data_str)

    def carregar_clientes(self):
        """Carrega os clientes com suas informações e status"""
        data_ref = self.data_ref_entry.get()
        if not self.validar_data(data_ref):
            messagebox.showerror("Erro", "Data inválida!")
            return

        try:
            # Limpar lista atual
            for item in self.tree_clientes.get_children():
                self.tree_clientes.delete(item)

            wb = load_workbook(ARQUIVO_CLIENTES)
            ws = wb['Clientes']
            
            print("\nIniciando carregamento de clientes...")
            data_ref_dt = datetime.strptime(data_ref, '%d/%m/%Y')
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Pula linhas vazias
                    continue
                    
                nome_cliente = row[0]
                arquivo_cliente = PASTA_CLIENTES / f"{nome_cliente}.xlsx"
                
                print(f"\nVerificando cliente: {nome_cliente}")
                
                if not os.path.exists(arquivo_cliente):
                    print(f"Arquivo não encontrado: {arquivo_cliente}")
                    continue
                    
                try:
                    wb_cliente = load_workbook(arquivo_cliente)
                    
                    if 'Contratos_ADM' not in wb_cliente.sheetnames:
                        print(f"Aba Contratos_ADM não encontrada para: {nome_cliente}")
                        wb_cliente.close()
                        continue
                    
                    ws_contratos = wb_cliente['Contratos_ADM']
                    ws_dados = wb_cliente['Dados']

                    # Primeiro verificar se já existe lançamento na data
                    tem_lancamento = False
                    for row in ws_dados.iter_rows(min_row=2, values_only=True):
                        if row[0] and isinstance(row[0], datetime):  # Data do lançamento
                            if (row[0].day == data_ref_dt.day and 
                                row[0].month == data_ref_dt.month and 
                                row[0].year == data_ref_dt.year and 
                                row[1] == 7):  # Tipo 7 = Taxa ADM percentual
                                tem_lancamento = True
                                print(f"Lançamento existente encontrado para {data_ref}")
                                break

                    # Se já tem lançamento, pular este cliente
                    if tem_lancamento:
                        print(f"Cliente {nome_cliente} já tem lançamento para esta data")
                        wb_cliente.close()
                        continue

                    # Calcular base para o valor (soma dos tipos 1 a 6 na data)
                    valor_total = 0
                    tem_lancamentos_base = False
                    for row in ws_dados.iter_rows(min_row=2, values_only=True):
                        if (row[0] and isinstance(row[0], datetime) and  # Data
                            row[0].day == data_ref_dt.day and
                            row[0].month == data_ref_dt.month and
                            row[0].year == data_ref_dt.year):
                            
                            tipo_lanc = row[1]
                            if isinstance(tipo_lanc, (int, float)) and 1 <= tipo_lanc <= 6:
                                try:
                                    valor = float(str(row[7]).replace(',', '.'))  # Valor na coluna H
                                    valor_total += valor
                                    tem_lancamentos_base = True
                                    print(f"Valor base encontrado: R$ {valor:.2f}")
                                except (ValueError, TypeError) as e:
                                    print(f"Erro ao processar valor: {e}")

                    # Se não tem lançamentos base, pular este cliente
                    if not tem_lancamentos_base:
                        print(f"Cliente {nome_cliente} não tem lançamentos base para cálculo")
                        wb_cliente.close()
                        continue
                    
                    # Verificar contratos ativos com taxa percentual no período
                    taxa_total = 0
                    for row in ws_contratos.iter_rows(min_row=3, values_only=True):
                        if row[0]:  # Contrato principal
                            num_contrato = row[0]
                            data_inicio = row[1]
                            data_fim = row[2]
                            status = row[3]
                            
                            # Verificar se o contrato está ativo e dentro do período
                            if (status == 'ATIVO' and
                                isinstance(data_inicio, datetime) and
                                isinstance(data_fim, datetime) and
                                data_inicio <= data_ref_dt <= data_fim):
                                
                                # Buscar taxas do contrato
                                for admin_row in ws_contratos.iter_rows(min_row=3, values_only=True):
                                    if (admin_row[6] == num_contrato and  # Mesmo contrato
                                        admin_row[9] == 'Percentual'):  # Tipo percentual
                                        
                                        try:
                                            taxa = float(str(admin_row[10]).replace(',', '.'))
                                            taxa_total += taxa
                                            print(f"Taxa encontrada para contrato {num_contrato}: {taxa}%")
                                        except (ValueError, TypeError) as e:
                                            print(f"Erro ao processar taxa: {e}")
                    
                    if taxa_total > 0:
                        valor_taxa = (valor_total * taxa_total) / 100

                        # Função local para formatar valor em Real
                        def formatar_valor_br(valor):
                            return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    
                        print(f"Cliente {nome_cliente}:")
                        print(f"Base de cálculo: R$ {valor_total:.2f}")
                        print(f"Taxa total: {taxa_total}%")
                        print(f"Valor da taxa: R$ {valor_taxa:.2f}")
                        
                        # Adicionar à tree
                        self.tree_clientes.insert('', 'end', values=(
                            nome_cliente,
                            f"{taxa_total:.1f}",  # Taxa em porcentagem
                            formatar_valor_br(valor_taxa),  # Valor formatado no padrão brasileiro
                            "PENDENTE"
                        ))
                    else:
                        print(f"Cliente {nome_cliente} não tem taxa percentual ativa no período")
                    
                    wb_cliente.close()
                    
                except Exception as e:
                    print(f"Erro ao processar cliente {nome_cliente}: {str(e)}")
                    if 'wb_cliente' in locals():
                        wb_cliente.close()
            
            wb.close()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar clientes: {str(e)}")
            if 'wb' in locals():
                wb.close()


    def verificar_lancamento_existente(self, ws_contratos, cliente, data_ref):
        """Verifica se já existe lançamento para o período"""
        try:
            # Se a data_ref é string, converter para datetime
            if isinstance(data_ref, str):
                data_ref = datetime.strptime(data_ref, '%d/%m/%Y')
                
            # Carregar arquivo do cliente
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws_dados = wb['Dados']
            
            # Procurar lançamentos
            for row in ws_dados.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Data do lançamento
                    data_lancamento = row[0]
                    if isinstance(data_lancamento, datetime):
                        if (data_lancamento.day == data_ref.day and 
                            data_lancamento.month == data_ref.month and 
                            data_lancamento.year == data_ref.year):
                            
                            tipo_desp = row[1]
                            if tipo_desp == 7:  # Tipo 7 = Taxa ADM
                                wb.close()
                                return True
            
            wb.close()
            return False
            
        except Exception as e:
            print(f"Erro ao verificar lançamento existente: {str(e)}")
            if 'wb' in locals():
                wb.close()
            return False
        

    def calcular_taxa_adm(self, cliente, data_ref):
        """Calcula o valor da taxa de administração para o cliente"""
        try:
            print("\n=== Iniciando cálculo de taxa ADM ===")
            print(f"Cliente: {cliente}")
            print(f"Data Ref: {data_ref}")
            
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            print(f"Carregando arquivo: {arquivo_cliente}")
            
            wb = load_workbook(arquivo_cliente)
            ws_dados = wb["Dados"]
            ws_contratos = wb["Contratos_ADM"]
            
            print("\nBuscando contratos ativos com taxa percentual...")
            
            # Primeiro, encontrar todos os números de contratos ativos
            contratos_ativos = set()
            for row in ws_contratos.iter_rows(min_row=3, values_only=True):
                if row[0] and row[3] == 'ATIVO':  # Número do contrato e Status
                    print(f"\nContrato ativo encontrado: {row[0]}")
                    contratos_ativos.add(row[0])
            
            print(f"\nContratos ativos encontrados: {contratos_ativos}")
            
            # Agora, para cada contrato ativo, procurar administradores com taxa percentual
            taxa_adm_total = 0
            for num_contrato in contratos_ativos:
                for row in ws_contratos.iter_rows(min_row=3, values_only=True):
                    if (row[6] == num_contrato and  # Número do contrato na coluna G
                        row[9] == 'Percentual'):    # Tipo na coluna J
                        try:
                            percentual = float(str(row[10]).replace(',', '.'))
                            print(f"Contrato {num_contrato}: Percentual encontrado = {percentual}%")
                            taxa_adm_total += percentual
                        except (ValueError, TypeError) as e:
                            print(f"Erro ao processar percentual do contrato {num_contrato}: {e}")
                            continue

            print(f"\nTaxa ADM total encontrada: {taxa_adm_total:.2f}%")
            
            if taxa_adm_total == 0:
                print("Nenhuma taxa percentual ativa encontrada")
                wb.close()
                return 0
                
            taxa_decimal = taxa_adm_total / 100
            print(f"Taxa em decimal: {taxa_decimal:.4f}")
                    
            # Converter data de referência
            data_ref = datetime.strptime(data_ref, '%d/%m/%Y') if isinstance(data_ref, str) else data_ref
            print(f"\nData de referência convertida: {data_ref.strftime('%d/%m/%Y')}")
            
            # Calcular valor base
            valor_base = 0
            lancamentos_encontrados = 0
            
            print("\nBuscando lançamentos do período...")
            for row in ws_dados.iter_rows(min_row=2, values_only=True):
                data_lancamento = row[0]
                if isinstance(data_lancamento, datetime):
                    if (data_lancamento.day == data_ref.day and 
                        data_lancamento.month == data_ref.month and 
                        data_lancamento.year == data_ref.year):
                        
                        tipo_desp = row[1]
                        if isinstance(tipo_desp, (int, float)) and 1 <= tipo_desp <= 6:
                            valor = row[7]  # Coluna H com o valor
                            if valor:
                                valor_numeric = float(str(valor).replace(',', '.'))
                                valor_base += valor_numeric
                                lancamentos_encontrados += 1
                                print(f"Lançamento encontrado - Tipo: {tipo_desp}, Valor: R$ {valor_numeric:.2f}")

            print(f"\nValor base total: R$ {valor_base:.2f}")
            print(f"Total de lançamentos encontrados: {lancamentos_encontrados}")
            
            valor_taxa = valor_base * taxa_decimal
            print(f"\nValor final da taxa: R$ {valor_taxa:.2f}")
            
            wb.close()
            return valor_taxa
                
        except Exception as e:
            print(f"\nERRO: {str(e)}")
            if 'wb' in locals():
                wb.close()
            raise Exception(f"Erro ao calcular taxa: {str(e)}")

        

    def lancar_taxa_adm(self, cliente, data_ref, valor_total):
        """Lança a taxa de administração para todos os administradores"""
        try:
            print(f"Lançando taxa ADM para {cliente}")
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws_contratos = wb['Contratos_ADM']
            ws_dados = wb["Dados"]
            
            # Buscar todos os administradores com taxa percentual
            administradores = {}  # Usar dicionário para evitar duplicatas
            contratos_ativos = set()

            # Primeiro identificar contratos ativos
            for row in ws_contratos.iter_rows(min_row=3, values_only=True):
                if row[0] and row[3] == 'ATIVO':  # Número do contrato e Status
                    contratos_ativos.add(row[0])
                    print(f"Contrato ativo encontrado: {row[0]}")

            taxa_total = 0
            # Para cada contrato ativo, buscar administradores únicos com taxa percentual
            for num_contrato in contratos_ativos:
                for row in ws_contratos.iter_rows(min_row=3, values_only=True):
                    if (row[6] == num_contrato and  # Número do contrato na coluna G
                        row[9] == 'Percentual'):    # Tipo na coluna J
                        
                        cnpj_cpf = row[7]  # CNPJ/CPF na coluna H
                        
                        # Se este administrador já foi processado, pular
                        if cnpj_cpf in administradores:
                            continue
                            
                        percentual = float(str(row[10]).replace(',', '.'))
                        taxa_total += percentual
                        
                        administradores[cnpj_cpf] = {
                            'cnpj_cpf': cnpj_cpf,
                            'nome': row[8],      # Nome na coluna I
                            'percentual': percentual
                        }
                        print(f"Administrador encontrado: {row[8]} - {percentual}%")

            if not administradores:
                raise Exception("Nenhum administrador com taxa percentual encontrado")

            print(f"\nTotal de administradores: {len(administradores)}")
            print(f"Taxa total: {taxa_total}%")

            # Determinar data de vencimento conforme as regras
            data = datetime.strptime(data_ref, '%d/%m/%Y') if isinstance(data_ref, str) else data_ref
            
            # Se dia 05, vencimento no próximo dia útil
            if data.day == 5:
                dt_vencto = data  # Mesmo dia 05
                
                # Se cair em fim de semana, ajustar para próximo dia útil
                while dt_vencto.weekday() >= 5:  # 5 = Sábado, 6 = Domingo
                    dt_vencto += relativedelta(days=1)
            
            # Se dia 20, vencimento no dia 20 mesmo
            else:  # data.day == 20
                dt_vencto = data  # Mantém dia 20 independente do dia da semana
            
            print(f"Data de vencimento calculada: {dt_vencto.strftime('%d/%m/%Y')}")
            
            # Determinar quinzena para referência
            quinzena = "1ª" if data.day == 5 else "2ª"
            
            # Lançar para cada administrador
            for cnpj_cpf, adm in administradores.items():
                # Verificar se já existe lançamento
                data_str = data.strftime('%d/%m/%Y')
                lancamentos_existentes = []
                for row in ws_dados.iter_rows(min_row=2, values_only=True):
                    if (row[0] and isinstance(row[0], datetime) and
                        row[0].strftime('%d/%m/%Y') == data_str and
                        row[1] == 7 and  # Tipo 7 = Taxa ADM percentual
                        row[2] == cnpj_cpf):  # CNPJ/CPF do administrador
                        lancamentos_existentes.append(row[2])
                
                # Pular se já existe lançamento para este administrador
                if cnpj_cpf in lancamentos_existentes:
                    print(f"Já existe lançamento para {adm['nome']} nesta data")
                    continue

                # Calcular valor proporcional
                valor_adm = (valor_total * adm['percentual']) / taxa_total
                print(f"\nLançando para {adm['nome']}")
                print(f"Percentual: {adm['percentual']}%")
                print(f"Valor: R$ {valor_adm:.2f}")

                proxima_linha = ws_dados.max_row + 1
                
                # Data
                ws_dados.cell(row=proxima_linha, column=1, value=data)
                ws_dados.cell(row=proxima_linha, column=1).number_format = 'DD/MM/YYYY'
                
                # Tipo 7 para taxa ADM percentual
                ws_dados.cell(row=proxima_linha, column=2, value=7)
                
                # CNPJ/CPF e Nome
                ws_dados.cell(row=proxima_linha, column=3, value=adm['cnpj_cpf'])
                ws_dados.cell(row=proxima_linha, column=4, value=adm['nome'])
                
                # Referência
                referencia = f"ADM. OBRA REF. {quinzena} QUINZ. {data.strftime('%m/%Y')}"
                ws_dados.cell(row=proxima_linha, column=5, value=referencia)
                
                # Valor unitário e total
                ws_dados.cell(row=proxima_linha, column=6, value=valor_adm)
                ws_dados.cell(row=proxima_linha, column=6).number_format = '#,##0.00'
                
                ws_dados.cell(row=proxima_linha, column=7, value=1)  # Dias
                
                ws_dados.cell(row=proxima_linha, column=8, value=valor_adm)
                ws_dados.cell(row=proxima_linha, column=8).number_format = '#,##0.00'
                
                # Data de vencimento
                ws_dados.cell(row=proxima_linha, column=9, value=dt_vencto)
                ws_dados.cell(row=proxima_linha, column=9).number_format = 'DD/MM/YYYY'
                
                # Campos adicionais
                ws_dados.cell(row=proxima_linha, column=10, value='ADM')  # Categoria
                ws_dados.cell(row=proxima_linha, column=11, value='')     # Dados bancários
                ws_dados.cell(row=proxima_linha, column=12, value='LANÇAMENTO AUTOMÁTICO')  # Observação

            wb.save(arquivo_cliente)
            print("\nLançamentos concluídos com sucesso!")
            
        except Exception as e:
            print(f"Erro durante o lançamento: {str(e)}")
            if 'wb' in locals():
                wb.close()
            raise Exception(f"Erro ao lançar taxa: {str(e)}")


    def obter_dados_fornecedor(self, cliente):
        """Obtém CNPJ/CPF e nome do administrador do contrato"""
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws_contratos = wb['Contratos_ADM']
            
            print("\nBuscando dados do administrador...")
            
            # Primeiro encontrar contratos ativos
            contratos_ativos = []
            for row in ws_contratos.iter_rows(min_row=3, values_only=True):
                if row[0] and row[3] == 'ATIVO':  # Número do contrato e Status
                    contratos_ativos.append(row[0])
                    print(f"Contrato ativo encontrado: {row[0]}")

            # Para cada contrato ativo, buscar administrador com taxa percentual
            for num_contrato in contratos_ativos:
                for row in ws_contratos.iter_rows(min_row=3, values_only=True):
                    if (row[6] == num_contrato and  # Número do contrato na coluna G
                        row[9] == 'Percentual'):    # Tipo na coluna J
                        
                        cnpj_cpf = row[7]  # CNPJ/CPF na coluna H
                        nome = row[8]      # Nome na coluna I
                        
                        print(f"Administrador encontrado:")
                        print(f"CNPJ/CPF: {cnpj_cpf}")
                        print(f"Nome: {nome}")
                        
                        wb.close()
                        return cnpj_cpf, nome
            
            wb.close()
            raise Exception("Nenhum administrador com taxa percentual encontrado")
            
        except Exception as e:
            if 'wb' in locals():
                wb.close()
            print(f"Erro ao obter dados do fornecedor: {str(e)}")
            raise Exception(f"Erro ao obter dados do fornecedor: {str(e)}")
        

    def processar_clientes(self):
        """Processa os clientes selecionados"""
        selecionados = self.tree_clientes.selection()
        if not selecionados:
            messagebox.showwarning("Aviso", "Selecione pelo menos um cliente!")
            return

        data_ref = self.data_ref_entry.get()
        data_ref_dt = datetime.strptime(data_ref, '%d/%m/%Y')
        processados = []
        ignorados = []

        for item in selecionados:
            valores = self.tree_clientes.item(item)['values']
            cliente = valores[0]
            tipo = valores[1]
            status = valores[3]

            if status == "JÁ LANÇADO":
                ignorados.append(f"{cliente} - Já possui lançamento para o período")
                continue

            try:
                arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
                wb = load_workbook(arquivo_cliente)
                ws_contratos = wb['Contratos_ADM']
                
                if tipo == "FIXO":
                    # Processar pagamento fixo
                    lancamentos = self.gestao_taxas.processar_lancamentos_fixos(cliente, data_ref_dt)
                    if lancamentos:
                        processados.append(f"{cliente} - Taxa Fixa processada")
                    else:
                        ignorados.append(f"{cliente} - Erro ao processar taxa fixa")
                else:
                    # Processar pagamento percentual
                    valor = self.calcular_taxa_adm(cliente, data_ref)
                    if valor > 0:
                        self.lancar_taxa_adm(cliente, data_ref, valor)
                        processados.append(f"{cliente} - Taxa Percentual processada (R$ {valor:,.2f})")
                    else:
                        ignorados.append(f"{cliente} - Valor zero calculado")
                
                wb.close()

            except Exception as e:
                ignorados.append(f"{cliente} - Erro: {str(e)}")
                if 'wb' in locals():
                    wb.close()

        # Atualizar interface
        self.carregar_clientes()
        self.mostrar_resultado_processamento(processados, ignorados)

    def mostrar_resultado_processamento(self, processados, ignorados):
        """Mostra janela com resultado do processamento"""
        resultado = tk.Toplevel(self.root)
        resultado.title("Resultado do Processamento")
        resultado.geometry("600x400")
        
        # Frame para processados
        frame_proc = ttk.LabelFrame(resultado, text="Clientes Processados")
        frame_proc.pack(fill='both', expand=True, padx=5, pady=5)
        
        texto_proc = tk.Text(frame_proc, height=5)
        texto_proc.pack(fill='both', expand=True)
        for cliente in processados:
            texto_proc.insert(tk.END, f"{cliente}\n")
        texto_proc.config(state='disabled')
        
        # Frame para ignorados
        frame_ign = ttk.LabelFrame(resultado, text="Clientes Ignorados")
        frame_ign.pack(fill='both', expand=True, padx=5, pady=5)
        
        texto_ign = tk.Text(frame_ign, height=5)
        texto_ign.pack(fill='both', expand=True)
        for cliente in ignorados:
            texto_ign.insert(tk.END, f"{cliente}\n")
        texto_ign.config(state='disabled')
        
        # Botão fechar
        ttk.Button(resultado, 
                  text="Fechar", 
                  command=resultado.destroy).pack(pady=10)


    
    def voltar_menu(self):
        """Fecha a janela e retorna ao menu principal"""
        self.root.destroy()
        if self.parent:
            self.parent.deiconify()  # Mostra a janela principal
            self.parent.lift()  # Traz a janela principal para frente
        


if __name__ == "__main__":
    root = tk.Tk()
    app = FinalizacaoQuinzena(root)
    root.mainloop()
