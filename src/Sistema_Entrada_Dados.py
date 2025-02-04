import tkinter as tk
from tkinter import ttk, messagebox, filedialog, Toplevel
from tkcalendar import DateEntry, Calendar
from validate_docbr import CPF, CNPJ
from datetime import datetime
from dateutil.relativedelta import relativedelta  # Adicionar este import
import openpyxl
from openpyxl import load_workbook
import os
import babel
import re
import sys
from pathlib import Path

# Ajusta o path para incluir o diretório raiz
sys.path.append(str(Path(__file__).parent.parent))

from config.utils import (
    validar_data,
    validar_data_quinzena,
    calcular_proxima_data_quinzena,
    validar_cnpj_cpf,
    formatar_cnpj_cpf,
    formatar_moeda,
    ARQUIVO_CLIENTES,
    ARQUIVO_FORNECEDORES,
    ARQUIVO_MODELO,
    BASE_PATH,
    PASTA_CLIENTES,
    buscar_fornecedor,  # Nova importação
    selecionar_fornecedor  # Nova importação
)

class VisualizadorLancamentos:
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal  # referência ao sistema principal
        self.janela = tk.Toplevel(sistema_principal.root)  # usar .root para o Toplevel
        self.janela.title("Visualização de Lançamentos Pendentes")
        self.janela.geometry("1000x600")
        self.alteracoes = False
        self.dados_para_incluir = []
        
        # Frame principal
        self.frame_principal = ttk.Frame(self.janela)
        self.frame_principal.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Criar Treeview para visualização
        self.criar_treeview()
        
        # Frame para resumo
        self.frame_resumo = ttk.LabelFrame(self.frame_principal, text="Resumo")
        self.frame_resumo.pack(fill='x', pady=5)
        
        self.lbl_total_lancamentos = ttk.Label(self.frame_resumo, text="Total de Lançamentos: 0")
        self.lbl_total_lancamentos.pack(side='left', padx=5)
        
        self.lbl_valor_total = ttk.Label(self.frame_resumo, text="Valor Total: R$ 0,00")
        self.lbl_valor_total.pack(side='left', padx=5)
        
        # Frame para botões
        self.frame_botoes = ttk.Frame(self.frame_principal)
        self.frame_botoes.pack(fill='x', pady=5)
        
        ttk.Button(self.frame_botoes, text="Editar", command=self.editar_lancamento).pack(side='left', padx=5)
        ttk.Button(self.frame_botoes, text="Remover", command=self.remover_lancamento).pack(side='left', padx=5)
        ttk.Button(self.frame_botoes, text="Salvar na Planilha", command=self.salvar_na_planilha).pack(side='left', padx=5)
        ttk.Button(self.frame_botoes, text="Fechar", command=self.janela.destroy).pack(side='right', padx=5)

    
        # Variável para rastrear se houve alterações
        self.alteracoes = False
        

    def criar_treeview(self):
        colunas = ('Data', 'Tipo', 'CNPJ/CPF', 'Nome', 'Referência', 'NF', 'Vr. Unit.', 
                   'Dias', 'Valor', 'Vencimento', 'Categoria', 'Dados Bancários', 'Observação')
        
        self.tree = ttk.Treeview(self.frame_principal, columns=colunas, show='headings')
        
        # Configurar cabeçalhos
        for col in colunas:
            self.tree.heading(col, text=col)
            # Ajustar largura baseado no conteúdo
            if col in ['CNPJ/CPF', 'Nome', 'Referência', 'Dados Bancários', 'Observação']:
                width = 150
            elif col in ['Data', 'Vencimento']:
                width = 100
            elif col in ['Vr. Unit.', 'Valor', 'NF']:
                width = 100
            else:
                width = 80
            self.tree.column(col, width=width)

        # Adicionar scrollbars
        scrolly = ttk.Scrollbar(self.frame_principal, orient='vertical', command=self.tree.yview)
        scrollx = ttk.Scrollbar(self.frame_principal, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        # Posicionar elementos
        self.tree.pack(fill='both', expand=True)
        scrolly.pack(side='right', fill='y')
        scrollx.pack(side='bottom', fill='x')

    def atualizar_dados(self, dados):
        """Atualiza os dados na visualização"""
        self.dados_para_incluir = dados.copy()  # Fazer uma cópia dos dados
        # Limpar dados existentes
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        # Inserir novos dados
        valor_total = 0
        for lancamento in self.dados_para_incluir:
            valores = (
                lancamento['data'],
                lancamento['tp_desp'],
                lancamento['cnpj_cpf'],
                lancamento['nome'],
                lancamento['referencia'],
                lancamento.get('nf', ''),  # Usar get() para compatibilidade
                lancamento['vr_unit'],
                lancamento['dias'],
                lancamento['valor'],
                lancamento['dt_vencto'],
                lancamento['categoria'],
                lancamento['dados_bancarios'],
                lancamento['observacao']
            )
            self.tree.insert('', 'end', values=valores)
            valor_total += float(lancamento['valor'])
        
        # Atualizar resumo
        self.lbl_total_lancamentos.config(text=f"Total de Lançamentos: {len(dados)}")
        self.lbl_valor_total.config(text=f"Valor Total: R$ {valor_total:,.2f}")


    def editar_lancamento(self):
        """Abre a janela de edição para o lançamento selecionado"""
        item_selecionado = self.tree.selection()
        if not item_selecionado:
            messagebox.showwarning("Aviso", "Selecione um lançamento para editar")
            return

        # Obter índice do item selecionado
        todos_items = self.tree.get_children()
        indice = todos_items.index(item_selecionado[0])
        
        # Obter valores atuais
        valores = self.tree.item(item_selecionado)['values']
        dados = {
            'data': valores[0],
            'tp_desp': valores[1],
            'cnpj_cpf': valores[2],
            'nome': valores[3],
            'referencia': valores[4],
            'nf': valores[5],
            'vr_unit': valores[6],
            'dias': valores[7],
            'valor': valores[8],
            'dt_vencto': valores[9],
            'categoria': valores[10],
            'dados_bancarios': valores[11],
            'observacao': valores[12] if len(valores) > 12 else ''
        }
        
        # Criar editor
        editor = EditorLancamento(self.janela, dados, indice, self.atualizar_lancamento)

    def atualizar_lancamento(self, indice, novos_dados):
        """Atualiza os dados de um lançamento específico"""
        try:
            # Formatar CNPJ/CPF baseado no número de dígitos
            cnpj_cpf = str(novos_dados['cnpj_cpf']).replace('.', '').replace('-', '').replace('/', '')
            novos_dados['cnpj_cpf'] = formatar_cnpj_cpf(cnpj_cpf)

            # Converter observação para maiúsculas
            novos_dados['observacao'] = novos_dados['observacao'].upper()

            # Atualizar na treeview
            item = self.tree.get_children()[indice]
            valores = (
                novos_dados['data'],
                novos_dados['tp_desp'],
                novos_dados['cnpj_cpf'],
                novos_dados['nome'],
                novos_dados['referencia'],
                novos_dados['nf'],
                novos_dados['vr_unit'],
                novos_dados['dias'],
                novos_dados['valor'],
                novos_dados['dt_vencto'],
                novos_dados['categoria'],
                novos_dados['dados_bancarios'],
                novos_dados['observacao']
            )
            
            # Atualizar dados na lista
            self.dados_para_incluir[indice] = novos_dados.copy()
            
            # Atualizar treeview
            self.tree.item(item, values=valores)
            
            # Atualizar resumo
            self.atualizar_resumo()
            
            return True
        except Exception as e:
            print(f"Erro ao atualizar lançamento: {str(e)}")
            return False


    def remover_lancamento(self):
        item_selecionado = self.tree.selection()
        if not item_selecionado:
            messagebox.showwarning("Aviso", "Selecione um lançamento para remover")
            return
            
        if messagebox.askyesno("Confirmação", "Deseja realmente remover este lançamento?"):
            # Obter índice do item selecionado
            todos_items = self.tree.get_children()
            indice = todos_items.index(item_selecionado[0])
            
            # Remover da lista de dados
            if 0 <= indice < len(self.dados_para_incluir):
                self.dados_para_incluir.pop(indice)
            
            # Remover da visualização
            self.tree.delete(item_selecionado)
            
            # Atualizar contadores e totais
            self.atualizar_resumo()

    def salvar_na_planilha(self):
        """Salva os dados diretamente na planilha"""
        try:
            # Atualizar dados do sistema principal
            self.sistema.dados_para_incluir = self.dados_para_incluir.copy()
            
            # Chamar o método enviar_dados do sistema principal
            self.sistema.enviar_dados()
            self.janela.destroy()  # Fecha o visualizador após salvar
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar dados: {str(e)}")

        

    def atualizar_resumo(self):
        items = self.tree.get_children()
        total_lancamentos = len(items)
        valor_total = sum(float(self.tree.item(item)['values'][8]) for item in items)
        
        self.lbl_total_lancamentos.config(text=f"Total de Lançamentos: {total_lancamentos}")
        self.lbl_valor_total.config(text=f"Valor Total: R$ {valor_total:,.2f}")


    def get_dados_atualizados(self):
        """Retorna todos os dados atualizados"""
        return self.dados_para_incluir.copy()


class EditorLancamento:
    def __init__(self, parent, dados, indice, callback_atualizacao):
        self.janela = tk.Toplevel(parent)
        self.janela.title("Editar Lançamento")
        self.janela.geometry("600x500")
        
        self.dados = dados
        self.indice = indice
        self.callback_atualizacao = callback_atualizacao
        
        # Frame principal
        frame = ttk.Frame(self.janela, padding="10")
        frame.pack(fill='both', expand=True)
        
        # Dados do Fornecedor (não editáveis)
        frame_fornecedor = ttk.LabelFrame(frame, text="Dados do Fornecedor")
        frame_fornecedor.pack(fill='x', pady=5)
        
        ttk.Label(frame_fornecedor, text="CNPJ/CPF:").grid(row=0, column=0, padx=5, pady=2)
        self.cnpj_cpf = ttk.Entry(frame_fornecedor, state='readonly')
        self.cnpj_cpf.grid(row=0, column=1, padx=5, pady=2)
        
        ttk.Label(frame_fornecedor, text="Nome:").grid(row=1, column=0, padx=5, pady=2)
        self.nome = ttk.Entry(frame_fornecedor, state='readonly')
        self.nome.grid(row=1, column=1, padx=5, pady=2)
        
        # Dados da Despesa (editáveis)
        frame_despesa = ttk.LabelFrame(frame, text="Dados da Despesa")
        frame_despesa.pack(fill='x', pady=5)
        
        # Data de Referência
        ttk.Label(frame_despesa, text="Data do Relatório:").grid(row=0, column=0, padx=5, pady=2)
        self.data_rel = DateEntry(frame_despesa, format='dd/mm/yyyy', locale='pt_BR')
        self.data_rel.grid(row=0, column=1, padx=5, pady=2)
        
        # Tipo de Despesa
        ttk.Label(frame_despesa, text="Tipo Despesa:").grid(row=1, column=0, padx=5, pady=2)
        self.tp_desp = ttk.Entry(frame_despesa)
        self.tp_desp.grid(row=1, column=1, padx=5, pady=2)
        
        # Referência
        ttk.Label(frame_despesa, text="Referência:").grid(row=2, column=0, padx=5, pady=2)
        self.referencia = ttk.Entry(frame_despesa)
        self.referencia.grid(row=2, column=1, padx=5, pady=2)
        
        # NF
        ttk.Label(frame_despesa, text="NF:").grid(row=3, column=0, padx=5, pady=2)
        self.nf = ttk.Entry(frame_despesa)  # Corrigido: campo NF
        self.nf.grid(row=3, column=1, padx=5, pady=2)
        
        # Valor Unitário
        ttk.Label(frame_despesa, text="Valor Unitário:").grid(row=4, column=0, padx=5, pady=2)
        self.vr_unit = ttk.Entry(frame_despesa)  # Corrigido: campo vr_unit
        self.vr_unit.grid(row=4, column=1, padx=5, pady=2)
        
        # Dias
        ttk.Label(frame_despesa, text="Dias:").grid(row=5, column=0, padx=5, pady=2)
        self.dias = ttk.Entry(frame_despesa)
        self.dias.grid(row=5, column=1, padx=5, pady=2)
        
        # Valor Total
        ttk.Label(frame_despesa, text="Valor Total:").grid(row=6, column=0, padx=5, pady=2)
        self.valor = ttk.Entry(frame_despesa, state='readonly')
        self.valor.grid(row=6, column=1, padx=5, pady=2)
        
        # Data de Vencimento
        ttk.Label(frame_despesa, text="Data Vencimento:").grid(row=7, column=0, padx=5, pady=2)
        self.dt_vencto = DateEntry(frame_despesa, format='dd/mm/yyyy', locale='pt_BR')
        self.dt_vencto.grid(row=7, column=1, padx=5, pady=2)
        
        # Observação
        ttk.Label(frame_despesa, text="Observação:").grid(row=8, column=0, padx=5, pady=2)
        self.observacao = ttk.Entry(frame_despesa)
        self.observacao.grid(row=8, column=1, padx=5, pady=2)
        
        # Eventos
        self.vr_unit.bind('<KeyRelease>', self.calcular_valor_total)
        self.dias.bind('<KeyRelease>', self.calcular_valor_total)
        
        # Botões
        frame_botoes = ttk.Frame(frame)
        frame_botoes.pack(fill='x', pady=10)
        
        ttk.Button(frame_botoes, text="Salvar", command=self.salvar).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Cancelar", command=self.janela.destroy).pack(side='left', padx=5)
        
        # Preencher dados
        self.preencher_dados()
        
    def preencher_dados(self):
        """Preenche os campos com os dados atuais"""
        self.cnpj_cpf.config(state='normal')
        self.cnpj_cpf.insert(0, self.dados['cnpj_cpf'])
        self.cnpj_cpf.config(state='readonly')
        
        self.nome.config(state='normal')
        self.nome.insert(0, self.dados['nome'])
        self.nome.config(state='readonly')
        
        self.data_rel.set_date(datetime.strptime(self.dados['data'], '%d/%m/%Y'))
        self.tp_desp.insert(0, self.dados['tp_desp'])
        self.referencia.insert(0, self.dados['referencia'])
        self.nf.insert(0, self.dados.get('nf', ''))
        self.vr_unit.insert(0, self.dados['vr_unit'])
        self.dias.insert(0, str(self.dados['dias']))
        
        self.valor.config(state='normal')
        self.valor.insert(0, self.dados['valor'])
        self.valor.config(state='readonly')
        
        self.dt_vencto.set_date(datetime.strptime(self.dados['dt_vencto'], '%d/%m/%Y'))
        self.observacao.insert(0, self.dados.get('observacao', ''))
        
    def calcular_valor_total(self, event=None):
        """Calcula o valor total baseado no valor unitário e dias"""
        try:
            vr_unit = float(self.vr_unit.get().replace(',', '.'))
            dias = int(self.dias.get() or 1)
            valor_total = vr_unit * dias
            
            self.valor.config(state='normal')
            self.valor.delete(0, tk.END)
            self.valor.insert(0, f"{valor_total:.2f}")
            self.valor.config(state='readonly')
            
        except (ValueError, AttributeError):
            self.valor.config(state='normal')
            self.valor.delete(0, tk.END)
            self.valor.config(state='readonly')
            
    def salvar(self):
        """Salva as alterações e fecha a janela"""
        try:
            # Validar campos obrigatórios
            if not all([self.tp_desp.get(), self.referencia.get(), self.vr_unit.get()]):
                messagebox.showerror("Erro", "Preencha todos os campos obrigatórios!")
                return
            
            # Validar datas
            for data_entry in [self.data_rel, self.dt_vencto]:
                data_str = data_entry.get()
                try:
                    datetime.strptime(data_str, '%d/%m/%Y')
                except ValueError:
                    messagebox.showerror("Erro", "Data inválida!")
                    return
            
            # Atualizar dados
            dados_atualizados = {
                'data': self.data_rel.get(),
                'tp_desp': self.tp_desp.get(),
                'cnpj_cpf': self.dados['cnpj_cpf'],
                'nome': self.dados['nome'],
                'referencia': self.referencia.get(),
                'nf': self.nf.get(),
                'vr_unit': self.vr_unit.get(),
                'dias': int(self.dias.get() or 1),
                'valor': self.valor.get(),
                'dt_vencto': self.dt_vencto.get(),
                'categoria': self.dados['categoria'],
                'dados_bancarios': self.dados['dados_bancarios'],
                'observacao': self.observacao.get()
            }
            
            # Chamar callback de atualização e verificar sucesso
            if self.callback_atualizacao(self.indice, dados_atualizados):
                messagebox.showinfo("Sucesso", "Alterações salvas com sucesso!")
                self.janela.destroy()
            else:
                messagebox.showerror("Erro", "Não foi possível salvar as alterações!")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar alterações: {str(e)}")

        

class SistemaEntradaDados:
    def __init__(self, parent=None):
        print("Inicializando SistemaEntradaDados...")  # Debug
        if parent:
            self.root = tk.Toplevel(parent)
            self.menu_principal = parent  # Guardar referência à janela principal
        else:
            self.root = tk.Tk()
            self.menu_principal = None
        
        self.root.title("Sistema de Entrada de Dados")
        self.dados_para_incluir = []
        self.data_rel = None
        self.cliente_atual = None
        self.visualizador = None
        self._gestor_parcelas = None  # Inicializa como None
        self.gestao_taxas = GestaoTaxasFixas(self)

        # Frame temporário para criar os entries
        temp_frame = ttk.Frame(self.root)

        # Inicializa os dicionários com Entry widgets temporários
        self.campos_fornecedor = {
            'cnpj_cpf': tk.Entry(temp_frame),
            'nome': tk.Entry(temp_frame),
            'categoria': tk.Entry(temp_frame),
            'dados_bancarios': tk.Entry(temp_frame)
        }
    
        self.campos_despesa = {
            'tp_desp': tk.Entry(temp_frame),
            'referencia': tk.Entry(temp_frame),
            'nf': tk.Entry(temp_frame),
            'vr_unit': tk.Entry(temp_frame),
            'dias': tk.Entry(temp_frame),
            'valor': tk.Entry(temp_frame),
            'dt_vencto': tk.Entry(temp_frame),
            'observacao': tk.Entry(temp_frame)
        }
    
        self.setup_gui()

    @property
    def gestor_parcelas(self):
        """Getter para gestor_parcelas - cria apenas quando necessário"""
        if self._gestor_parcelas is None:
            print("Criando nova instância do GestorParcelas")  # Debug
            self._gestor_parcelas = GestorParcelas(self)
        return self._gestor_parcelas

    @gestor_parcelas.setter
    def gestor_parcelas(self, valor):
        """Setter para gestor_parcelas"""
        self._gestor_parcelas = valor        

    def voltar_menu(self):
        """Retorna ao menu principal verificando dados não salvos"""
        if self.dados_para_incluir and messagebox.askyesno(
            "Confirmação", 
            "Existem dados não salvos. Deseja salvá-los antes de sair?"):
            self.enviar_dados()
        
        self.root.destroy()  # Fecha a janela atual
        
        # Se tiver referência ao menu principal, mostra ele
        if self.menu_principal:
            self.menu_principal.deiconify()
            self.menu_principal.lift()
            self.menu_principal.focus_force()

    def sair_sistema(self):
        """Fecha o sistema verificando dados não salvos"""
        if self.dados_para_incluir and messagebox.askyesno(
            "Confirmação", 
            "Existem dados não salvos. Deseja salvá-los antes de sair?"):
            self.enviar_dados()
        self.root.destroy()
        sys.exit()    
    

    def setup_gui(self):
        # Frame principal com abas
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=5)

        # Criar abas
        self.aba_selecao = ttk.Frame(self.notebook)
        self.aba_fornecedor = ttk.Frame(self.notebook)
        self.aba_dados = ttk.Frame(self.notebook)

        self.notebook.add(self.aba_selecao, text='Seleção de Cliente')
        self.notebook.add(self.aba_fornecedor, text='Fornecedor')
        self.notebook.add(self.aba_dados, text='Entrada de Dados')

        self.setup_aba_selecao()
        self.setup_aba_fornecedor()
        self.setup_aba_dados()

    def setup_aba_selecao(self):
        """Configura a aba de seleção de cliente"""
        # Frame principal para organização
        frame_principal = ttk.Frame(self.aba_selecao)
        frame_principal.pack(expand=True, fill='both', padx=10, pady=5)

        # Lista de clientes
        ttk.Label(frame_principal, text="Selecione o Cliente:").pack(pady=5)
        self.cliente_combobox = ttk.Combobox(frame_principal)
        self.cliente_combobox.pack(pady=5)
        
        # Frame para botões de gerenciamento de clientes
        frame_gerenciar = ttk.Frame(frame_principal)
        frame_gerenciar.pack(pady=5)
        
        ttk.Button(frame_gerenciar, 
                  text="Novo Cliente", 
                  command=self.criar_novo_cliente).pack(side='left', padx=5)
                 
        ttk.Button(frame_gerenciar,
                  text="Editar Cliente",
                  command=self.editar_cliente).pack(side='left', padx=5)

        ttk.Button(frame_gerenciar, 
                  text="Gerir Contratos",
                  command=self.abrir_gestao_contratos).pack(side='left', padx=5)

        # Botão continuar (inicialmente desabilitado)
        self.btn_continuar = ttk.Button(frame_gerenciar,
                                      text="Continuar →",
                                      command=self.continuar_para_fornecedor,
                                      state='disabled')
        self.btn_continuar.pack(side='left', padx=5)
        
        # Carregar clientes existentes
        self.atualizar_lista_clientes()
        
        # Binding para seleção de cliente
        self.cliente_combobox.bind('<<ComboboxSelected>>', self.selecionar_cliente)

        # Frame de botões
        frame_botoes_selecao = ttk.Frame(frame_principal)
        frame_botoes_selecao.pack(fill='x', side='bottom', pady=5)

        ttk.Button(frame_botoes_selecao, 
                   text="Voltar ao Menu", 
                   command=self.voltar_menu).pack(side='left', padx=5)
        ttk.Button(frame_botoes_selecao, 
                   text="Sair", 
                   command=self.sair_sistema).pack(side='left', padx=5)

    def abrir_gestao_contratos(self):
        """Abre a gestão de contratos para o cliente atual"""
        if not self.cliente_atual:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro!")
            return
            
        gestor = GestaoContratos(self.root)
        gestor.abrir_janela_contrato(self.cliente_atual)


    def selecionar_cliente(self, event):
        """Atualiza seleção de cliente e habilita botão de continuar"""
        self.cliente_atual = self.cliente_combobox.get()
        # Atualiza label na aba de dados
        self.cliente_label.config(text=f"Cliente: {self.cliente_atual}")
        # Habilita o botão continuar
        self.btn_continuar.config(state='normal')
        # Não muda de aba automaticamente


    def continuar_para_fornecedor(self):
        """Avança para a aba de fornecedor após confirmar seleção"""
        if self.cliente_atual:
            self.notebook.select(1)  # Vai para aba de fornecedor
        else:
            messagebox.showwarning("Aviso", "Selecione um cliente primeiro!")



    def criar_arquivo_clientes(self):
        """Cria arquivo base de clientes se não existir"""
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Clientes'
            
            # Adicionar cabeçalhos - somente campos básicos agora
            headers = ['Nome', 'Endereco', 'Data_Inicial', 'Observacoes']
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
            
            caminho_base = ARQUIVO_CLIENTES
            workbook.save(caminho_base)
            messagebox.showinfo("Informação", "Arquivo de clientes criado com sucesso!")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao criar arquivo de clientes: {str(e)}")


 
    def criar_novo_cliente(self):
        """Abre janela para cadastro de novo cliente"""
        janela_cliente = tk.Toplevel(self.root)
        janela_cliente.title("Novo Cliente")
        janela_cliente.geometry("500x400")  # Reduzido pois terá menos campos

        # Campos do formulário
        tk.Label(janela_cliente, text="Nome do Cliente:*").pack(pady=5)
        nome_entry = tk.Entry(janela_cliente, width=50)
        nome_entry.pack(pady=5)

        tk.Label(janela_cliente, text="Endereço:*").pack(pady=5)
        endereco_entry = tk.Entry(janela_cliente, width=50)
        endereco_entry.pack(pady=5)

        tk.Label(janela_cliente, text="Data Inicial:* (Dia 5 ou 20)").pack(pady=5)
        data_entry = DateEntry(
            janela_cliente,
            width=20,
            date_pattern='yyyy-mm-dd',
            locale='pt_BR'
        )
        data_entry.pack(pady=5)

        def validar_data(*args):
            """Valida se a data selecionada é dia 5 ou 20"""
            data = data_entry.get_date()
            if data.day not in [5, 20]:
                messagebox.showwarning(
                    "Data Inválida",
                    "A data inicial deve ser dia 5 ou 20 do mês.\n"
                    "Por favor, selecione uma data válida."
                )
                # Encontrar o próximo dia 5 ou 20
                if data.day < 5:
                    data = data.replace(day=5)
                elif data.day < 20:
                    data = data.replace(day=20)
                else:
                    if data.month == 12:
                        data = data.replace(year=data.year + 1, month=1, day=5)
                    else:
                        data = data.replace(month=data.month + 1, day=5)
                data_entry.set_date(data)

        # Adicionar validação quando a data é alterada
        data_entry.bind("<<DateEntrySelected>>", validar_data)

        tk.Label(janela_cliente, text="Observações:").pack(pady=5)
        obs_entry = tk.Entry(janela_cliente, width=50)
        obs_entry.pack(pady=5)
        

        def salvar_cliente():
            nome = nome_entry.get().strip()
            endereco = endereco_entry.get().strip()
            data = data_entry.get()
            observacoes = obs_entry.get().strip()
            
            if not nome or not endereco:
                messagebox.showerror("Erro", "Nome e Endereço são obrigatórios!")
                return

            # Verificar se a data é válida
            try:
                data = datetime.strptime(data, '%Y-%m-%d').date()
                if data.day not in [5, 20]:
                    messagebox.showerror("Erro", "A data inicial deve ser dia 5 ou 20 do mês!")
                    return
            except ValueError:
                messagebox.showerror("Erro", "Data inválida!")
                return

            try:
                # Salvar no arquivo de clientes
                caminho_base = ARQUIVO_CLIENTES
                workbook = load_workbook(caminho_base)
                sheet = workbook['Clientes']

                # Verificar se cliente já existe
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[0].upper() == nome.upper():
                        messagebox.showerror("Erro", "Cliente já cadastrado!")
                        return

                # Adicionar novo cliente
                proxima_linha = sheet.max_row + 1
                sheet.cell(row=proxima_linha, column=1, value=nome.upper())
                sheet.cell(row=proxima_linha, column=2, value=endereco.upper())
                sheet.cell(row=proxima_linha, column=3, value=data)
                sheet.cell(row=proxima_linha, column=4, value=observacoes.upper())

                workbook.save(caminho_base)

                # Criar arquivo do cliente baseado no modelo
                if self.criar_arquivo_cliente(nome.upper(), endereco.upper()):
                    messagebox.showinfo("Sucesso", "Cliente cadastrado com sucesso!")
                    self.atualizar_lista_clientes()
                    janela_cliente.destroy()

            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao cadastrar cliente: {str(e)}")

        tk.Button(janela_cliente, text="Salvar", command=salvar_cliente).pack(pady=10)
        tk.Button(janela_cliente, text="Cancelar", 
                 command=janela_cliente.destroy).pack(pady=5)




    def criar_arquivo_clientes(self):
        """Cria arquivo base de clientes se não existir"""
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Clientes'
            
            # Adicionar cabeçalhos - somente campos básicos agora
            headers = ['Nome', 'Endereco', 'Data_Inicial', 'Observacoes']
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)
            
            caminho_base = ARQUIVO_CLIENTES
            workbook.save(caminho_base)
            messagebox.showinfo("Informação", "Arquivo de clientes criado com sucesso!")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao criar arquivo de clientes: {str(e)}")
        


    def criar_arquivo_cliente(self, nome_cliente, endereco):
        """
        Cria um novo arquivo Excel para o cliente baseado no MODELO.xlsx
        """
        try:
            modelo_path = ARQUIVO_MODELO
            novo_arquivo = PASTA_CLIENTES / f"{nome_cliente}.xlsx"
            
            if os.path.exists(novo_arquivo):
                raise Exception("Arquivo do cliente já existe!")
                
            # Buscar data inicial do arquivo clientes.xlsx
            wb_clientes = load_workbook(ARQUIVO_CLIENTES)
            ws_clientes = wb_clientes['Clientes']
            
            data_inicial = None
            # Procurar o cliente e sua data inicial
            for row in range(2, ws_clientes.max_row + 1):
                if ws_clientes.cell(row=row, column=1).value == nome_cliente:
                    data_valor = ws_clientes.cell(row=row, column=3).value  # Coluna C
                    if not data_valor:
                        raise Exception("Data inicial não informada no cadastro do cliente")
                        
                    if isinstance(data_valor, datetime):
                        data_inicial = data_valor.date()
                    else:
                        try:
                            data_inicial = datetime.strptime(str(data_valor), '%Y-%m-%d').date()
                        except ValueError:
                            raise Exception("Data inicial deve estar no formato AAAA-MM-DD")
                    break
            
            if not data_inicial:
                raise Exception("Cliente não encontrado no cadastro")
                
            # Validar se é dia 5 ou 20
            if data_inicial.day not in [5, 20]:
                raise Exception("A data inicial deve ser dia 5 ou 20 do mês")
                
            # Copiar o arquivo modelo
            from shutil import copy2
            copy2(modelo_path, novo_arquivo)
            
            # Abrir o novo arquivo para edição
            workbook = load_workbook(novo_arquivo)
            
            # Atualizar planilha RESUMO
            resumo_sheet = workbook["RESUMO"]
            
            # Informações básicas
            resumo_sheet["A3"] = nome_cliente
            resumo_sheet["A4"] = endereco
            
            # Descrições das células
            resumo_sheet["K3"] = "Data Inicial"
            
            # Adicionar data inicial
            resumo_sheet["L3"] = data_inicial
            resumo_sheet["L3"].number_format = 'dd/mm/yyyy'
            
            # Gerar as 96 datas quinzenais
            data_atual = data_inicial
            datas_geradas = []
            
            for i in range(96):  # 4 anos = 96 relatórios
                row = i + 9  # Começar na linha 9
                
                # Verificar se a data já foi usada
                if data_atual in datas_geradas:
                    raise Exception(f"Data duplicada detectada: {data_atual.strftime('%d/%m/%Y')}")
                datas_geradas.append(data_atual)
                
                # Adicionar data e número do relatório
                resumo_sheet.cell(row=row, column=1, value=data_atual)
                resumo_sheet.cell(row=row, column=1).number_format = 'dd/mm/yyyy'
                resumo_sheet.cell(row=row, column=2, value=i + 1)
                
                # Próxima data
                if data_atual.day == 5:
                    data_atual = data_atual.replace(day=20)
                else:  # day == 20
                    if data_atual.month == 12:
                        data_atual = data_atual.replace(year=data_atual.year + 1, month=1, day=5)
                    else:
                        data_atual = data_atual.replace(month=data_atual.month + 1, day=5)

            # Criar aba Contratos_ADM
            contratos_sheet = workbook.create_sheet("Contratos_ADM")
            
            # Definir os blocos na linha 1
            blocos = ["CONTRATOS", "", "", "", "", "",
                     "ADMINISTRADORES_CONTRATO", "", "", "", "", "", "",
                     "ADITIVOS", "", "", "",
                     "ADMINISTRADORES_ADITIVO", "", "", "", "", "", "",
                     "PARCELAS", "", "", "", "", "", "", ""]
            
            for col, valor in enumerate(blocos, 1):
                contratos_sheet.cell(row=1, column=col, value=valor)
            
            # Definir cabeçalhos na linha 2
            headers = [
                # CONTRATOS
                "Nº Contrato", "Data Início", "Data Fim", "Status", "Observações", "",
                # ADMINISTRADORES_CONTRATO
                "Nº Contrato", "CNPJ/CPF", "Nome/Razão Social", "Tipo", "Valor/Percentual", "Valor Total", "Nº Parcelas", 
                # ADITIVOS
                "Nº Contrato", "Nº Aditivo", "Data Início", "Data Fim",
                # ADMINISTRADORES_ADITIVO
                "Nº Contrato", "Nº Aditivo", "CNPJ/CPF", "Nome/Razão Social", "Tipo", "Valor/Percentual", "Valor Total",
                # PARCELAS
                "Referência", "Número", "CNPJ/CPF", "Nome", "Data Vencimento", "Valor", "Status", "Data Pagamento"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = contratos_sheet.cell(row=2, column=col, value=header)
                # Formatação do cabeçalho
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            # Ajustar largura das colunas
            for col in range(1, len(headers) + 1):
                contratos_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            
            # Salvar alterações
            workbook.save(novo_arquivo)
            wb_clientes.close()
            
            return True
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao criar arquivo do cliente: {str(e)}")
            if 'wb_clientes' in locals():
                wb_clientes.close()
            return False



    def editar_cliente(self):
        """Edita o cliente selecionado"""
        cliente_selecionado = self.cliente_combobox.get()
        if not cliente_selecionado:
            messagebox.showwarning("Aviso", "Selecione um cliente para editar")
            return

        try:
            # Carregar dados do cliente
            wb = load_workbook(ARQUIVO_CLIENTES)
            ws = wb['Clientes']
            
            dados_cliente = None
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == cliente_selecionado:
                    dados_cliente = {
                        'nome': row[0],
                        'endereco': row[1],
                        'data_inicial': row[2],
                        'observacoes': row[3]
                    }
                    break
            
            wb.close()
            
            if not dados_cliente:
                messagebox.showerror("Erro", "Cliente não encontrado!")
                return
                
            # Criar janela de edição
            janela_edicao = tk.Toplevel(self.root)
            janela_edicao.title(f"Editar Cliente - {cliente_selecionado}")
            janela_edicao.geometry("500x300")

            # Frame principal
            frame = ttk.Frame(janela_edicao, padding="10")
            frame.pack(fill='both', expand=True)

            # Frame para dados básicos
            frame_dados = ttk.LabelFrame(frame, text="Dados do Cliente")
            frame_dados.pack(fill='x', pady=5)

            # Nome
            ttk.Label(frame_dados, text="Nome do Cliente:*").grid(row=0, column=0, padx=5, pady=2)
            nome_entry = ttk.Entry(frame_dados, width=50)
            nome_entry.insert(0, dados_cliente['nome'])
            nome_entry.grid(row=0, column=1, padx=5, pady=2)

            # Endereço
            ttk.Label(frame_dados, text="Endereço:*").grid(row=1, column=0, padx=5, pady=2)
            endereco_entry = ttk.Entry(frame_dados, width=50)
            endereco_entry.insert(0, dados_cliente['endereco'])
            endereco_entry.grid(row=1, column=1, padx=5, pady=2)

            # Data Inicial
            ttk.Label(frame_dados, text="Data Inicial:").grid(row=2, column=0, padx=5, pady=2)
            data_entry = DateEntry(
                frame_dados,
                width=20,
                date_pattern='yyyy-mm-dd',
                locale='pt_BR'
            )
            if dados_cliente['data_inicial']:
                data_entry.set_date(dados_cliente['data_inicial'])
            data_entry.grid(row=2, column=1, padx=5, pady=2)

            # Observações
            ttk.Label(frame_dados, text="Observações:").grid(row=3, column=0, padx=5, pady=2)
            obs_entry = ttk.Entry(frame_dados, width=50)
            obs_entry.insert(0, dados_cliente['observacoes'] or '')
            obs_entry.grid(row=3, column=1, padx=5, pady=2)

            def salvar_alteracoes():
                try:
                    nome = nome_entry.get().strip()
                    endereco = endereco_entry.get().strip()
                    
                    if not nome or not endereco:
                        messagebox.showerror("Erro", "Nome e Endereço são obrigatórios!")
                        return

                    wb = load_workbook(ARQUIVO_CLIENTES)
                    ws = wb['Clientes']

                    # Remover registros antigos do cliente
                    linhas_para_remover = []
                    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                        if row[0].value == cliente_selecionado:
                            linhas_para_remover.append(idx)

                    for linha in reversed(linhas_para_remover):
                        ws.delete_rows(linha)

                    # Adicionar novo registro
                    proxima_linha = ws.max_row + 1
                    ws.cell(row=proxima_linha, column=1, value=nome.upper())
                    ws.cell(row=proxima_linha, column=2, value=endereco.upper())
                    ws.cell(row=proxima_linha, column=3, value=data_entry.get_date())
                    ws.cell(row=proxima_linha, column=4, value=obs_entry.get().upper())

                    wb.save(ARQUIVO_CLIENTES)
                    
                    # Atualizar nome do arquivo do cliente se mudou
                    if nome.upper() != cliente_selecionado:
                        caminho_antigo = PASTA_CLIENTES / f"{cliente_selecionado}.xlsx"
                        caminho_novo = PASTA_CLIENTES / f"{nome.upper()}.xlsx"
                        if os.path.exists(caminho_antigo):
                            os.rename(caminho_antigo, caminho_novo)

                    messagebox.showinfo("Sucesso", "Cliente atualizado com sucesso!")
                    self.atualizar_lista_clientes()
                    janela_edicao.destroy()

                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao salvar alterações: {str(e)}")

            # Frame para botões
            frame_botoes = ttk.Frame(frame)
            frame_botoes.pack(fill='x', pady=10)

            ttk.Button(frame_botoes, 
                      text="Salvar", 
                      command=salvar_alteracoes).pack(side='left', padx=5)
            ttk.Button(frame_botoes, 
                      text="Cancelar", 
                      command=janela_edicao.destroy).pack(side='left', padx=5)

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir editor: {str(e)}")   
    

    def atualizar_lista_clientes(self):
        """Atualiza a lista de clientes baseado nos arquivos Excel disponíveis"""
        try:
            # Carregar arquivo de clientes
            caminho_base = ARQUIVO_CLIENTES
            workbook = load_workbook(caminho_base)
            sheet = workbook['Clientes']  # Assumindo que existe uma aba chamada 'Clientes'
            
            # Limpar lista atual
            self.cliente_combobox['values'] = []
            
            # Pegar todos os clientes (pulando o cabeçalho)
            clientes = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Nome do cliente está na primeira coluna
                    clientes.append(row[0])
            
            # Atualizar combobox
            self.cliente_combobox['values'] = sorted(clientes)
            workbook.close()
            
        except FileNotFoundError:
            # Se o arquivo não existir, criar novo
            self.criar_arquivo_clientes()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar clientes: {str(e)}")





    def setup_aba_fornecedor(self):
        # Frame de busca
        frame_busca = ttk.LabelFrame(self.aba_fornecedor, text="Busca de Fornecedor")
        frame_busca.pack(fill='x', padx=5, pady=5)

        # Campo de busca
        tk.Label(frame_busca, text="Nome:").pack(side='left', padx=5)
        self.busca_entry = tk.Entry(frame_busca)
        self.busca_entry.pack(side='left', padx=5)
        self.busca_entry.bind('<Return>', lambda e: self.buscar_fornecedor())

        # Botão de busca
        tk.Button(frame_busca, 
                 text="Buscar", 
                 command=self.buscar_fornecedor).pack(side='left', padx=5)

        # Lista de resultados
        self.tree_fornecedores = ttk.Treeview(self.aba_fornecedor, 
                                             columns=('CNPJ/CPF', 'Nome', 'Categoria'),
                                             show='headings')
        
        self.tree_fornecedores.heading('CNPJ/CPF', text='CNPJ/CPF')
        self.tree_fornecedores.heading('Nome', text='Nome')
        self.tree_fornecedores.heading('Categoria', text='Categoria')
        self.tree_fornecedores.pack(fill='both', expand=True, padx=5, pady=5)

        # Botões de ação
        frame_acoes = ttk.Frame(self.aba_fornecedor)
        frame_acoes.pack(fill='x', padx=5, pady=5)

        tk.Button(frame_acoes, 
                 text="Novo Fornecedor", 
                 command=self.novo_fornecedor).pack(side='left', padx=5)
        tk.Button(frame_acoes, 
                 text="Editar Fornecedor", 
                 command=self.editar_fornecedor).pack(side='left', padx=5)
        tk.Button(frame_acoes, 
                 text="Selecionar", 
                 command=self.selecionar_fornecedor).pack(side='left', padx=5)

        # Adicionar frame de botões na parte inferior
        frame_botoes_fornecedor = ttk.Frame(self.aba_fornecedor)
        frame_botoes_fornecedor.pack(fill='x', padx=5, pady=5)

        ttk.Button(frame_botoes_fornecedor, 
                   text="Visualizar Lançamentos", 
                   command=self.visualizar_lancamentos).pack(side='left', padx=5)
        ttk.Button(frame_botoes_fornecedor, 
                   text="Enviar Registros", 
                   command=self.enviar_dados).pack(side='left', padx=5)
        ttk.Button(frame_botoes_fornecedor, 
                   text="Voltar ao Menu", 
                   command=self.voltar_menu).pack(side='left', padx=5)
        ttk.Button(frame_botoes_fornecedor, 
                   text="Sair", 
                   command=self.sair_sistema).pack(side='left', padx=5)



    def validar_tipo_despesa(self, P):
        """
        Valida entrada do tipo de despesa
        Args:
            P: valor proposto após a modificação
        """
        if P == "": return True  # Permite campo vazio
        if not P.isdigit(): return False  # Permite apenas dígitos
        return 1 <= int(P) <= 6  # Permite apenas valores entre 1 e 6


    def setup_aba_dados(self):
        frame_cliente = ttk.Frame(self.aba_dados)
        frame_cliente.pack(fill='x', padx=5, pady=5)
        
    
        self.cliente_label = tk.Label(frame_cliente, 
                                    text="Cliente: Nenhum selecionado", 
                                    font=('Arial', 14, 'bold'),
                                    anchor='w')
        self.cliente_label.pack(side='left')
        
        # Frame para data de referência
        frame_data = ttk.LabelFrame(self.aba_dados, text="Data de Referência")
        frame_data.pack(fill='x', padx=5, pady=5)
    
        def calcular_data_rel():
            hoje = datetime.now()
            if 6 <= hoje.day <= 20:
                data_rel = hoje.replace(day=20)
            else:
                if hoje.day > 20:
                    data_rel = (hoje + relativedelta(months=1)).replace(day=5)
                else:
                    data_rel = hoje.replace(day=5)
            return data_rel

        self.data_rel_entry = DateEntry(
            frame_data,
            format='dd/mm/yyyy',
            locale='pt_BR',
            font=('Helvetica', 12),
            background='darkblue',
            foreground='black',
            borderwidth=2,
            selectbackground='darkblue',
            selectforeground='white',
            normalbackground='white',
            normalforeground='black',
            headersbackground='darkblue',
            headersforeground='gray'
        )
        self.data_rel_entry.pack(side='left', padx=5, pady=5)
        
        # Definir data de referência inicial
        data_rel_inicial = calcular_data_rel()
        self.data_rel_entry.set_date(data_rel_inicial)
    
        def validar_entrada_data(event=None):
            data = self.data_rel_entry.get()
            if not validar_data(data):
                messagebox.showerror("Erro", "Data inválida! Use o formato dd/mm/aaaa")
                self.data_rel_entry.delete(0, tk.END)
                self.data_rel_entry.insert(0, datetime.now().strftime('%d/%m/%Y'))
                return False
            return True
    
        self.data_rel_entry.bind('<FocusOut>', validar_entrada_data)  # Valida quando perde o foco
    

        frame_fornecedor = ttk.LabelFrame(self.aba_dados, text="Dados do Fornecedor")
        frame_fornecedor.pack(fill='x', padx=5, pady=5)

        self.campos_fornecedor = {}
        campos = [('cnpj_cpf', 'CNPJ/CPF:'), 
                 ('nome', 'Nome:'), 
                 ('categoria', 'Categoria:'), 
                 ('dados_bancarios', 'Dados Bancários:')]
    
        for row, (campo, label) in enumerate(campos):
            tk.Label(frame_fornecedor, text=label).grid(row=row, column=0, padx=5, pady=2, sticky='e')
            entry = tk.Entry(frame_fornecedor, width=40)
            entry.grid(row=row, column=1, padx=5, pady=2, sticky='ew')
            if campo != 'categoria':
                entry.config(state='readonly')
            self.campos_fornecedor[campo] = entry

        frame_despesa = ttk.LabelFrame(self.aba_dados, text="Dados da Despesa")
        frame_despesa.pack(fill='x', padx=5, pady=5)

        # Frame para botões de parcelamento
        frame_parcelamento = ttk.Frame(self.aba_dados)
        frame_parcelamento.pack(fill='x', padx=5, pady=5)

        # Inicializar o gestor de parcelas com a janela root
        self.gestor_parcelas = GestorParcelas(self)
        
        ttk.Button(
            frame_parcelamento,
            text="Parcelar Despesa",
            command=self.abrir_parcelamento
        ).pack(side='left', padx=5)

        

         # Adicionar as opções de referência para tipo 1
        self.opcoes_referencia_tipo1 = [
            'DIÁRIA', 'SALÁRIO', 'TRANSPORTE', 
            'FÉRIAS', '13º SALÁRIO', 'RESCISÃO', 'CAFÉ'
        ]

        self.campos_despesa = {}
        campos_despesa = [
            ('tp_desp', 'Tipo Despesa (1-7):'),
            ('referencia', 'Referência:'),
            ('nf', 'NF:'),  # Novo campo
            ('vr_unit', 'Valor Unitário:'),
            ('dias', 'Dias:'),
            ('valor', 'Valor Total:'),
            ('dt_vencto', 'Data Vencimento:'),
            ('observacao', 'Observação:')
        ]

        for row, (campo, label) in enumerate(campos_despesa):
            tk.Label(frame_despesa, text=label).grid(row=row, column=0, padx=5, pady=2, sticky='e')
        
            if campo == 'dt_vencto':
                entry = DateEntry(
                    frame_despesa,
                    format='dd/mm/yyyy',
                    locale='pt_BR',
                    font=('Helvetica', 12),
                    background='darkblue',
                    foreground='black',
                    borderwidth=2,
                    selectbackground='darkblue',
                    selectforeground='white',
                    normalbackground='white',
                    normalforeground='black',
                    headersbackground='darkblue',
                    headersforeground='gray'
                )
                # Inicializa o campo vazio
                entry.delete(0, tk.END)
            elif campo == 'valor':
                entry = tk.Entry(frame_despesa, state='readonly')
            elif campo == 'referencia':
                entry = ttk.Combobox(frame_despesa)
                entry['values'] = self.opcoes_referencia_tipo1  # Define as opções na criação
                entry.bind('<<ComboboxSelected>>', lambda e: self.calcular_valor_total())  # Atualiza valor ao selecionar
            elif campo == 'tp_desp':
                vcmd = (frame_despesa.register(self.validar_tipo_despesa), '%P')
                entry = tk.Entry(frame_despesa, validate='key', validatecommand=vcmd)
            else:
                entry = tk.Entry(frame_despesa)
            
            entry.grid(row=row, column=1, padx=5, pady=2, sticky='ew')
            self.campos_despesa[campo] = entry

        self.campos_despesa['dias'].insert(0, "1")
        self.campos_despesa['vr_unit'].bind('<KeyRelease>', self.calcular_valor_total)
        self.campos_despesa['dias'].bind('<KeyRelease>', self.calcular_valor_total)
        self.campos_despesa['tp_desp'].bind('<KeyRelease>', self.verificar_tipo_despesa)
        

        frame_botoes = ttk.Frame(self.aba_dados)
        frame_botoes.pack(fill='x', padx=5, pady=10)

        ttk.Button(frame_botoes, text="Adicionar", command=self.adicionar_dados).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Visualizar Lançamentos", command=self.visualizar_lancamentos).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Enviar", command=self.enviar_dados).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Cancelar", command=self.cancelar_entrada).pack(side='left', padx=5)



    def visualizar_lancamentos(self):
        """Abre a janela de visualização de lançamentos pendentes"""
        if hasattr(self, 'visualizador') and self.visualizador:
            self.visualizador.janela.destroy()
        
        # Criar nova instância do visualizador
        self.visualizador = VisualizadorLancamentos(self)
        
        # Configurar callback para quando a janela for fechada
        self.visualizador.janela.protocol("WM_DELETE_WINDOW", self.on_visualizador_close)
        
        # Atualizar dados
        self.visualizador.dados_para_incluir = self.dados_para_incluir.copy()
        self.visualizador.atualizar_dados(self.dados_para_incluir)

    def on_visualizador_close(self):
        """Manipula o fechamento da janela do visualizador"""
        # Atualizar dados_para_incluir com os dados mais recentes do visualizador
        if self.visualizador:
            self.dados_para_incluir = self.visualizador.get_dados_atualizados()
            self.visualizador.janela.destroy()
            self.visualizador = None



    def processar_parcelas(self):
        """Processa as parcelas geradas mantendo os dados do fornecedor"""
        print("Iniciando processamento de parcelas...")
        
        # Verificar se há parcelas para processar
        if not hasattr(self, 'gestor_parcelas') or not self.gestor_parcelas.parcelas:
            print("Nenhuma parcela para processar")
            return False
            
        # Validar se há fornecedor selecionado
        if not self.campos_fornecedor['cnpj_cpf'].get():
            messagebox.showerror("Erro", "Selecione um fornecedor antes de processar as parcelas!")
            return False
            
        # Guardar dados do fornecedor atual
        dados_fornecedor = {
            'cnpj_cpf': self.campos_fornecedor['cnpj_cpf'].get(),
            'nome': self.campos_fornecedor['nome'].get(),
            'categoria': self.campos_fornecedor['categoria'].get(),
            'dados_bancarios': self.campos_fornecedor['dados_bancarios'].get()
        }
        
        print(f"Dados do fornecedor capturados: {dados_fornecedor}")
        total_parcelas = len(self.gestor_parcelas.parcelas)
        print(f"Total de parcelas a processar: {total_parcelas}")
        
        try:
            processadas = 0
            for i, parcela in enumerate(self.gestor_parcelas.parcelas, 1):
                print(f"\nProcessando parcela {i} de {total_parcelas}")
                
                # Restaurar dados do fornecedor antes de cada parcela
                for campo, valor in dados_fornecedor.items():
                    entry = self.campos_fornecedor[campo]
                    entry.config(state='normal')
                    entry.delete(0, tk.END)
                    entry.insert(0, valor)
                    if campo != 'categoria':
                        entry.config(state='readonly')
                
                print(f"Dados do fornecedor restaurados para parcela {i}")

                
                # Preencher dados da parcela
                self.data_rel_entry.set_date(datetime.strptime(parcela['data_rel'], '%d/%m/%Y'))
                self.campos_despesa['tp_desp'].delete(0, tk.END)
                self.campos_despesa['tp_desp'].insert(0, self.gestor_parcelas.tipo_despesa_valor)
                self.campos_despesa['nf'].delete(0, tk.END)
                self.campos_despesa['nf'].insert(0, parcela['nf'])
                
                if isinstance(self.campos_despesa['referencia'], ttk.Combobox):
                    self.campos_despesa['referencia'].set(parcela['referencia'])
                else:
                    self.campos_despesa['referencia'].delete(0, tk.END)
                    self.campos_despesa['referencia'].insert(0, parcela['referencia'])
                
                self.campos_despesa['vr_unit'].delete(0, tk.END)
                self.campos_despesa['vr_unit'].insert(0, f"{parcela['valor']:.2f}")
                
                self.campos_despesa['valor'].config(state='normal')
                self.campos_despesa['valor'].delete(0, tk.END)
                self.campos_despesa['valor'].insert(0, f"{parcela['valor']:.2f}")
                self.campos_despesa['valor'].config(state='readonly')
                
                self.campos_despesa['dt_vencto'].set_date(
                    datetime.strptime(parcela['dt_vencto'], '%d/%m/%Y')
                )
                
                # Adicionar à lista de dados e verificar sucesso
                if self.adicionar_dados(eh_parcelamento=True):
                    processadas += 1
                    print(f"Parcela {i} processada com sucesso")
                else:
                    print(f"Falha ao processar parcela {i}")
            
            # Relatório final
            if processadas == total_parcelas:
                messagebox.showinfo("Sucesso", 
                                  f"Todas as {total_parcelas} parcelas foram processadas com sucesso!")
            else:
                messagebox.showwarning("Aviso", 
                                     f"Apenas {processadas} de {total_parcelas} parcelas foram processadas.")
            
            return processadas == total_parcelas
            
        except Exception as e:
            erro_msg = f"Erro ao processar parcelas: {str(e)}"
            print(erro_msg)
            messagebox.showerror("Erro", erro_msg)
            return False
            
        finally:
            self.limpar_campos_despesa()
            print("Processamento de parcelas finalizado")



    def abrir_parcelamento(self):
        """Abre a janela de parcelamento e processa os dados após o fechamento"""
        print("\nIniciando processo de parcelamento...")
        
        # Verificar se há fornecedor selecionado
        cnpj_cpf = self.campos_fornecedor['cnpj_cpf'].get()
        if not cnpj_cpf:
            print("Erro: Fornecedor não selecionado")
            messagebox.showerror("Erro", "Selecione um fornecedor antes de criar parcelas!")
            return

        print("\nCapturando dados do fornecedor...")
        dados_fornecedor = {
            'cnpj_cpf': cnpj_cpf,
            'nome': self.campos_fornecedor['nome'].get(),
            'categoria': self.campos_fornecedor['categoria'].get(),
            'dados_bancarios': self.campos_fornecedor['dados_bancarios'].get()
        }
        print(f"Dados capturados: {dados_fornecedor}")
        
        # Validar se todos os campos do fornecedor estão preenchidos
        if not all(dados_fornecedor.values()):
            print("Erro: Dados do fornecedor incompletos")
            messagebox.showerror("Erro", "Dados do fornecedor incompletos!")
            return

        print("Abrindo janela de parcelamento...")
        self.gestor_parcelas.abrir_janela_parcelas()
        self.root.wait_window(self.gestor_parcelas.janela_parcelas)

        if hasattr(self.gestor_parcelas, 'parcelas') and self.gestor_parcelas.parcelas:
            print(f"Processando {len(self.gestor_parcelas.parcelas)} parcelas...")
            
            success = True
            for i, parcela in enumerate(self.gestor_parcelas.parcelas, 1):
                try:
                    print(f"\nProcessando parcela {i}")
                    
                    # Restaurar dados do fornecedor
                    for campo, valor in dados_fornecedor.items():
                        entry = self.campos_fornecedor[campo]
                        entry.config(state='normal')
                        entry.delete(0, tk.END)
                        entry.insert(0, valor)
                        if campo != 'categoria':
                            entry.config(state='readonly')
                    
                    # Preencher dados da parcela
                    self.data_rel_entry.set_date(datetime.strptime(parcela['data_rel'], '%d/%m/%Y'))
                    
                    self.campos_despesa['tp_desp'].delete(0, tk.END)
                    self.campos_despesa['tp_desp'].insert(0, self.gestor_parcelas.tipo_despesa_valor)
                    self.campos_despesa['nf'].delete(0, tk.END)
                    self.campos_despesa['nf'].insert(0, parcela['nf'])
                    
                    if isinstance(self.campos_despesa['referencia'], ttk.Combobox):
                        self.campos_despesa['referencia'].set(parcela['referencia'])
                    else:
                        self.campos_despesa['referencia'].delete(0, tk.END)
                        self.campos_despesa['referencia'].insert(0, parcela['referencia'])
                    
                    self.campos_despesa['vr_unit'].delete(0, tk.END)
                    self.campos_despesa['vr_unit'].insert(0, f"{parcela['valor']:.2f}")
                    
                    self.campos_despesa['dias'].delete(0, tk.END)
                    self.campos_despesa['dias'].insert(0, '1')
                    
                    self.campos_despesa['valor'].config(state='normal')
                    self.campos_despesa['valor'].delete(0, tk.END)
                    self.campos_despesa['valor'].insert(0, f"{parcela['valor']:.2f}")
                    self.campos_despesa['valor'].config(state='readonly')
                    
                    self.campos_despesa['dt_vencto'].set_date(
                        datetime.strptime(parcela['dt_vencto'], '%d/%m/%Y')
                    )
                    
                    # Adicionar à lista de dados
                    if not self.adicionar_dados(eh_parcelamento=True):
                        print(f"Falha ao adicionar parcela {i}")
                        success = False
                        break
                    
                    print(f"Parcela {i} processada com sucesso")
                    
                except Exception as e:
                    success = False
                    print(f"Erro ao processar parcela {i}: {str(e)}")
                    messagebox.showerror("Erro", f"Erro ao processar parcela {i}: {str(e)}")
                    break
            
            if success:
                messagebox.showinfo("Sucesso", 
                                  f"Todas as {len(self.gestor_parcelas.parcelas)} parcelas foram processadas!")
                # Calcular a data de referência padrão
                hoje = datetime.now()
                if 6 <= hoje.day <= 20:
                    data_rel = hoje.replace(day=20)
                else:
                    if hoje.day > 20:
                        data_rel = (hoje + relativedelta(months=1)).replace(day=5)
                    else:
                        data_rel = hoje.replace(day=5)
                
                # Restaurar a data de referência padrão
                self.data_rel_entry.set_date(data_rel)
                
                self.limpar_campos_despesa()
                self.notebook.select(1)  # Volta para aba fornecedor
            else:
                messagebox.showerror("Erro", "Houve um erro no processamento das parcelas.")
        else:
            print("Nenhuma parcela para processar")



    def abrir_calendario(self):
        try:
            top = Toplevel(self.root)
            top.title("Selecionar Data")
            top.geometry("300x250")
            top.grab_set()  # Torna a janela modal
        
            cal = Calendar(top, selectmode='day', 
                          date_pattern='dd/mm/yyyy',
                          locale='pt_BR')
            cal.pack(padx=10, pady=10)
        
            def selecionar_data():
                data = cal.get_date()
                self.data_rel_entry.delete(0, tk.END)
                self.data_rel_entry.insert(0, data)
                top.destroy()
        
            ttk.Button(top, text="OK", command=selecionar_data).pack(pady=5)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir calendário: {str(e)}")

    def atualizar_campo_referencia(self, event=None):
        """Atualiza o campo de referência baseado no tipo de despesa"""
        tp_desp = self.campos_despesa['tp_desp'].get().strip()
    
        try:
            if tp_desp == '1':
                # Redefine as opções e configura como readonly
                self.campos_despesa['referencia']['values'] = self.opcoes_referencia_tipo1
                self.campos_despesa['referencia'].config(state='readonly')
                # Seleciona o primeiro item como padrão
                if self.opcoes_referencia_tipo1:
                    self.campos_despesa['referencia'].set(self.opcoes_referencia_tipo1[0])
            else:
                # Para outros tipos, limpa a seleção e permite digitação
                self.campos_despesa['referencia'].set('')
                self.campos_despesa['referencia']['values'] = []
                self.campos_despesa['referencia'].config(state='normal')
            
        except Exception as e:
            print(f"Erro ao atualizar campo referência: {str(e)}")

    def atualizar_dados_bancarios(self, event=None):
        """Atualiza os dados bancários baseado no tipo de despesa"""
        tp_desp = self.campos_despesa['tp_desp'].get().strip()
        cnpj_cpf = self.campos_fornecedor['cnpj_cpf'].get().strip()
    
        if not cnpj_cpf:  # Se não houver fornecedor selecionado
            return
        
        fornecedor_completo = self.buscar_fornecedor_completo(cnpj_cpf)
        if not fornecedor_completo:
            return
        
        self.campos_fornecedor['dados_bancarios'].config(state='normal')
        self.campos_fornecedor['dados_bancarios'].delete(0, tk.END)
    
        if tp_desp not in ['3', '5']:  # Mostra dados bancários para todos EXCETO tipos 3 e 5
            if fornecedor_completo['chave_pix']:
                dados_bancarios = f"PIX: {fornecedor_completo['chave_pix']}"
            else:
                dados_bancarios = (f"{fornecedor_completo['banco'] or ''} "
                                 f"{fornecedor_completo['op'] or ''} - "
                                 f"{fornecedor_completo['agencia'] or ''} "
                                 f"{fornecedor_completo['conta'] or ''}").strip()
            if dados_bancarios.strip() in ['', ' - ']:
                dados_bancarios = 'DADOS BANCÁRIOS NÃO CADASTRADOS'
        else:
            dados_bancarios = ''
        
        self.campos_fornecedor['dados_bancarios'].insert(0, dados_bancarios)
        self.campos_fornecedor['dados_bancarios'].config(state='readonly')
    

    def cancelar_entrada(self):
        """Cancela a entrada de dados atual e retorna à aba fornecedor"""
        if any(self.campos_despesa[campo].get() for campo in ['tp_desp', 'referencia', 'vr_unit']):
            if messagebox.askyesno("Confirmação", "Deseja descartar os dados atuais?"):
                self.limpar_campos_despesa()
                self.notebook.select(1)  # Volta para aba fornecedor
        else:
            self.notebook.select(1)  # Volta para aba fornecedor





    def buscar_fornecedor(self):
        termo = self.busca_entry.get()
        buscar_fornecedor(self.tree_fornecedores, termo)

    def novo_fornecedor(self):
        """Abre janela para cadastro de novo fornecedor"""
        self.janela_fornecedor = tk.Toplevel(self.root)
        self.janela_fornecedor.title("Novo Fornecedor")
        self.setup_formulario_fornecedor()

    def editar_fornecedor(self):
        """Abre janela para edição de fornecedor existente"""
        selecionado = self.tree_fornecedores.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um fornecedor para editar")
            return

        # Buscar dados completos do fornecedor
        fornecedor = self.buscar_fornecedor_completo(
            self.tree_fornecedores.item(selecionado)['values'][0]
        )
        if not fornecedor:
            messagebox.showerror("Erro", "Fornecedor não encontrado")
            return

        # Criar janela de edição
        self.janela_fornecedor = tk.Toplevel(self.root)
        self.janela_fornecedor.title("Editar Fornecedor")
        self.setup_formulario_fornecedor(modo_edicao=True)

        try:
            # Determinar tipo de pessoa baseado no tamanho do CNPJ/CPF
            cnpj_cpf = str(fornecedor['cnpj_cpf']).strip()
            tipo_pessoa = 'PJ' if len(cnpj_cpf) > 11 else 'PF'

            # Preencher e configurar campos não editáveis
            # CNPJ/CPF
            self.campos_form['cnpj_cpf'].insert(0, cnpj_cpf.zfill(14 if tipo_pessoa == 'PJ' else 11))
            self.campos_form['cnpj_cpf'].config(state='readonly')
            
            # Tipo Pessoa
            self.campos_form['tipo_pessoa'].set(tipo_pessoa)
            self.campos_form['tipo_pessoa'].config(state='disabled')
            
            # Razão Social
            self.campos_form['razao_social'].insert(0, fornecedor['razao_social'] or '')
            self.campos_form['razao_social'].config(state='readonly')
            
            # Preencher campos editáveis
            self.campos_form['nome'].insert(0, fornecedor['nome'] or '')
            self.campos_form['telefone'].insert(0, fornecedor['telefone'] or '')
            self.campos_form['email'].insert(0, fornecedor['email'] or '')
            self.campos_form['banco'].insert(0, fornecedor['banco'] or '')
            self.campos_form['op'].insert(0, fornecedor['op'] or '')
            self.campos_form['agencia'].insert(0, fornecedor['agencia'] or '')
            self.campos_form['conta'].insert(0, fornecedor['conta'] or '')
            self.campos_form['chave_pix'].insert(0, fornecedor['chave_pix'] or '')
            
            # Categoria (pode ser combobox)
            if isinstance(self.campos_form['categoria'], ttk.Combobox):
                self.campos_form['categoria'].set(fornecedor['categoria'] or '')
            else:
                self.campos_form['categoria'].insert(0, fornecedor['categoria'] or '')
                
            self.campos_form['especificacao'].insert(0, fornecedor['especificacao'] or '')
            self.campos_form['vinculo'].insert(0, fornecedor['vinculo'] or '')

            # Centralizar a janela
            self.janela_fornecedor.update_idletasks()
            width = self.janela_fornecedor.winfo_width()
            height = self.janela_fornecedor.winfo_height()
            x = (self.janela_fornecedor.winfo_screenwidth() // 2) - (width // 2)
            y = (self.janela_fornecedor.winfo_screenheight() // 2) - (height // 2)
            self.janela_fornecedor.geometry('{}x{}+{}+{}'.format(width, height, x, y))
            
            # Tornar a janela modal
            self.janela_fornecedor.transient(self.root)
            self.janela_fornecedor.grab_set()

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados do fornecedor: {str(e)}")
            self.janela_fornecedor.destroy()

    def selecionar_fornecedor(self):
        """Seleciona o fornecedor e preenche seus dados"""
        fornecedor = selecionar_fornecedor(
            self.tree_fornecedores, 
            self.campos_fornecedor,
            self.campos_despesa,
            self.notebook,
            self.buscar_fornecedor_completo
        )
        if fornecedor:
            # Formatar CNPJ/CPF
            cnpj_cpf = str(fornecedor[0]).strip()
            self.campos_fornecedor['cnpj_cpf'].config(state='normal')
            self.campos_fornecedor['cnpj_cpf'].delete(0, tk.END)
            self.campos_fornecedor['cnpj_cpf'].insert(0, formatar_cnpj_cpf(cnpj_cpf))
            self.campos_fornecedor['cnpj_cpf'].config(state='readonly')
            
            # Carregar dados completos do fornecedor
            fornecedor_completo = self.buscar_fornecedor_completo(cnpj_cpf)
            if fornecedor_completo:
                self.campos_fornecedor['categoria'].delete(0, tk.END)
                self.campos_fornecedor['categoria'].insert(0, fornecedor_completo['categoria'])
                
                self.campos_fornecedor['dados_bancarios'].config(state='normal')
                self.campos_fornecedor['dados_bancarios'].delete(0, tk.END)
                
                # Construir dados bancários
                if fornecedor_completo['chave_pix']:
                    dados_bancarios = f"PIX: {fornecedor_completo['chave_pix']}"
                else:
                    dados_bancarios = (f"{fornecedor_completo['banco'] or ''} "
                                     f"{fornecedor_completo['op'] or ''} - "
                                     f"{fornecedor_completo['agencia'] or ''} "
                                     f"{fornecedor_completo['conta'] or ''}").strip()
                if dados_bancarios.strip() in ['', ' - ']:
                    dados_bancarios = 'DADOS BANCÁRIOS NÃO CADASTRADOS'
                
                self.campos_fornecedor['dados_bancarios'].insert(0, dados_bancarios)
                self.campos_fornecedor['dados_bancarios'].config(state='readonly')
            
            self.notebook.select(2)  # Vai para aba de dados
            

    def buscar_dados_bancarios(self, cnpj_cpf):
        try:
            wb = load_workbook(ARQUIVO_FORNECEDORES, data_only=True)
            ws = wb['Fornecedores']
        
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == cnpj_cpf:
                    print(f"CNPJ/CPF encontrado: {cnpj_cpf}")
                    print(f"Dados da linha: {row}")
                    if row[14]:  # coluna O com dados bancários consolidados
                        return row[14]
                    return ""
            return ""
        except Exception as e:
            print(f"Erro ao buscar dados bancários: {e}")
            return ""

    

    def buscar_fornecedor_completo(self, cnpj_cpf):
        """Busca todos os dados de um fornecedor"""
        try:
            wb = load_workbook(ARQUIVO_FORNECEDORES, data_only=True)
            ws = wb['Fornecedores']
        
            cnpj_cpf = str(cnpj_cpf).zfill(14)  # Preenche com zeros à esquerda
            for row in ws.iter_rows(min_row=2):
                # Garante que o CNPJ/CPF da planilha também tenha 14 dígitos
                row_cnpj = str(row[0].value or '').zfill(14)
                if row_cnpj == cnpj_cpf:
                    fornecedor = {
                        'cnpj_cpf': row[0].value,
                        'tipo_pessoa': row[1].value,
                        'razao_social': row[2].value,
                        'nome': row[3].value,
                        'telefone': row[4].value,
                        'email': row[5].value,
                        'banco': row[6].value,
                        'op': row[7].value,
                        'agencia': row[8].value,
                        'conta': row[9].value,
                        'chave_pix': row[10].value,
                        'categoria': row[11].value,
                        'especificacao': row[12].value,
                        'vinculo': row[13].value,
                    }
                    return fornecedor
            return None
        except Exception as e:
            print(f"Erro ao buscar fornecedor: {e}")
            return None

        
    def setup_formulario_fornecedor(self, modo_edicao=False):
        """Configura o formulário de cadastro/edição de fornecedor"""
        formulario = ttk.Frame(self.janela_fornecedor)
        formulario.pack(padx=10, pady=5, fill='both', expand=True)

        # Campos principais
        campos_principais = ttk.LabelFrame(formulario, text="Dados Principais")
        campos_principais.pack(fill='x', pady=5)

        self.campos_form = {}

        # CNPJ/CPF na mesma linha
        tk.Label(campos_principais, text="CNPJ/CPF:*").grid(row=0, column=0, padx=5, pady=2)
        self.campos_form['cnpj_cpf'] = tk.Entry(campos_principais)
        self.campos_form['cnpj_cpf'].grid(row=0, column=1, padx=5, pady=2, sticky='ew')
        self.campos_form['cnpj_cpf'].bind('<FocusOut>', self.atualizar_tipo_pessoa)
        
        tk.Label(campos_principais, text="Tipo:*").grid(row=0, column=2, padx=5, pady=2)
        self.campos_form['tipo_pessoa'] = ttk.Combobox(campos_principais, 
                                                      values=['PF', 'PJ'],
                                                      state='readonly',
                                                      width=5)
        self.campos_form['tipo_pessoa'].grid(row=0, column=3, padx=5, pady=2, sticky='w')
        
        # Razão Social e Nome
        tk.Label(campos_principais, text="Razão Social:*").grid(row=1, column=0, padx=5, pady=2)
        self.campos_form['razao_social'] = tk.Entry(campos_principais)
        self.campos_form['razao_social'].grid(row=1, column=1, columnspan=3, padx=5, pady=2, sticky='ew')
        self.campos_form['razao_social'].bind('<FocusOut>', self.copiar_para_nome)
        
        tk.Label(campos_principais, text="Nome Fantasia:*").grid(row=2, column=0, padx=5, pady=2)
        self.campos_form['nome'] = tk.Entry(campos_principais)
        self.campos_form['nome'].grid(row=2, column=1, columnspan=3, padx=5, pady=2, sticky='ew')


        # Contatos
        campos_contato = ttk.LabelFrame(formulario, text="Contato")
        campos_contato.pack(fill='x', pady=5)

        tk.Label(campos_contato, text="Telefone:").grid(row=0, column=0, padx=5, pady=2)
        self.campos_form['telefone'] = tk.Entry(campos_contato)
        self.campos_form['telefone'].grid(row=0, column=1, padx=5, pady=2, sticky='ew')

        tk.Label(campos_contato, text="Email:").grid(row=1, column=0, padx=5, pady=2)
        self.campos_form['email'] = tk.Entry(campos_contato)
        self.campos_form['email'].grid(row=1, column=1, padx=5, pady=2, sticky='ew')

        # Dados Bancários
        campos_bancarios = ttk.LabelFrame(formulario, text="Dados Bancários")
        campos_bancarios.pack(fill='x', pady=5)

        tk.Label(campos_bancarios, text="Banco:").grid(row=0, column=0, padx=5, pady=2)
        self.campos_form['banco'] = tk.Entry(campos_bancarios)
        self.campos_form['banco'].grid(row=0, column=1, padx=5, pady=2, sticky='ew')

        tk.Label(campos_bancarios, text="Operação:").grid(row=1, column=0, padx=5, pady=2)
        self.campos_form['op'] = tk.Entry(campos_bancarios)
        self.campos_form['op'].grid(row=1, column=1, padx=5, pady=2, sticky='ew')

        tk.Label(campos_bancarios, text="Agência:").grid(row=2, column=0, padx=5, pady=2)
        self.campos_form['agencia'] = tk.Entry(campos_bancarios)
        self.campos_form['agencia'].grid(row=2, column=1, padx=5, pady=2, sticky='ew')

        tk.Label(campos_bancarios, text="Conta:").grid(row=3, column=0, padx=5, pady=2)
        self.campos_form['conta'] = tk.Entry(campos_bancarios)
        self.campos_form['conta'].grid(row=3, column=1, padx=5, pady=2, sticky='ew')

        # PIX
        campos_pix = ttk.LabelFrame(formulario, text="Chave PIX")
        campos_pix.pack(fill='x', pady=5)

        # Tipo de chave PIX
        tk.Label(campos_pix, text="Tipo de Chave:").grid(row=0, column=0, padx=5, pady=2)
        self.tipo_pix = ttk.Combobox(campos_pix, 
                                    values=['Selecione', 'CNPJ/CPF', 'Telefone', 'Email'])
        self.tipo_pix.grid(row=0, column=1, padx=5, pady=2, sticky='ew')
        self.tipo_pix.set('Selecione')
        if not modo_edicao:  # Só adiciona o binding se não estiver em modo edição
            self.tipo_pix.bind('<<ComboboxSelected>>', self.atualizar_chave_pix)

        tk.Label(campos_pix, text="Chave:").grid(row=1, column=0, padx=5, pady=2)
        self.campos_form['chave_pix'] = tk.Entry(campos_pix)
        self.campos_form['chave_pix'].grid(row=1, column=1, padx=5, pady=2, sticky='ew')

        # Classificação
        campos_class = ttk.LabelFrame(formulario, text="Classificação")
        campos_class.pack(fill='x', pady=5)

        # Categoria
        tk.Label(campos_class, text="Categoria:*").grid(row=0, column=0, padx=5, pady=2)
        self.campos_form['categoria'] = ttk.Combobox(campos_class, 
                                                    values=['ADM', 'DIV', 'LOC', 'MAT', 'MO', 'SERV', 'TP'])
        self.campos_form['categoria'].grid(row=0, column=1, padx=5, pady=2, sticky='ew')

        # Especificação
        tk.Label(campos_class, text="Especificação:").grid(row=1, column=0, padx=5, pady=2)
        self.campos_form['especificacao'] = tk.Entry(campos_class)
        self.campos_form['especificacao'].grid(row=1, column=1, padx=5, pady=2, sticky='ew')

        # Vínculo
        tk.Label(campos_class, text="Vínculo:").grid(row=2, column=0, padx=5, pady=2)
        self.campos_form['vinculo'] = tk.Entry(campos_class)
        self.campos_form['vinculo'].grid(row=2, column=1, padx=5, pady=2, sticky='ew')

        # Botões de ação
        frame_botoes = ttk.Frame(formulario)
        frame_botoes.pack(fill='x', pady=10)

        ttk.Button(frame_botoes, 
                   text="Salvar", 
                   command=self.salvar_fornecedor).pack(side='left', padx=5)
        ttk.Button(frame_botoes, 
                   text="Cancelar", 
                   command=self.janela_fornecedor.destroy).pack(side='left', padx=5)

    def atualizar_chave_pix(self, event=None):
        """Atualiza o campo de chave PIX baseado no tipo selecionado"""
        tipo_selecionado = self.tipo_pix.get()
        self.campos_form['chave_pix'].delete(0, tk.END)
        
        if tipo_selecionado == 'CNPJ/CPF':
            self.campos_form['chave_pix'].insert(0, self.campos_form['cnpj_cpf'].get())
        elif tipo_selecionado == 'Telefone':
            self.campos_form['chave_pix'].insert(0, self.campos_form['telefone'].get())
        elif tipo_selecionado == 'Email':
            self.campos_form['chave_pix'].insert(0, self.campos_form['email'].get())


    def atualizar_tipo_pessoa(self, event=None):
        """Determina automaticamente o tipo de pessoa baseado no CNPJ/CPF"""
        cnpj_cpf = self.campos_form['cnpj_cpf'].get().strip()
        # Remove caracteres não numéricos
        cnpj_cpf = ''.join(filter(str.isdigit, cnpj_cpf))
        
        if len(cnpj_cpf) <= 11:
            self.campos_form['tipo_pessoa'].set('PF')
        else:
            self.campos_form['tipo_pessoa'].set('PJ')

    def copiar_para_nome(self, event=None):
        """Copia a razão social para o nome se este estiver vazio"""
        razao_social = self.campos_form['razao_social'].get().strip()
        nome_atual = self.campos_form['nome'].get().strip()
        
        if razao_social and not nome_atual:
            self.campos_form['nome'].insert(0, razao_social)



    def salvar_fornecedor(self):
        """Salva os dados do fornecedor"""
        # Validar campos obrigatórios
        campos_obrigatorios = ['tipo_pessoa', 'cnpj_cpf', 'razao_social', 'nome', 'categoria']
        for campo in campos_obrigatorios:
            if not self.campos_form[campo].get().strip():
                messagebox.showerror("Erro", f"O campo {campo} é obrigatório!")
                return

        # Validar CNPJ/CPF
        tipo_pessoa = self.campos_form['tipo_pessoa'].get()
        cnpj_cpf = self.campos_form['cnpj_cpf'].get().strip()
        
        if not validar_cnpj_cpf(cnpj_cpf):
            messagebox.showerror("Erro", f"{'CPF' if tipo_pessoa == 'PF' else 'CNPJ'} inválido!")
            return
            
            cnpj_cpf = formatar_cnpj_cpf(cnpj_cpf)

        # Montar dados bancários
        if self.campos_form['chave_pix'].get():
            dados_bancarios = f"PIX: {self.campos_form['chave_pix'].get()}"
        else:
            dados_bancarios = (f"{self.campos_form['banco'].get()} "
                             f"{self.campos_form['op'].get()} - "
                             f"{self.campos_form['agencia'].get()} "
                             f"{self.campos_form['conta'].get()}").strip()

        # Preparar dados para salvar
        dados = {
            'tipo_pessoa': tipo_pessoa,
            'cnpj_cpf': cnpj_cpf,
            'razao_social': self.campos_form['razao_social'].get().upper(),
            'nome': self.campos_form['nome'].get().upper(),
            'telefone': self.campos_form['telefone'].get(),
            'email': self.campos_form['email'].get(),
            'banco': self.campos_form['banco'].get(),
            'op': self.campos_form['op'].get(),
            'agencia': self.campos_form['agencia'].get(),
            'conta': self.campos_form['conta'].get(),
            'chave_pix': self.campos_form['chave_pix'].get(),
            'categoria': self.campos_form['categoria'].get().upper(),
            'especificacao': self.campos_form['especificacao'].get().upper(),
            'vinculo': self.campos_form['vinculo'].get().upper(),
            'dados_bancarios': dados_bancarios
        }

        try:
            self.salvar_na_base_fornecedores(dados)
            messagebox.showinfo("Sucesso", "Fornecedor salvo com sucesso!")
            self.janela_fornecedor.destroy()
            self.buscar_fornecedor()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar fornecedor: {str(e)}")

    def salvar_na_base_fornecedores(self, dados):
        """Salva os dados na planilha de fornecedores"""
        try:
            wb = load_workbook(ARQUIVO_FORNECEDORES)
            ws = wb['Fornecedores']
            
            # Verificar se o fornecedor já existe
            for row in ws.iter_rows(min_row=2):
                if row[0].value == dados['cnpj_cpf']:  # Atualiza se já existe
                    self.atualizar_linha_fornecedor(row, dados)
                    break
            else:  # Adiciona novo se não existe
                proxima_linha = ws.max_row + 1
                self.adicionar_linha_fornecedor(ws, proxima_linha, dados)
            
            wb.save(ARQUIVO_FORNECEDORES)
        except Exception as e:
            raise Exception(f"Erro ao salvar na planilha: {str(e)}")

    def atualizar_linha_fornecedor(self, row, dados):
        """Atualiza uma linha existente com novos dados"""
        row[0].value = dados['cnpj_cpf']
        row[1].value = dados['tipo_pessoa']  # Nova coluna para tipo de pessoa
        row[2].value = dados['razao_social']
        row[3].value = dados['nome']
        row[4].value = dados['telefone']
        row[5].value = dados['email']
        row[6].value = dados['banco']
        row[7].value = dados['op']
        row[8].value = dados['agencia']
        row[9].value = dados['conta']
        row[10].value = dados['chave_pix']
        row[11].value = dados['categoria']
        row[12].value = dados['especificacao']
        row[13].value = dados['vinculo']
        row[14].value = dados['dados_bancarios']

    def adicionar_linha_fornecedor(self, ws, linha, dados):
        """Adiciona uma nova linha com os dados do fornecedor"""
        ws.cell(row=linha, column=1, value=dados['cnpj_cpf'])
        ws.cell(row=linha, column=2, value=dados['tipo_pessoa'])
        ws.cell(row=linha, column=3, value=dados['razao_social'])
        ws.cell(row=linha, column=4, value=dados['nome'])
        ws.cell(row=linha, column=5, value=dados['telefone'])
        ws.cell(row=linha, column=6, value=dados['email'])
        ws.cell(row=linha, column=7, value=dados['banco'])
        ws.cell(row=linha, column=8, value=dados['op'])
        ws.cell(row=linha, column=9, value=dados['agencia'])
        ws.cell(row=linha, column=10, value=dados['conta'])
        ws.cell(row=linha, column=11, value=dados['chave_pix'])
        ws.cell(row=linha, column=12, value=dados['categoria'])
        ws.cell(row=linha, column=13, value=dados['especificacao'])
        ws.cell(row=linha, column=14, value=dados['vinculo'])
        ws.cell(row=linha, column=15, value=dados['dados_bancarios'])

        
    def atualizar_fornecedor(self):
        """Atualiza dados do fornecedor existente"""
        # Validações semelhantes ao salvar_fornecedor
        campos_obrigatorios = ['razao_social', 'nome', 'categoria']
        for campo in campos_obrigatorios:
            if not self.campos_form[campo].get().strip():
                messagebox.showerror("Erro", f"O campo {campo} é obrigatório!")
                return

        try:
            wb = load_workbook(ARQUIVO_FORNECEDORES)
            ws = wb['Fornecedores']
            
            cnpj_cpf = self.campos_form['cnpj_cpf'].get()
            for row in ws.iter_rows(min_row=2):
                if row[0].value == cnpj_cpf:
                    # Atualizar dados na linha existente
                    row[1].value = self.campos_form['tipo_pessoa'].get().upper()
                    row[2].value = self.campos_form['razao_social'].get().upper()
                    row[3].value = self.campos_form['nome'].get().upper()
                    row[4].value = self.campos_form['telefone'].get()
                    row[5].value = self.campos_form['email'].get()
                    row[6].value = self.campos_form['banco'].get()
                    row[7].value = self.campos_form['op'].get()
                    row[8].value = self.campos_form['agencia'].get()
                    row[9].value = self.campos_form['conta'].get()
                    row[10].value = self.campos_form['chave_pix'].get()
                    row[11].value = self.campos_form['categoria'].get()
                    row[12].value = self.campos_form['especificacao'].get().upper()
                    row[13].value = self.campos_form['vinculo'].get().upper()
                    row[14].value = self.campos_form['dados_bancarios'].get().upper()
                    break

            wb.save(ARQUIVO_FORNECEDORES)
            messagebox.showinfo("Sucesso", "Fornecedor atualizado com sucesso!")
            self.janela_fornecedor.destroy()
            self.buscar_fornecedor()  # Atualiza a lista
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao atualizar fornecedor: {str(e)}")


    def preencher_dados_fornecedor(self, dados):
        """Preenche os campos do fornecedor na aba de entrada"""
        self.campos_fornecedor['cnpj_cpf'].delete(0, tk.END)
        self.campos_fornecedor['cnpj_cpf'].insert(0, dados[0])
        
        self.campos_fornecedor['nome'].delete(0, tk.END)
        self.campos_fornecedor['nome'].insert(0, dados[1])
        
        self.campos_fornecedor['categoria'].delete(0, tk.END)
        self.campos_fornecedor['categoria'].insert(0, dados[2])

    def calcular_valor_total(self, event=None):
        """Calcula o valor total baseado no tipo de despesa"""
        try:
            # Pegar valor unitário
            vr_unit_str = self.campos_despesa['vr_unit'].get().strip()
            if not vr_unit_str:
                self.campos_despesa['valor'].config(state='normal')
                self.campos_despesa['valor'].delete(0, tk.END)
                self.campos_despesa['valor'].config(state='readonly')
                return
                
            vr_unit = float(vr_unit_str.replace(',', '.'))
            
            # Pegar tipo de despesa
            tp_desp = self.campos_despesa['tp_desp'].get()
            
            # Calcular com base no tipo
            if tp_desp == '1':  # Tipo que usa dias
                dias_str = self.campos_despesa['dias'].get().strip()
                dias = int(dias_str) if dias_str else 1
                valor_total = vr_unit * dias
            else:
                valor_total = vr_unit
                
            # Atualizar campo de valor
            self.campos_despesa['valor'].config(state='normal')
            self.campos_despesa['valor'].delete(0, tk.END)
            self.campos_despesa['valor'].insert(0, f"{valor_total:.2f}")
            self.campos_despesa['valor'].config(state='readonly')
            
        except ValueError:
            # Em caso de erro, limpa o campo valor
            self.campos_despesa['valor'].config(state='normal')
            self.campos_despesa['valor'].delete(0, tk.END)
            self.campos_despesa['valor'].config(state='readonly')


    def verificar_tipo_despesa(self, event=None):
        """Verifica o tipo de despesa e ajusta campos conforme necessário"""
        tp_desp = self.campos_despesa['tp_desp'].get().strip()

        if not tp_desp.isdigit():
            self.campos_despesa['tp_desp'].delete(0, tk.END)
            return

        tp_desp_num = int(tp_desp)
        if not (1 <= tp_desp_num <= 6):
            self.campos_despesa['tp_desp'].delete(0, tk.END)
            return
        # Configura o campo dias
        if tp_desp == '1':
            self.campos_despesa['dias'].config(state='normal')
        else:
            self.campos_despesa['dias'].config(state='disabled')
            self.campos_despesa['dias'].delete(0, tk.END)
            self.campos_despesa['dias'].insert(0, '1')


        # Configura o campo nf
        if tp_desp != '1':
            self.campos_despesa['nf'].config(state='normal')
        else:
            self.campos_despesa['nf'].config(state='disabled')
            self.campos_despesa['nf'].delete(0, tk.END)
            


        # Atualiza o campo referência
        self.atualizar_campo_referencia(event)

        # Move para o campo referência
        self.campos_despesa['referencia'].focus()
        

    def adicionar_dados(self, eh_parcelamento=False):
        """Adiciona dados à lista temporária e retorna à aba fornecedor"""
        print("Iniciando adição de dados...")

        if not self.validar_campos():
            print("Falha na validação dos campos")
            return False # Importante retornar False aqui
        try:
            # Coleta do primeiro conjunto de dados
            vr_unit_str = self.campos_despesa['vr_unit'].get().strip()
            if not vr_unit_str:
                messagebox.showerror("Erro", "Valor unitário é obrigatório!")
                return
            vr_unit = float(vr_unit_str.replace(',', '.'))
        
            valor_str = self.campos_despesa['valor'].get().strip()
            if not valor_str:
                messagebox.showerror("Erro", "Valor total não foi calculado!")
                return
            valor = float(valor_str.replace(',', '.'))

            # Coletar dados do lançamento
            dados = {
                'data': self.data_rel_entry.get(),
                'cnpj_cpf': self.campos_fornecedor['cnpj_cpf'].get(),
                'nome': self.campos_fornecedor['nome'].get(),
                'categoria': self.campos_fornecedor['categoria'].get().upper(),
                'tp_desp': self.campos_despesa['tp_desp'].get(),
                'referencia': self.campos_despesa['referencia'].get().upper(),
                'nf': self.campos_despesa['nf'].get().upper(),  # Novo campo
                'vr_unit': f"{vr_unit:.2f}",
                'dias': int(self.campos_despesa['dias'].get() or 1),
                'valor': f"{valor:.2f}",
                'dt_vencto': self.campos_despesa['dt_vencto'].get(),
                'dados_bancarios': self.campos_fornecedor['dados_bancarios'].get(),
                'observacao': self.campos_despesa['observacao'].get().upper()
            }
            self.dados_para_incluir.append(dados)

            # Verificar se é um lançamento de TRANSPORTE e criar lançamento automático de CAFÉ
            if dados['tp_desp'] == '1' and dados['referencia'] == 'TRANSPORTE':
                # Calcular valores para o CAFÉ
                vr_unit_cafe = 4.0
                dias_cafe = int(dados['dias'])
                valor_cafe = vr_unit_cafe * dias_cafe
                # Criar dados do lançamento do CAFÉ
                dados_cafe = dados.copy()
                dados_cafe.update({
                    'referencia': 'CAFÉ',
                    'vr_unit': f"{vr_unit_cafe:.2f}",
                    'valor': f"{valor_cafe:.2f}"
                })
                self.dados_para_incluir.append(dados_cafe)
                messagebox.showinfo("Informação", "Lançamento de CAFÉ adicionado automaticamente!")

            # Só limpa os campos e mostra mensagem se não for parcelamento
            if not eh_parcelamento:
                self.limpar_campos_despesa()
                
                # Limpar campos do fornecedor
                for campo, entry in self.campos_fornecedor.items():
                    entry.config(state='normal')
                    entry.delete(0, tk.END)
                    if campo != 'categoria':
                        entry.config(state='readonly')
                
                messagebox.showinfo("Sucesso", "Dados adicionados com sucesso!")
                
                # Voltar para a aba fornecedor
                self.notebook.select(1)
                self.tree_fornecedores.selection_remove(self.tree_fornecedores.selection())
                self.busca_entry.delete(0, tk.END)
            
            return True  # Importante retornar True aqui
            
        except ValueError as e:
            messagebox.showerror("Erro", f"Erro ao processar valores: {str(e)}")
            return False  # Importante retornar False aqui

        
    def validar_campos(self):
        """Valida os campos antes de adicionar/enviar dados"""
        # Validar data
        if not self.data_rel_entry.get():
            messagebox.showerror("Erro", "Data de referência é obrigatória!")
            return False

        # Validar fornecedor
        if not self.campos_fornecedor['cnpj_cpf'].get():
            messagebox.showerror("Erro", "Selecione um fornecedor!")
            return False

        # Validar tipo de despesa
        tp_desp = self.campos_despesa['tp_desp'].get().strip()
        if not tp_desp or not tp_desp.isdigit() or not (1 <= int(tp_desp) <= 7):
            messagebox.showerror("Erro", "Tipo de despesa deve ser um número entre 1 e 7!")
            return False

        # Validar valor unitário
        vr_unit = self.campos_despesa['vr_unit'].get().strip()
        if not vr_unit:
            messagebox.showerror("Erro", "Valor unitário é obrigatório!")
            return False
        try:
            float(vr_unit.replace(',', '.'))
        except ValueError:
            messagebox.showerror("Erro", "Valor unitário inválido!")
            return False

        # Validar dias para tipo de despesa 1
        if tp_desp == '1':
            dias = self.campos_despesa['dias'].get().strip()
            if not dias:
                messagebox.showerror("Erro", "Quantidade de dias é obrigatória para tipo 1!")
                return False
            try:
                if int(dias) <= 0:
                    messagebox.showerror("Erro", "Quantidade de dias deve ser maior que zero!")
                    return False
            except ValueError:
                messagebox.showerror("Erro", "Quantidade de dias inválida!")
                return False

        # Validar referência
        if not self.campos_despesa['referencia'].get().strip():
            messagebox.showerror("Erro", "Referência é obrigatória!")
            return False

        # Validar data de vencimento
        if not self.campos_despesa['dt_vencto'].get():
            messagebox.showerror("Erro", "Data de vencimento é obrigatória!")
            return False

        return True

    


    def limpar_campos_despesa(self):
        """Limpa os campos de despesa mantendo alguns valores padrão"""
        # Limpar todos os campos
        self.campos_despesa['tp_desp'].delete(0, tk.END)
        self.campos_despesa['referencia'].set('')  # Para Combobox
        self.campos_despesa['nf'].delete(0, tk.END)  # Novo campo
        self.campos_despesa['vr_unit'].delete(0, tk.END)
        self.campos_despesa['dias'].delete(0, tk.END)
        self.campos_despesa['dias'].insert(0, '1')  # Valor padrão
        self.campos_despesa['valor'].config(state='normal')
        self.campos_despesa['valor'].delete(0, tk.END)
        self.campos_despesa['valor'].config(state='readonly')
        self.campos_despesa['observacao'].delete(0, tk.END)
        
        # Definir data de vencimento igual à data de referência
        self.campos_despesa['dt_vencto'].set_date(self.data_rel_entry.get_date())

        # Resetar estado do campo referência
        self.campos_despesa['referencia'].config(state='normal')
        self.campos_despesa['referencia']['values'] = []

    def enviar_dados(self):
        """Salva os dados na planilha existente do cliente"""
        if not self.cliente_atual:
            messagebox.showerror("Erro", "Selecione um cliente!")
            return
        
        try:
            # Verificar se temos dados para processar
            dados_para_processar = []
            if hasattr(self, 'visualizador') and self.visualizador and self.visualizador.tree.winfo_exists():
                dados_para_processar = self.visualizador.get_dados_atualizados()
            elif self.dados_para_incluir:
                dados_para_processar = self.dados_para_incluir.copy()
                
            if not dados_para_processar:
                messagebox.showwarning("Aviso", "Não há dados para enviar!")
                return

            # Atualizar lista principal com dados mais recentes
            self.dados_para_incluir = dados_para_processar

            arquivo_cliente = PASTA_CLIENTES / f"{self.cliente_atual}.xlsx"
            
            try:
                workbook = load_workbook(arquivo_cliente)
            except PermissionError:
                messagebox.showerror(
                    "Erro", 
                    f"A planilha '{self.cliente_atual}.xlsx' está aberta!\n\n"
                    "Por favor:\n"
                    "1. Feche a planilha\n"
                    "2. Clique em OK\n"
                    "3. Tente enviar novamente"
                )
                return
            
            sheet = workbook["Dados"]

            if sheet.tables:
                table_name = list(sheet.tables.keys())[0]
                sheet.tables.pop(table_name)
                
            # Processar registros
            for dados in dados_para_processar:
                proxima_linha = sheet.max_row + 1
                
                # Converter e salvar data de referência
                data_rel = datetime.strptime(dados['data'], '%d/%m/%Y')
                data_cell = sheet.cell(row=proxima_linha, column=1, value=data_rel)
                data_cell.number_format = 'DD/MM/YYYY'

                # Converter tipo de despesa para número
                tp_desp_cell = sheet.cell(row=proxima_linha, column=2, value=int(dados['tp_desp']))
                tp_desp_cell.number_format = '0'

                sheet.cell(row=proxima_linha, column=3, value=dados['cnpj_cpf'])
                sheet.cell(row=proxima_linha, column=4, value=dados['nome'])
                sheet.cell(row=proxima_linha, column=5, value=dados['referencia'])
                sheet.cell(row=proxima_linha, column=6, value=dados['nf'])

                vr_unit_cell = sheet.cell(row=proxima_linha, column=7, 
                                        value=float(str(dados['vr_unit']).replace(',', '.')))
                vr_unit_cell.number_format = '#,##0.00'

                sheet.cell(row=proxima_linha, column=8, value=dados['dias'])

                valor_cell = sheet.cell(row=proxima_linha, column=9, 
                                      value=float(str(dados['valor']).replace(',', '.')))
                valor_cell.number_format = '#,##0.00'

                dt_vencto = datetime.strptime(dados['dt_vencto'], '%d/%m/%Y')
                dt_vencto_cell = sheet.cell(row=proxima_linha, column=10, value=dt_vencto)
                dt_vencto_cell.number_format = 'DD/MM/YYYY'

                sheet.cell(row=proxima_linha, column=11, value=dados['categoria'])
                sheet.cell(row=proxima_linha, column=12, value=dados['dados_bancarios'])
                sheet.cell(row=proxima_linha, column=13, value=dados['observacao'])

            try:
                # Tentar salvar o arquivo
                workbook.save(arquivo_cliente)
                messagebox.showinfo("Sucesso", "Dados salvos com sucesso!")
                    
                # Limpar após salvar
                self.dados_para_incluir.clear()
                if hasattr(self, 'visualizador') and self.visualizador:
                    self.visualizador.janela.destroy()
                    self.visualizador = None
                    
                # Criar uma janela de diálogo personalizada
                dialog = tk.Toplevel(self.root)
                dialog.title("Continuar")
                dialog.geometry("300x250")
                dialog.transient(self.root)
                dialog.grab_set()
                
                # Centralizar a janela
                dialog.update_idletasks()
                width = dialog.winfo_width()
                height = dialog.winfo_height()
                x = (dialog.winfo_screenwidth() // 2) - (width // 2)
                y = (dialog.winfo_screenheight() // 2) - (height // 2)
                dialog.geometry(f'{width}x{height}+{x}+{y}')
                
                ttk.Label(dialog, 
                         text="O que você deseja fazer?",
                         font=('Helvetica', 10, 'bold')).pack(pady=10)
                
                def continuar_entrada():
                    dialog.destroy()
                    self.limpar_campos_despesa()
                    self.notebook.select(1)  # Volta para aba de fornecedor
                    
                def voltar_menu_local():  # Renomeada para evitar conflito
                    dialog.destroy()
                    self.root.destroy()
                    if hasattr(self, 'menu_principal'):
                        self.menu_principal.deiconify()
                    
                def sair_sistema():
                    dialog.destroy()
                    self.root.destroy()
                    sys.exit()

                # Frame para os botões
                btn_frame = ttk.Frame(dialog)
                btn_frame.pack(fill='x', pady=10)
                
                # Botões com a função local correta
                ttk.Button(btn_frame, 
                          text="Continuar Entrada de Dados", 
                          command=continuar_entrada).pack(pady=5, padx=10, fill='x')
                
                ttk.Button(btn_frame, 
                          text="Voltar ao Menu Principal", 
                          command=voltar_menu_local).pack(pady=5, padx=10, fill='x')
                
                ttk.Button(btn_frame, 
                          text="Sair do Sistema", 
                          command=sair_sistema).pack(pady=5, padx=10, fill='x')

            except PermissionError:
                messagebox.showerror(
                    "Erro", 
                    f"Não foi possível salvar! A planilha '{self.cliente_atual}.xlsx' está aberta.\n\n"
                    "Por favor:\n"
                    "1. Feche a planilha\n"
                    "2. Clique em OK\n"
                    "3. Tente enviar novamente"
                )
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao salvar arquivo: {str(e)}")
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar dados: {str(e)}")

class EditorCliente:
    def __init__(self, parent):
        self.parent = parent
        self.root = tk.Toplevel(parent)
        self.root.title("Editor de Clientes")
        self.root.geometry("800x600")
        
        self.setup_gui()
        self.carregar_clientes()

    def setup_gui(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill='both', expand=True)

        # Lista de clientes
        frame_clientes = ttk.LabelFrame(main_frame, text="Clientes")
        frame_clientes.pack(fill='both', expand=True, pady=5)

        self.tree_clientes = ttk.Treeview(frame_clientes, 
                                        columns=('Nome', 'Endereço', 'Taxa ADM'),
                                        show='headings')
        self.tree_clientes.heading('Nome', text='Nome')
        self.tree_clientes.heading('Endereço', text='Endereço')
        self.tree_clientes.heading('Taxa ADM', text='Taxa ADM (%)')
        self.tree_clientes.pack(fill='both', expand=True, padx=5, pady=5)

        # Frame para edição
        frame_edicao = ttk.LabelFrame(main_frame, text="Edição")
        frame_edicao.pack(fill='x', pady=5)

        ttk.Label(frame_edicao, text="Taxa de Administração (%):").pack(side='left', padx=5)
        self.taxa_entry = ttk.Entry(frame_edicao, width=10)
        self.taxa_entry.pack(side='left', padx=5)

        # Botões
        frame_botoes = ttk.Frame(main_frame)
        frame_botoes.pack(fill='x', pady=5)

        ttk.Button(frame_botoes, 
                  text="Atualizar Taxa", 
                  command=self.atualizar_taxa).pack(side='left', padx=5)
        ttk.Button(frame_botoes, 
                  text="Remover Taxa", 
                  command=self.remover_taxa).pack(side='left', padx=5)
        ttk.Button(frame_botoes, 
                  text="Fechar", 
                  command=self.root.destroy).pack(side='right', padx=5)

    def carregar_clientes(self):
        """Carrega a lista de clientes do arquivo Excel"""
        try:
            wb = load_workbook(ARQUIVO_CLIENTES)
            ws = wb['Clientes']
            
            # Limpar lista atual
            for item in self.tree_clientes.get_children():
                self.tree_clientes.delete(item)
            
            # Adicionar clientes
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Nome não vazio
                    self.tree_clientes.insert('', 'end', values=(
                        row[0],  # Nome
                        row[1],  # Endereço
                        row[6] if row[6] else "0.00"  # Taxa ADM
                    ))
            
            wb.close()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar clientes: {str(e)}")

    def atualizar_taxa(self):
        """Atualiza a taxa de administração do cliente selecionado"""
        selecionado = self.tree_clientes.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um cliente")
            return

        try:
            taxa = float(self.taxa_entry.get().replace(',', '.'))
            if not (0 <= taxa <= 100):
                messagebox.showerror("Erro", "Taxa deve estar entre 0 e 100")
                return
                
            cliente = self.tree_clientes.item(selecionado)['values'][0]
            
            # Atualizar no arquivo
            wb = load_workbook(ARQUIVO_CLIENTES)
            ws = wb['Clientes']
            
            for row in ws.iter_rows(min_row=2):
                if row[0].value == cliente:
                    row[6].value = taxa  # Coluna da taxa de administração
                    
            wb.save(ARQUIVO_CLIENTES)
            
            # Atualizar na treeview
            self.tree_clientes.set(selecionado, 'Taxa ADM', f"{taxa:.2f}")
            messagebox.showinfo("Sucesso", "Taxa atualizada com sucesso!")
            
        except ValueError:
            messagebox.showerror("Erro", "Taxa inválida")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao atualizar taxa: {str(e)}")

    def remover_taxa(self):
        """Remove a taxa de administração do cliente selecionado"""
        selecionado = self.tree_clientes.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um cliente")
            return

        if messagebox.askyesno("Confirmar", "Deseja remover a taxa de administração?"):
            try:
                cliente = self.tree_clientes.item(selecionado)['values'][0]
                
                # Atualizar no arquivo
                wb = load_workbook(ARQUIVO_CLIENTES)
                ws = wb['Clientes']
                
                for row in ws.iter_rows(min_row=2):
                    if row[0].value == cliente:
                        row[6].value = None  # Remover taxa
                        
                wb.save(ARQUIVO_CLIENTES)
                
                # Atualizar na treeview
                self.tree_clientes.set(selecionado, 'Taxa ADM', "0.00")
                messagebox.showinfo("Sucesso", "Taxa removida com sucesso!")
                
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao remover taxa: {str(e)}")



class GestaoContratos:
    def __init__(self, parent):
        self.parent = parent
        self.arquivo_cliente = None
        self.cliente_atual = None

    def abrir_janela_contrato(self, cliente):
        """Abre janela para gestão de contratos"""
        self.cliente_atual = cliente
        self.arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
        
        try:
            # Verificar se o arquivo existe
            if not os.path.exists(self.arquivo_cliente):
                messagebox.showerror("Erro", f"Arquivo do cliente {cliente} não encontrado!")
                return

            # Abrir arquivo e verificar aba
            wb = load_workbook(self.arquivo_cliente)
            if 'Contratos_ADM' not in wb.sheetnames:
                # Se não existir a aba, criar
                print(f"Criando aba Contratos_ADM para {cliente}")
                ws = wb.create_sheet("Contratos_ADM")
                
                # Definir os blocos na linha 1
                blocos = ["CONTRATOS", "", "", "", "", "",
                         "ADMINISTRADORES_CONTRATO", "", "", "", "", "", "",
                         "ADITIVOS", "", "", "",
                         "ADMINISTRADORES_ADITIVO", "", "", "", "", "", "",
                         "PARCELAS", "", "", "", ""]
                
                for col, valor in enumerate(blocos, 1):
                    ws.cell(row=1, column=col, value=valor)
                
                # Definir cabeçalhos na linha 2
                headers = [
                    # CONTRATOS
                    "Nº Contrato", "Data Início", "Data Fim", "Status", "Observações", "",
                    # ADMINISTRADORES_CONTRATO
                    "Nº Contrato", "CNPJ/CPF", "Nome/Razão Social", "Tipo", "Valor/Percentual", "Valor Total", "Nº Parcelas", 
                    # ADITIVOS
                    "Nº Contrato", "Nº Aditivo", "Data Início", "Data Fim",
                    # ADMINISTRADORES_ADITIVO
                    "Nº Contrato", "Nº Aditivo", "CNPJ/CPF", "Nome/Razão Social", "Tipo", "Valor/Percentual", "Valor Total",
                    # PARCELAS
                    "Referência", "Número", "CNPJ/CPF", "Nome", "Data Vencimento", "Valor", "Status", "Data Pagamento"
                ]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=2, column=col, value=header)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                
                # Ajustar largura das colunas
                for col in range(1, len(headers) + 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
                
                # Salvar as alterações
                wb.save(self.arquivo_cliente)

            # Criar a janela principal
            janela = tk.Toplevel(self.parent)
            janela.title(f"Gestão de Contratos - {cliente}")
            janela.geometry("800x650")

            # Frame principal
            frame_principal = ttk.Frame(janela, padding="10")
            frame_principal.pack(fill='both', expand=True)

            # Frame para lista de contratos existentes
            frame_contratos = ttk.LabelFrame(frame_principal, text="Contratos Existentes")
            frame_contratos.pack(fill='both', expand=True, pady=5)

            # Treeview para contratos
            colunas = ('Nº Contrato', 'Data Início', 'Data Fim', 'Status')
            self.tree_contratos = ttk.Treeview(frame_contratos, columns=colunas, show='headings')
            for col in colunas:
                self.tree_contratos.heading(col, text=col)
                self.tree_contratos.column(col, width=100)
            
            # Adicionar scrollbars
            scroll_y = ttk.Scrollbar(frame_contratos, orient='vertical', command=self.tree_contratos.yview)
            scroll_x = ttk.Scrollbar(frame_contratos, orient='horizontal', command=self.tree_contratos.xview)
            self.tree_contratos.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
            
            self.tree_contratos.pack(fill='both', expand=True, padx=5, pady=5)
            scroll_y.pack(side='right', fill='y')
            scroll_x.pack(side='bottom', fill='x')

            # Frame para lista de administradores do contrato selecionado
            frame_admins = ttk.LabelFrame(frame_principal, text="Administradores do Contrato")
            frame_admins.pack(fill='both', expand=True, pady=5)

            # Treeview para administradores
            colunas_adm = ('CNPJ/CPF', 'Nome', 'Tipo', 'Valor/Percentual', 'Valor Total', 'Nº Parcelas', 'Data Inicial Pagamento')
            self.tree_adm_contrato = ttk.Treeview(frame_admins, columns=colunas_adm, show='headings')
            for col in colunas_adm:
                self.tree_adm_contrato.heading(col, text=col)
                self.tree_adm_contrato.column(col, width=100)
            
            # Adicionar scrollbars para administradores
            scroll_y_adm = ttk.Scrollbar(frame_admins, orient='vertical', command=self.tree_adm_contrato.yview)
            scroll_x_adm = ttk.Scrollbar(frame_admins, orient='horizontal', command=self.tree_adm_contrato.xview)
            self.tree_adm_contrato.configure(yscrollcommand=scroll_y_adm.set, xscrollcommand=scroll_x_adm.set)
            
            self.tree_adm_contrato.pack(fill='both', expand=True, padx=5, pady=5)
            scroll_y_adm.pack(side='right', fill='y')
            scroll_x_adm.pack(side='bottom', fill='x')

            # Botões de ação
            frame_botoes = ttk.Frame(frame_principal)
            frame_botoes.pack(fill='x', pady=5)

            ttk.Button(frame_botoes, text="Novo Contrato", 
                      command=lambda: self.criar_novo_contrato(janela)).pack(side='left', padx=5)
            ttk.Button(frame_botoes, text="Editar Contrato", 
                      command=self.editar_contrato).pack(side='left', padx=5)
            ttk.Button(frame_botoes, text="Excluir Contrato", 
                      command=self.excluir_contrato).pack(side='left', padx=5)
            ttk.Button(frame_botoes, text="Fechar", 
                      command=janela.destroy).pack(side='right', padx=5)

            # Carregar contratos existentes
            self.carregar_contratos()

            # Binding para atualizar administradores quando selecionar contrato
            self.tree_contratos.bind('<<TreeviewSelect>>', self.mostrar_administradores)

            # Centralizar a janela
            janela.update_idletasks()
            width = janela.winfo_width()
            height = janela.winfo_height()
            x = (janela.winfo_screenwidth() // 2) - (width // 2)
            y = (janela.winfo_screenheight() // 2) - (height // 2)
            janela.geometry(f'{width}x{height}+{x}+{y}')

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir janela de contratos: {str(e)}")
            if 'wb' in locals():
                wb.close()


    def carregar_contratos(self):
        try:
            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Contratos_ADM']
            
            for item in self.tree_contratos.get_children():
                self.tree_contratos.delete(item)
            
            contratos_processados = set()
            for row in ws.iter_rows(min_row=3, values_only=True):
                num_contrato = row[0]
                if num_contrato and num_contrato not in contratos_processados:
                    # Processar datas
                    data_inicio = ''
                    if row[1]:
                        try:
                            if isinstance(row[1], datetime):
                                data_inicio = row[1].strftime('%d/%m/%Y')
                            else:
                                data_temp = datetime.strptime(str(row[1]), '%Y-%m-%d')
                                data_inicio = data_temp.strftime('%d/%m/%Y')
                        except ValueError:
                            data_inicio = str(row[1])

                    data_fim = ''
                    if row[2]:
                        try:
                            if isinstance(row[2], datetime):
                                data_fim = row[2].strftime('%d/%m/%Y')
                            else:
                                data_temp = datetime.strptime(str(row[2]), '%Y-%m-%d')
                                data_fim = data_temp.strftime('%d/%m/%Y')
                        except ValueError:
                            data_fim = str(row[2])

                    
                    self.tree_contratos.insert('', 'end', values=(
                        num_contrato,
                        data_inicio,
                        data_fim,
                        row[3] or ''
                    ))
                    contratos_processados.add(num_contrato)
            
            wb.close()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar contratos: {str(e)}")

    def mostrar_administradores(self, event=None):
        """Mostra administradores do contrato selecionado"""
        selecionado = self.tree_contratos.selection()
        if not selecionado:
            return
            
        try:
            # Limpar lista atual
            for item in self.tree_adm_contrato.get_children():
                self.tree_adm_contrato.delete(item)
                
            num_contrato = self.tree_contratos.item(selecionado)['values'][0]
            
            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Contratos_ADM']
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[6] == num_contrato:  # Coluna G - Nº Contrato
                    if row[26]:  # Data Inicial de Pagamento
                        data_inicial = row[26].strftime('%d/%m/%Y') if isinstance(row[26], datetime) else str(row[26])
                    else:
                        data_inicial = ''
                        
                    self.tree_adm_contrato.insert('', 'end', values=(
                        row[7],   # CNPJ/CPF
                        row[8],   # Nome
                        row[9],   # Tipo
                        row[10],  # Valor/Percentual
                        row[11],  # Valor Total
                        row[12],  # Nº Parcelas
                        data_inicial  # Data Inicial de Pagamento
                    ))
            
            wb.close()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar administradores: {str(e)}")
        

    def criar_novo_contrato(self, janela_principal):
        """Abre janela para criar novo contrato"""
        janela = tk.Toplevel(self.parent)
        janela.title(f"Novo Contrato - {self.cliente_atual}")
        janela.geometry("600x500")

        # Frame principal
        frame = ttk.Frame(janela, padding="10")
        frame.pack(fill='both', expand=True)

        # Frame para dados do contrato
        frame_contrato = ttk.LabelFrame(frame, text="Dados do Contrato")
        frame_contrato.pack(fill='x', pady=5)

        # Número do Contrato
        ttk.Label(frame_contrato, text="Nº Contrato:*").grid(row=0, column=0, padx=5, pady=2)
        num_contrato = ttk.Entry(frame_contrato)
        num_contrato.grid(row=0, column=1, padx=5, pady=2)

        # Datas
        ttk.Label(frame_contrato, text="Data Início:*").grid(row=1, column=0, padx=5, pady=2)
        data_inicio = DateEntry(frame_contrato, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
        data_inicio.grid(row=1, column=1, padx=5, pady=2)

        ttk.Label(frame_contrato, text="Data Fim:*").grid(row=2, column=0, padx=5, pady=2)
        data_fim = DateEntry(frame_contrato, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
        data_fim.grid(row=2, column=1, padx=5, pady=2)

        # Observações
        ttk.Label(frame_contrato, text="Observações:").grid(row=3, column=0, padx=5, pady=2)
        observacoes = ttk.Entry(frame_contrato, width=50)
        observacoes.grid(row=3, column=1, padx=5, pady=2)

        # Frame para Administradores
        frame_adm = ttk.LabelFrame(frame, text="Administradores")
        frame_adm.pack(fill='both', expand=True, pady=5)

        # Lista de Administradores
        colunas = ('CNPJ/CPF', 'Nome', 'Tipo', 'Valor/Percentual', 'Valor Total', 'Nº Parcelas', 'Data Inicial')
        self.tree_adm = ttk.Treeview(frame_adm, columns=colunas, show='headings', height=5)
        
        for col in colunas:
            self.tree_adm.heading(col, text=col)
            self.tree_adm.column(col, width=100)
        
        # Adicionar scrollbars
        scroll_y = ttk.Scrollbar(frame_adm, orient='vertical', command=self.tree_adm.yview)
        scroll_x = ttk.Scrollbar(frame_adm, orient='horizontal', command=self.tree_adm.xview)
        self.tree_adm.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        
        self.tree_adm.pack(fill='both', expand=True, padx=5, pady=5)
        scroll_y.pack(side='right', fill='y')
        scroll_x.pack(side='bottom', fill='x')

        # Botões para administradores
        frame_botoes_adm = ttk.Frame(frame_adm)
        frame_botoes_adm.pack(fill='x', pady=5)

        ttk.Button(frame_botoes_adm, 
                  text="Adicionar Administrador",
                  command=lambda: self.adicionar_administrador(self.tree_adm)).pack(side='left', padx=5)
        
        ttk.Button(frame_botoes_adm, 
                  text="Remover Administrador",
                  command=lambda: self.remover_administrador(self.tree_adm)).pack(side='left', padx=5)

        # Botões principais
        frame_botoes = ttk.Frame(frame)
        frame_botoes.pack(fill='x', pady=10)

        def salvar():
            self.salvar_contrato(
                num_contrato.get(),
                data_inicio.get_date(),
                data_fim.get_date(),
                observacoes.get(),
                janela
            )
            janela_principal.focus_set()
            self.carregar_contratos()

        ttk.Button(frame_botoes, text="Salvar", command=salvar).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Cancelar", command=janela.destroy).pack(side='left', padx=5)


    def editar_contrato(self):
        """Edita o contrato selecionado"""
        selecionado = self.tree_contratos.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um contrato para editar")
            return

        try:
            dados_contrato = self.tree_contratos.item(selecionado)['values']
            
            janela = tk.Toplevel(self.parent)
            janela.title(f"Editar Contrato - {self.cliente_atual}")
            janela.geometry("600x500")

            # Frame principal
            frame = ttk.Frame(janela, padding="10")
            frame.pack(fill='both', expand=True)

            # Dados do Contrato
            frame_contrato = ttk.LabelFrame(frame, text="Dados do Contrato")
            frame_contrato.pack(fill='x', pady=5)

            # Número do Contrato (readonly)
            ttk.Label(frame_contrato, text="Nº Contrato:").grid(row=0, column=0, padx=5, pady=2)
            num_contrato = ttk.Entry(frame_contrato, state='readonly')
            num_contrato.grid(row=0, column=1, padx=5, pady=2)
            num_contrato.insert(0, dados_contrato[0])

            # Datas
            ttk.Label(frame_contrato, text="Data Início:").grid(row=1, column=0, padx=5, pady=2)
            data_inicio = DateEntry(frame_contrato, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
            data_inicio.grid(row=1, column=1, padx=5, pady=2)
            data_inicio.set_date(datetime.strptime(dados_contrato[1], '%d/%m/%Y'))

            ttk.Label(frame_contrato, text="Data Fim:").grid(row=2, column=0, padx=5, pady=2)
            data_fim = DateEntry(frame_contrato, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
            data_fim.grid(row=2, column=1, padx=5, pady=2)
            data_fim.set_date(datetime.strptime(dados_contrato[2], '%d/%m/%Y'))

            # Status
            ttk.Label(frame_contrato, text="Status:").grid(row=3, column=0, padx=5, pady=2)
            status_combo = ttk.Combobox(frame_contrato, values=['ATIVO', 'INATIVO'], state='readonly')
            status_combo.grid(row=3, column=1, padx=5, pady=2)
            status_combo.set(dados_contrato[3])

            def salvar_alteracoes():
                try:
                    wb = load_workbook(self.arquivo_cliente)
                    ws = wb['Contratos_ADM']
                    
                    # Atualizar dados do contrato
                    for row in ws.iter_rows(min_row=2):
                        if row[0].value == dados_contrato[0]:
                            row[1].value = data_inicio.get_date()
                            row[2].value = data_fim.get_date()
                            row[3].value = status_combo.get()
                    
                    wb.save(self.arquivo_cliente)
                    messagebox.showinfo("Sucesso", "Contrato atualizado com sucesso!")
                    janela.destroy()
                    self.carregar_contratos()
                    
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao salvar alterações: {str(e)}")

            # Botões
            frame_botoes = ttk.Frame(frame)
            frame_botoes.pack(fill='x', pady=10)

            ttk.Button(frame_botoes, text="Salvar", command=salvar_alteracoes).pack(side='left', padx=5)
            ttk.Button(frame_botoes, text="Cancelar", command=janela.destroy).pack(side='left', padx=5)

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir edição: {str(e)}")



    def adicionar_administrador(self, tree):
        """Abre janela para adicionar novo administrador"""
        janela = tk.Toplevel(self.parent)
        janela.title("Adicionar Administrador")
        janela.geometry("600x650")

        frame = ttk.Frame(janela, padding="10")
        frame.pack(fill='both', expand=True)

        # Frame de busca
        frame_busca = ttk.LabelFrame(frame, text="Buscar Fornecedor")
        frame_busca.pack(fill='x', padx=5, pady=5)

        ttk.Label(frame_busca, text="Nome:").pack(side='left', padx=5)
        busca_entry = ttk.Entry(frame_busca, width=40)
        busca_entry.pack(side='left', padx=5)

        # Lista de fornecedores
        frame_lista = ttk.LabelFrame(frame, text="Fornecedores")
        frame_lista.pack(fill='x', padx=5, pady=5)

        tree_fornecedores = ttk.Treeview(frame_lista, 
                                        columns=('CNPJ/CPF', 'Nome', 'Categoria'),
                                        show='headings',
                                        height=3)
        tree_fornecedores.heading('CNPJ/CPF', text='CNPJ/CPF')
        tree_fornecedores.heading('Nome', text='Nome')
        tree_fornecedores.heading('Categoria', text='Categoria')
        tree_fornecedores.pack(fill='both', expand=True, padx=5, pady=5)

        # Frame para dados do administrador
        frame_dados = ttk.LabelFrame(frame, text="Dados do Administrador")
        frame_dados.pack(fill='x', padx=5, pady=5)

        # CNPJ/CPF
        ttk.Label(frame_dados, text="CNPJ/CPF:*").grid(row=0, column=0, padx=5, pady=2)
        cnpj_cpf_entry = ttk.Entry(frame_dados, state='readonly')
        cnpj_cpf_entry.grid(row=0, column=1, padx=5, pady=2)

        # Nome
        ttk.Label(frame_dados, text="Nome/Razão Social:*").grid(row=1, column=0, padx=5, pady=2)
        nome_entry = ttk.Entry(frame_dados, state='readonly')
        nome_entry.grid(row=1, column=1, padx=5, pady=2)

        # Tipo
        ttk.Label(frame_dados, text="Tipo:*").grid(row=2, column=0, padx=5, pady=2)
        tipo_combo = ttk.Combobox(frame_dados, values=['Percentual', 'Fixo'], state='readonly')
        tipo_combo.grid(row=2, column=1, padx=5, pady=2)

        # Frame para valores fixos
        frame_fixo = ttk.Frame(frame_dados)
        frame_fixo.grid(row=4, column=0, columnspan=2, pady=5)

        # Valor Total
        ttk.Label(frame_fixo, text="Valor Total do Contrato:*").grid(row=0, column=0, padx=5, pady=2)
        valor_total_entry = ttk.Entry(frame_fixo)
        valor_total_entry.grid(row=0, column=1, padx=5, pady=2)

        # Checkbox para entrada
        tem_entrada = tk.BooleanVar()
        check_entrada = ttk.Checkbutton(frame_fixo, text="Possui Entrada?", variable=tem_entrada,
                                       command=lambda: atualizar_campos_entrada())
        check_entrada.grid(row=1, column=0, columnspan=2, padx=5, pady=2)

        # Frame para entrada
        frame_entrada = ttk.Frame(frame_fixo)
        frame_entrada.grid(row=2, column=0, columnspan=2, pady=5)

        ttk.Label(frame_entrada, text="Valor da Entrada:*").grid(row=0, column=0, padx=5, pady=2)
        valor_entrada_entry = ttk.Entry(frame_entrada)
        valor_entrada_entry.grid(row=0, column=1, padx=5, pady=2)
        valor_entrada_entry.insert(0, "0.00")

        ttk.Label(frame_entrada, text="Data da Entrada:*").grid(row=1, column=0, padx=5, pady=2)
        data_entrada = DateEntry(frame_entrada, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
        data_entrada.grid(row=1, column=1, padx=5, pady=2)

        # Número de Parcelas
        ttk.Label(frame_fixo, text="Nº Parcelas:*").grid(row=3, column=0, padx=5, pady=2)
        parcelas_entry = ttk.Entry(frame_fixo)
        parcelas_entry.grid(row=3, column=1, padx=5, pady=2)

        # Data Inicial
        ttk.Label(frame_fixo, text="Data Inicial Pagamento:*").grid(row=4, column=0, padx=5, pady=2)
        data_inicial_pagto = DateEntry(frame_fixo, width=20, date_pattern='dd/mm/yyyy', locale='pt_BR')
        data_inicial_pagto.grid(row=4, column=1, padx=5, pady=2)

        def atualizar_campos_entrada():
            """Atualiza visibilidade dos campos de entrada"""
            if tem_entrada.get():
                frame_entrada.grid()
            else:
                frame_entrada.grid_remove()
                valor_entrada_entry.delete(0, tk.END)
                valor_entrada_entry.insert(0, "0.00")

        def calcular_valor_parcela(*args):
            """Calcula o valor das parcelas considerando entrada"""
            try:
                valor_total = float(valor_total_entry.get().replace(',', '.'))
                valor_entrada = float(valor_entrada_entry.get().replace(',', '.') or "0")
                num_parcelas = int(parcelas_entry.get() or "0")

                if valor_entrada >= valor_total:
                    messagebox.showerror("Erro", "Valor da entrada não pode ser maior ou igual ao valor total!")
                    valor_entrada_entry.delete(0, tk.END)
                    valor_entrada_entry.insert(0, "0.00")
                    return

                valor_restante = valor_total - valor_entrada
                if num_parcelas > 0:
                    valor_parcela = valor_restante / num_parcelas
                    return valor_parcela
                return 0

            except (ValueError, ZeroDivisionError):
                return 0

        def confirmar():
            """Confirma a adição do administrador"""
            try:
                if not cnpj_cpf_entry.get() or not nome_entry.get() or not tipo_combo.get():
                    messagebox.showerror("Erro", "Preencha todos os campos obrigatórios!")
                    return

                if tipo_combo.get() == 'Fixo':
                    if not valor_total_entry.get() or not parcelas_entry.get():
                        messagebox.showerror("Erro", "Preencha valor total e número de parcelas!")
                        return

                    valor_entrada = float(valor_entrada_entry.get().replace(',', '.'))
                    if tem_entrada.get() and valor_entrada > 0:
                        # Adicionar registro de entrada
                        valores_entrada = (
                            cnpj_cpf_entry.get(),
                            nome_entry.get(),
                            tipo_combo.get(),
                            f"{valor_entrada:.2f}",
                            valor_total_entry.get(),
                            "1",  # Uma parcela para entrada
                            data_entrada.get()
                        )
                        tree.insert('', 'end', values=valores_entrada, tags=('entrada',))

                    # Calcular valor da parcela
                    valor_parcela = calcular_valor_parcela()
                    if valor_parcela > 0:
                        # Adicionar registro de parcelas
                        valores_parcelas = (
                            cnpj_cpf_entry.get(),
                            nome_entry.get(),
                            tipo_combo.get(),
                            f"{valor_parcela:.2f}",
                            valor_total_entry.get(),
                            parcelas_entry.get(),
                            data_inicial_pagto.get()
                        )
                        tree.insert('', 'end', values=valores_parcelas, tags=('parcelas',))

                janela.destroy()

            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao confirmar: {str(e)}")

        def atualizar_campos_fixo(*args):
            """Atualiza campos baseado no tipo selecionado"""
            if tipo_combo.get() == 'Fixo':
                frame_fixo.grid()  # Mostra frame de valores fixos
                
                # Se tem entrada, mostra frame de entrada também
                if tem_entrada.get():
                    frame_entrada.grid()
                else:
                    frame_entrada.grid_remove()
            else:
                frame_fixo.grid_remove()  # Esconde todos os campos de valor fixo
                frame_entrada.grid_remove()  # Esconde campos de entrada

        # Configurar eventos
        valor_total_entry.bind('<KeyRelease>', lambda e: calcular_valor_parcela())
        parcelas_entry.bind('<KeyRelease>', lambda e: calcular_valor_parcela())
        valor_entrada_entry.bind('<KeyRelease>', lambda e: calcular_valor_parcela())
        tipo_combo.bind('<<ComboboxSelected>>', atualizar_campos_fixo)

        # Esconder campos de entrada inicialmente
        frame_entrada.grid_remove()

        # Botões
        frame_botoes = ttk.Frame(frame)
        frame_botoes.pack(fill='x', pady=10)
        ttk.Button(frame_botoes, text="Confirmar", command=confirmar).pack(side='left', padx=5)
        ttk.Button(frame_botoes, text="Cancelar", command=janela.destroy).pack(side='left', padx=5)

        # Função de busca
        def busca_local():
            termo = busca_entry.get()
            buscar_fornecedor(tree_fornecedores, termo)

        ttk.Button(frame_busca, text="Buscar", command=busca_local).pack(side='left', padx=5)
        busca_entry.bind('<Return>', lambda e: busca_local())

        def selecionar_e_preencher(event=None):
            selecionado = tree_fornecedores.selection()
            if not selecionado:
                return

            valores = tree_fornecedores.item(selecionado)['values']
            cnpj_cpf_entry.config(state='normal')
            nome_entry.config(state='normal')
            
            cnpj_cpf_entry.delete(0, tk.END)
            cnpj_cpf_entry.insert(0, str(valores[0]).zfill(14))
            
            nome_entry.delete(0, tk.END)
            nome_entry.insert(0, valores[1])
            
            cnpj_cpf_entry.config(state='readonly')
            nome_entry.config(state='readonly')

        tree_fornecedores.bind('<Double-1>', selecionar_e_preencher)          

        


    def salvar_contrato(self, num_contrato, data_inicio, data_fim, observacoes, janela):
        """Salva os dados do contrato e seus administradores"""
        if not num_contrato or not data_inicio or not data_fim:
            messagebox.showerror("Erro", "Preencha todos os campos obrigatórios!")
            return

        try:
            if not self.tree_adm.get_children():
                messagebox.showerror("Erro", "Adicione pelo menos um administrador!")
                return

            wb = load_workbook(self.arquivo_cliente)
            ws = wb['Contratos_ADM']
            ws_dados = wb['Dados']

            # Verificar se o contrato já existe
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == num_contrato:
                    messagebox.showerror("Erro", "Número de contrato já existe!")
                    return

            # Salvar dados do contrato
            proxima_linha = ws.max_row + 1
            ws.cell(row=proxima_linha, column=1, value=num_contrato)
            ws.cell(row=proxima_linha, column=2, value=data_inicio)
            ws.cell(row=proxima_linha, column=3, value=data_fim)
            ws.cell(row=proxima_linha, column=4, value='ATIVO')
            ws.cell(row=proxima_linha, column=5, value=observacoes)

            # Processar administradores
            for item in self.tree_adm.get_children():
                valores = self.tree_adm.item(item)['values']
                tags = self.tree_adm.item(item)['tags']
                
                # Formatação do CNPJ/CPF
                cnpj_cpf = str(valores[0]).strip()
                cnpj_cpf = formatar_cnpj_cpf(cnpj_cpf)
                nome_admin = valores[1]

                # Registrar administrador no contrato apenas uma vez
                proxima_linha = ws.max_row + 1
                ws.cell(row=proxima_linha, column=7, value=num_contrato)
                ws.cell(row=proxima_linha, column=8, value=cnpj_cpf)
                ws.cell(row=proxima_linha, column=9, value=nome_admin)
                ws.cell(row=proxima_linha, column=10, value=valores[2])  # Tipo
                ws.cell(row=proxima_linha, column=11, value=valores[3])  # Valor/Percentual
                ws.cell(row=proxima_linha, column=12, value=valores[4])  # Valor Total
                ws.cell(row=proxima_linha, column=13, value=valores[5])  # Nº Parcelas
                
                if valores[2] == 'Fixo' and 'entrada' in (tags or ()):
                    # Se é uma entrada, registra apenas na aba Dados
                    try:
                        # Calcular data de relatório para entrada
                        hoje = datetime.now()
                        if hoje.day <= 5:
                            data_rel = hoje.replace(day=5)
                        elif hoje.day <= 20:
                            data_rel = hoje.replace(day=20)
                        else:
                            data_rel = (hoje + relativedelta(months=1)).replace(day=5)

                        data_rel_str = data_rel.strftime('%d/%m/%Y')
                        data_vencto = datetime.strptime(valores[6], '%d/%m/%Y')
                        data_vencto_str = data_vencto.strftime('%d/%m/%Y')

                        # Registrar entrada na aba Dados
                        proxima_linha_dados = ws_dados.max_row + 1
                        ws_dados.cell(row=proxima_linha_dados, column=1, value=data_rel_str)
                        ws_dados.cell(row=proxima_linha_dados, column=2, value=2)  # Tipo despesa 2
                        ws_dados.cell(row=proxima_linha_dados, column=3, value=cnpj_cpf)
                        ws_dados.cell(row=proxima_linha_dados, column=4, value=nome_admin)
                        ws_dados.cell(row=proxima_linha_dados, column=5, value=f"ADM OBRA - ENTRADA - CONTRATO {num_contrato}")
                        ws_dados.cell(row=proxima_linha_dados, column=7, value=valores[3])  # Valor da entrada
                        ws_dados.cell(row=proxima_linha_dados, column=8, value=1)
                        ws_dados.cell(row=proxima_linha_dados, column=9, value=valores[3])
                        ws_dados.cell(row=proxima_linha_dados, column=10, value=data_vencto_str)
                        ws_dados.cell(row=proxima_linha_dados, column=11, value='ADM')
                        ws_dados.cell(row=proxima_linha_dados, column=12, value='')
                        ws_dados.cell(row=proxima_linha_dados, column=13, value='LANÇAMENTO AUTOMÁTICO')

                    except (ValueError, TypeError) as e:
                        messagebox.showwarning(
                            "Aviso", 
                            f"Erro ao processar entrada para {nome_admin}: {str(e)}"
                        )

                elif valores[2] == 'Fixo' and not 'entrada' in (tags or ()):
                    # Se não é entrada, registra apenas as parcelas normais
                    try:
                        data_inicial = datetime.strptime(valores[6], '%d/%m/%Y')
                        num_parcelas = int(valores[5])
                        valor_parcela = float(str(valores[3]).replace(',', '.'))

                        for i in range(num_parcelas):
                            if data_inicial.month + i > 12:
                                ano = data_inicial.year + ((data_inicial.month + i - 1) // 12)
                                mes = ((data_inicial.month + i - 1) % 12) + 1
                            else:
                                ano = data_inicial.year
                                mes = data_inicial.month + i
                            
                            data_vencto = data_inicial.replace(year=ano, month=mes, day=5)
                            data_vencto_str = data_vencto.strftime('%d/%m/%Y')
                            
                            proxima_linha = ws.max_row + 1
                            ws.cell(row=proxima_linha, column=25, value=num_contrato)
                            ws.cell(row=proxima_linha, column=26, value=i + 1)
                            ws.cell(row=proxima_linha, column=27, value=cnpj_cpf)
                            ws.cell(row=proxima_linha, column=28, value=nome_admin)
                            ws.cell(row=proxima_linha, column=29, value=data_vencto)
                            ws.cell(row=proxima_linha, column=30, value=valor_parcela)
                            ws.cell(row=proxima_linha, column=31, value='PENDENTE')

                    except (ValueError, TypeError) as e:
                        messagebox.showwarning(
                            "Aviso", 
                            f"Erro ao processar parcelas para {nome_admin}: {str(e)}"
                        )

            wb.save(self.arquivo_cliente)
            messagebox.showinfo("Sucesso", "Contrato cadastrado com sucesso!")
            janela.destroy()
            self.carregar_contratos()

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar contrato: {str(e)}")
            if 'wb' in locals():
                wb.close()  
                

    def excluir_contrato(self):
        """Exclui o contrato selecionado"""
        selecionado = self.tree_contratos.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um contrato para excluir")
            return
            
        if messagebox.askyesno("Confirmação", 
                              "Deseja realmente excluir este contrato e seus administradores?"):
            try:
                num_contrato = self.tree_contratos.item(selecionado)['values'][0]
                
                wb = load_workbook(self.arquivo_cliente)
                ws = wb['Contratos_ADM']
                
                # Marcar contrato como inativo
                for row in ws.iter_rows(min_row=2):
                    if row[0].value == num_contrato:
                        row[3].value = 'INATIVO'  # Coluna D - Status
                
                wb.save(self.arquivo_cliente)
                self.carregar_contratos()
                messagebox.showinfo("Sucesso", "Contrato marcado como inativo")
                
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao excluir contrato: {str(e)})")

            
class GestaoTaxasFixas:
    def __init__(self, sistema_principal):
        self.sistema = sistema_principal
        self.gestor_parcelas = GestorParcelas(self)

    def processar_lancamentos_fixos(self, cliente, data_ref):
        """Processa os lançamentos de taxas fixas para a data de referência"""
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            wb = load_workbook(arquivo_cliente)
            ws = wb['Contratos_ADM']
            
            lancamentos_gerados = []
            
            # Buscar contratos ativos com taxa fixa
            for row in ws.iter_rows(min_row=3, values_only=True):
                # Verifica se é registro de administrador e tipo fixo
                if (row[6] and  # Tem nº contrato na coluna G
                    row[9] == 'Fixo' and  # É tipo fixo
                    self.contrato_ativo(ws, row[6])):  # Contrato está ativo
                    
                    # Verificar se já tem lançamento para este período
                    if not self.tem_lancamento(ws, row[6], row[7], data_ref):
                        # Preparar dados para o lançamento
                        dados_lancamento = {
                            'data_rel': data_ref,
                            'cnpj_cpf': row[7],  # CNPJ/CPF
                            'nome': row[8],      # Nome/Razão Social
                            'referencia': f'ADM FIXA REF. {data_ref.strftime("%m/%Y")}',
                            'valor': float(row[10].replace(',', '.')),  # Valor/Parcela
                            'dt_vencto': self.calcular_vencimento(data_ref)
                        }
                        
                        # Registrar lançamento no sistema
                        self.sistema.dados_para_incluir.append(dados_lancamento)
                        lancamentos_gerados.append(dados_lancamento)
                        
                        # Registrar na aba de controle
                        self.registrar_lancamento(ws, dados_lancamento)
                        
            wb.save(arquivo_cliente)
            return lancamentos_gerados
            
        except Exception as e:
            raise Exception(f"Erro ao processar lançamentos fixos: {str(e)}")

    def contrato_ativo(self, ws, num_contrato):
        """Verifica se o contrato está ativo"""
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] == num_contrato:  # Coluna A (Nº Contrato)
                return row[3] == 'ATIVO'  # Coluna D (Status)
        return False

    def tem_lancamento(self, ws, num_contrato, cnpj_cpf, data_ref):
        """Verifica se já existe lançamento para o período"""
        data_str = data_ref.strftime("%d/%m/%Y")
        for row in ws.iter_rows(min_row=3, values_only=True):
            if (row[25] and  # Tem referência na coluna PARCELAS
                row[24] == num_contrato and  # Mesmo contrato
                row[26] == cnpj_cpf and  # Mesmo CNPJ/CPF
                row[28] == data_str):  # Mesma data
                return True
        return False

    def calcular_vencimento(self, data_ref):
        """Calcula data de vencimento (dia 5 do mês seguinte)"""
        if data_ref.day == 5:
            vencto = data_ref.replace(day=20)
        else:  # day == 20
            if data_ref.month == 12:
                vencto = data_ref.replace(year=data_ref.year + 1, month=1, day=5)
            else:
                vencto = data_ref.replace(month=data_ref.month + 1, day=5)
        return vencto

    def registrar_lancamento(self, ws, dados):
        """Registra o lançamento na aba de controle"""
        proxima_linha = ws.max_row + 1
        ws.cell(row=proxima_linha, column=26, value=dados['cnpj_cpf'])
        ws.cell(row=proxima_linha, column=27, value=dados['nome'])
        ws.cell(row=proxima_linha, column=28, value=dados['data_rel'])
        ws.cell(row=proxima_linha, column=29, value=dados['valor'])
        ws.cell(row=proxima_linha, column=30, value='LANÇADO')


class GestaoAdministradores:
    def __init__(self, parent):
        self.parent = parent
        self.administradores = []  # Lista para armazenar os administradores
        
    def abrir_janela_admin(self):
        """Abre janela para gestão de administradores"""
        self.janela_admin = tk.Toplevel(self.parent)
        self.janela_admin.title("Gestão de Administradores")
        self.janela_admin.geometry("800x600")
        
        # Frame para busca de fornecedor
        frame_busca = ttk.LabelFrame(self.janela_admin, text="Buscar Fornecedor")
        frame_busca.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(frame_busca, text="Nome:").pack(side='left', padx=5)
        self.busca_entry = ttk.Entry(frame_busca, width=40)
        self.busca_entry.pack(side='left', padx=5)
        busca_entry.bind('<Return>', lambda e: buscar())
        tree_fornecedores.bind('<<TreeviewSelect>>', selecionar)
        ttk.Button(frame_busca, text="Buscar", command=buscar).pack(side='left', padx=5)

        
        # Frame para lista de fornecedores
        frame_fornecedores = ttk.LabelFrame(self.janela_admin, text="Fornecedores")
        frame_fornecedores.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.tree_fornecedores = ttk.Treeview(frame_fornecedores, 
                                             columns=('CNPJ/CPF', 'Nome', 'Categoria'),
                                             show='headings',
                                             height=5)
        self.tree_fornecedores.heading('CNPJ/CPF', text='CNPJ/CPF')
        self.tree_fornecedores.heading('Nome', text='Nome')
        self.tree_fornecedores.heading('Categoria', text='Categoria')
        self.tree_fornecedores.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Frame para percentual
        frame_percentual = ttk.Frame(self.janela_admin)
        frame_percentual.pack(fill='x', padx=5, pady=5)
        
        ttk.Label(frame_percentual, text="Percentual (%):").pack(side='left', padx=5)
        self.percentual_entry = ttk.Entry(frame_percentual, width=10)
        self.percentual_entry.pack(side='left', padx=5)
        
        ttk.Button(frame_percentual, 
                  text="Adicionar Administrador", 
                  command=self.adicionar_administrador).pack(side='left', padx=5)
        
        # Frame para lista de administradores
        frame_lista = ttk.LabelFrame(self.janela_admin, text="Administradores Cadastrados")
        frame_lista.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.tree_admin = ttk.Treeview(frame_lista, 
                                     columns=('CNPJ/CPF', 'Nome', 'Percentual'),
                                     show='headings')
        self.tree_admin.heading('CNPJ/CPF', text='CNPJ/CPF')
        self.tree_admin.heading('Nome', text='Nome')
        self.tree_admin.heading('Percentual', text='Percentual (%)')
        self.tree_admin.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Frame para botões de ação
        frame_botoes = ttk.Frame(self.janela_admin)
        frame_botoes.pack(fill='x', padx=5, pady=5)
        
        ttk.Button(frame_botoes, 
                  text="Remover Administrador", 
                  command=self.remover_administrador).pack(side='left', padx=5)
        ttk.Button(frame_botoes, 
                  text="Concluir", 
                  command=self.finalizar_gestao).pack(side='right', padx=5)

    def buscar_fornecedor(self):
        termo = self.busca_entry.get()
        buscar_fornecedor(self.tree_fornecedores, termo)
            
    def adicionar_administrador(self):
        """Adiciona um fornecedor selecionado como administrador"""
        selecionado = self.tree_fornecedores.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um fornecedor")
            return
            
        fornecedor = self.tree_fornecedores.item(selecionado)['values']
        percentual = self.percentual_entry.get().strip()
        
        # Validar percentual
        if not percentual:
            messagebox.showerror("Erro", "Informe o percentual!")
            return
            
        try:
            percentual_float = float(percentual.replace(',', '.'))
            if percentual_float <= 0 or percentual_float > 100:
                messagebox.showerror("Erro", "Percentual deve estar entre 0 e 100!")
                return
        except ValueError:
            messagebox.showerror("Erro", "Percentual inválido!")
            return
            
        # Formatar CNPJ/CPF como string
        cnpj_cpf = str(fornecedor[0]).strip()  # Converter para string e remover espaços
        
        # Verificar se o fornecedor já está na lista
        for admin in self.administradores:
            if admin[0] == cnpj_cpf:  # Compara CNPJ/CPF
                messagebox.showerror("Erro", "Este fornecedor já está cadastrado como administrador!")
                return
                
        # Verificar se o total de percentuais não excede 100%
        total_atual = sum(float(item[2].replace(',', '.')) 
                         for item in self.administradores)
        if total_atual + percentual_float > 100:
            messagebox.showerror("Erro", "Soma dos percentuais excede 100%!")
            return
            
        # Adicionar à lista e à treeview usando o CNPJ/CPF como string
        self.administradores.append((cnpj_cpf, fornecedor[1], percentual))
        self.tree_admin.insert('', 'end', values=(cnpj_cpf, fornecedor[1], percentual))
        
        # Limpar campo de percentual
        self.percentual_entry.delete(0, tk.END)
        
    def remover_administrador(self):
        """Remove o administrador selecionado"""
        selecionado = self.tree_admin.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um administrador para remover")
            return
        
        self.tree_admin.delete(selecionado)
        valores = self.tree_admin.item(selecionado)['values']
        self.administradores = [(cnpj, nome, perc) for cnpj, nome, perc 
                              in self.administradores 
                              if cnpj != valores[0]]
        
    def finalizar_gestao(self):
        """Finaliza a gestão de administradores"""
        total = sum(float(perc.replace(',', '.')) 
                   for _, _, perc in self.administradores)
        if total > 100:
            messagebox.showerror("Erro", "Soma dos percentuais excede 100%!")
            return
        
        self.janela_admin.destroy()
        
    def get_administradores(self):
        """Retorna a lista de administradores configurados"""
        return self.administradores.copy()        



class GestorParcelas:
    def __init__(self, parent):
        print("Inicializando GestorParcelas")  # Debug
        self.parent = parent
        self.parcelas = []
        self.tipo_despesa_valor = '3'
        self.janela_parcelas = None
        self._var_tem_entrada = None  # Inicializa como None
        # Limpar referências de widgets
        self.frame_modalidade = None
        self.frame_valor_entrada = None
        self.lbl_entrada = None
        self.valor_entrada = None
        self.modalidade_entrada = None

    @property
    def tem_entrada(self):
        """Getter para tem_entrada - cria apenas quando necessário"""
        if self._var_tem_entrada is None:
            self._var_tem_entrada = tk.BooleanVar(master=self.parent.root, value=False)
        return self._var_tem_entrada


    # Interface e Controles
    def abrir_janela_parcelas(self):
        print("Abrindo janela de parcelas")  # Debug
        # Criar janela como Toplevel do parent
        self.janela_parcelas = tk.Toplevel(self.parent.root)
        self.janela_parcelas.title("Configuração de Parcelas")
        self.janela_parcelas.geometry("600x700")
        
        # Garantir que a janela seja modal
        self.janela_parcelas.transient(self.parent.root)
        self.janela_parcelas.grab_set()
        
        frame = ttk.Frame(self.janela_parcelas, padding="10")
        frame.pack(fill='both', expand=True)

        # Frame para entrada
        frame_entrada = ttk.LabelFrame(frame, text="Entrada")
        frame_entrada.grid(row=0, column=0, columnspan=2, sticky='ew', padx=5, pady=5)
        
        print("Criando Checkbutton")  # Debug
        check = ttk.Checkbutton(
            frame_entrada, 
            text="Possui entrada?",
            variable=self.tem_entrada,
            command=self.atualizar_campos_entrada
        )
        check.grid(row=0, column=0, padx=5, pady=5)

        # Frame para modalidades de entrada
        print("Criando frame modalidade")  # Debug
        self.frame_modalidade = ttk.Frame(frame_entrada)
        self.frame_modalidade.grid(row=1, column=0, columnspan=2, sticky='ew', padx=5, pady=5)
        
        ttk.Label(self.frame_modalidade, text="Modalidade de Entrada:").grid(row=0, column=0, padx=5, pady=2)
        self.modalidade_entrada = ttk.Combobox(self.frame_modalidade, state='readonly', width=40)
        self.modalidade_entrada['values'] = [
            "Percentual do valor total na primeira parcela",
            "Primeira parcela igual às demais (arredonda no final)",
            "Valor específico na primeira parcela"
        ]
        self.modalidade_entrada.grid(row=0, column=1, padx=5, pady=2)
        

        # Garantir que o frame modalidade começa oculto
        print("Ocultando frame modalidade inicialmente")  # Debug
        self.frame_modalidade.grid_remove()
        
        # Frame para valor da entrada (dinâmico baseado na modalidade)
        self.frame_valor_entrada = ttk.Frame(frame_entrada)
        self.frame_valor_entrada.grid(row=2, column=0, columnspan=2, sticky='ew', padx=5, pady=5)
        
        # Ocultar frames inicialmente
        self.frame_modalidade.grid_remove()
        self.frame_valor_entrada.grid_remove()
        
        # Tipo de Despesa
        ttk.Label(frame, text="Tipo de Despesa:").grid(row=1, column=0, padx=5, pady=5)
        self.tipo_despesa = ttk.Combobox(frame, values=['2', '3'], state='readonly', width=5)
        self.tipo_despesa.grid(row=1, column=1, sticky='w', padx=5, pady=5)
        self.tipo_despesa.set('3')  # Tipo 3 como padrão

        # Tipo de Parcelamento
        ttk.Label(frame, text="Tipo de Parcelamento:").grid(row=2, column=0, padx=5, pady=5)
        self.tipo_parcelamento = ttk.Combobox(frame, values=[
            "Prazo Fixo em Dias",
            "Datas Específicas",
            "Cartão de Crédito"
        ], state="readonly")
        self.tipo_parcelamento.grid(row=2, column=1, padx=5, pady=5)
        self.tipo_parcelamento.set("Prazo Fixo em Dias")
        self.tipo_parcelamento.bind('<<ComboboxSelected>>', self.atualizar_campos_parcelamento)

        # Frame para campos dinâmicos
        self.frame_dinamico = ttk.Frame(frame)
        self.frame_dinamico.grid(row=3, column=0, columnspan=2, pady=10)

        # Campos comuns
        ttk.Label(frame, text="Data da Despesa:").grid(row=4, column=0, padx=5, pady=5)
        self.data_despesa = DateEntry(
            frame,
            format='dd/mm/yyyy',
            locale='pt_BR',
            background='darkblue',
            foreground='white',
            borderwidth=2
        )
        
        self.data_despesa.grid(row=4, column=1, padx=5, pady=5)
        self.data_despesa.configure(state='normal')
        self.configurar_calendario(self.data_despesa)

        ttk.Label(frame, text="Valor Original:").grid(row=5, column=0, padx=5, pady=5)
        self.valor_original = ttk.Entry(frame)
        self.valor_original.grid(row=5, column=1, padx=5, pady=5)

        # Alterar o label do número de parcelas para ser mais claro
        if self.tem_entrada.get():
            ttk.Label(frame, text="Número de Parcelas (além da entrada):").grid(row=6, column=0, padx=5, pady=5)
        else:
            ttk.Label(frame, text="Número de Parcelas:").grid(row=6, column=0, padx=5, pady=5)
        self.num_parcelas = ttk.Entry(frame)
        self.num_parcelas.grid(row=6, column=1, padx=5, pady=5)

        # Adicionar um label informativo
        self.lbl_info_parcelas = ttk.Label(frame, text="")
        self.lbl_info_parcelas.grid(row=7, column=0, columnspan=2, padx=5, pady=2)

        # Frame específico para informação sobre parcelas
        frame_info_parcelas = ttk.Frame(frame)
        frame_info_parcelas.grid(row=7, column=0, columnspan=2, padx=5, pady=5, sticky='ew')
        
        self.lbl_info_parcelas = ttk.Label(
            frame_info_parcelas, 
            text="",
            wraplength=500,  # Permitir quebra de linha se necessário
            justify='center'
        )
        self.lbl_info_parcelas.pack(fill='x', padx=5)

        # Referência Base (já existe)
        ttk.Label(frame, text="Referência Base:").grid(row=8, column=0, padx=5, pady=5)
        self.referencia_base = ttk.Entry(frame)
        self.referencia_base.grid(row=8, column=1, padx=5, pady=5, sticky='ew')

        # Adicionar campo NF
        ttk.Label(frame, text="NF:").grid(row=9, column=0, padx=5, pady=5)
        self.campos_nf = ttk.Entry(frame)
        self.campos_nf.grid(row=9, column=1, padx=5, pady=5, sticky='ew')

    

        # Botões
        frame_botoes = ttk.Frame(frame)
        frame_botoes.grid(row=11, column=0, columnspan=2, pady=20)

        ttk.Button(frame_botoes, 
                  text="Gerar Parcelas", 
                  command=self.gerar_parcelas).pack(side='left', padx=5)
        ttk.Button(frame_botoes, 
                  text="Cancelar", 
                  command=self.cancelar_parcelamento).pack(side='left', padx=5)

        # Inicializar campos do tipo padrão
        self.atualizar_campos_parcelamento(None)

        # Fazer a janela modal
        self.janela_parcelas.transient(self.parent.root)
        self.janela_parcelas.grab_set()

        # Centralizar a janela
        self.janela_parcelas.update_idletasks()
        width = self.janela_parcelas.winfo_width()
        height = self.janela_parcelas.winfo_height()
        x = (self.janela_parcelas.winfo_screenwidth() // 2) - (width // 2)
        y = (self.janela_parcelas.winfo_screenheight() // 2) - (height // 2)
        self.janela_parcelas.geometry(f'{width}x{height}+{x}+{y}')


    def atualizar_campos_entrada(self):
        """Mostra/oculta campos relacionados à entrada e atualiza labels"""
        if self.tem_entrada.get():
            # Mostrar frame modalidade
            if self.frame_modalidade:
                self.frame_modalidade.grid()
                
                # Criar campos se não existirem
                if not hasattr(self, 'valor_entrada') or not self.valor_entrada:
                    if not self.frame_valor_entrada:
                        self.frame_valor_entrada = ttk.Frame(self.frame_modalidade)
                        self.frame_valor_entrada.grid(row=1, column=0, columnspan=2, sticky='ew', padx=5, pady=5)
                    
                    self.lbl_entrada = ttk.Label(self.frame_valor_entrada, text="Valor:")
                    self.lbl_entrada.grid(row=0, column=0, padx=5, pady=2)
                    
                    self.valor_entrada = ttk.Entry(self.frame_valor_entrada)
                    self.valor_entrada.grid(row=0, column=1, padx=5, pady=2)
                
                if self.frame_valor_entrada:
                    self.frame_valor_entrada.grid()
        else:
            # Ocultar frames
            if self.frame_modalidade:
                self.frame_modalidade.grid_remove()
            if self.frame_valor_entrada:
                self.frame_valor_entrada.grid_remove()
            
            # Restaurar label original
            for widget in self.janela_parcelas.winfo_children():
                if isinstance(widget, ttk.Label) and widget.cget("text").startswith("Número de Parcelas"):
                    widget.config(text="Número de Parcelas:")
            self.lbl_info_parcelas.config(text="")

    def atualizar_campos_modalidade(self, event=None):
        """Atualiza campos baseado na modalidade selecionada"""
        modalidade = self.modalidade_entrada.get()
        
        if not hasattr(self, 'frame_valor_entrada') or not hasattr(self, 'lbl_entrada'):
            return
            
        self.frame_valor_entrada.grid()
        
        if modalidade == "Percentual do valor total na primeira parcela":
            self.lbl_entrada.config(text="Percentual (%): ")
            self.valor_entrada.delete(0, tk.END)
        elif modalidade == "Primeira parcela igual às demais (arredonda no final)":
            self.frame_valor_entrada.grid_remove()
        elif modalidade == "Valor específico na primeira parcela":
            self.lbl_entrada.config(text="Valor (R$): ")
            self.valor_entrada.delete(0, tk.END)
            
    def atualizar_campos_parcelamento(self, event):
        # Limpar frame dinâmico
        for widget in self.frame_dinamico.winfo_children():
            widget.destroy()

        tipo = self.tipo_parcelamento.get()
        
        if tipo == "Prazo Fixo em Dias":
            ttk.Label(self.frame_dinamico, text="Prazo entre Parcelas (dias):").grid(row=0, column=0, padx=5, pady=5)
            self.prazo_dias = ttk.Entry(self.frame_dinamico)
            self.prazo_dias.grid(row=0, column=1, padx=5, pady=5)
            self.prazo_dias.insert(0, "30")  # Valor padrão

        elif tipo == "Datas Específicas":
            num_parcelas_txt = "parcelas após a entrada" if self.tem_entrada.get() else "parcelas"
            
            ttk.Label(self.frame_dinamico, 
                     text=f"Informe as datas de vencimento das {num_parcelas_txt}:").grid(
                         row=0, column=0, columnspan=2, padx=5, pady=5)
            
            self.texto_datas = tk.Text(self.frame_dinamico, height=4, width=30)
            self.texto_datas.grid(row=1, column=0, columnspan=2, padx=5, pady=5)
            
            ttk.Label(self.frame_dinamico, 
                     text="Digite uma data por linha no formato dd/mm/aaaa\n"
                          "(não inclua a data da entrada)").grid(
                         row=2, column=0, columnspan=2, padx=5, pady=5)

        elif tipo == "Cartão de Crédito":
            ttk.Label(self.frame_dinamico, text="Dia do Vencimento:").grid(row=0, column=0, padx=5, pady=5)
            self.dia_vencimento = ttk.Entry(self.frame_dinamico, width=5)
            self.dia_vencimento.grid(row=0, column=1, padx=5, pady=5)
            self.dia_vencimento.insert(0, "10")  # Valor padrão


    # Métodos de geração e validação de parcelas
    def validar_dados_entrada(self, valor_original, num_parcelas, referencia_base, tipo):
        """Valida os dados básicos antes de gerar parcelas"""
        if not referencia_base or num_parcelas <= 0:
            messagebox.showerror("Erro", "Preencha todos os campos obrigatórios!")
            return False

        # Validações específicas por tipo de parcelamento
        if tipo == "Prazo Fixo em Dias":
            if not hasattr(self, 'prazo_dias') or not self.prazo_dias.get():
                messagebox.showerror("Erro", "Informe o prazo entre as parcelas!")
                return False
        elif tipo == "Datas Específicas":
            if not hasattr(self, 'texto_datas'):
                messagebox.showerror("Erro", "Configure as datas específicas!")
                return False
        elif tipo == "Cartão de Crédito":
            if not hasattr(self, 'dia_vencimento') or not self.dia_vencimento.get():
                messagebox.showerror("Erro", "Informe o dia do vencimento!")
                return False
            try:
                dia_vencimento = int(self.dia_vencimento.get())
                if not (1 <= dia_vencimento <= 31):
                    messagebox.showerror("Erro", "Dia de vencimento deve estar entre 1 e 31!")
                    return False
            except ValueError:
                messagebox.showerror("Erro", "Dia de vencimento inválido!")
                return False

        return True

    def gerar_parcelas(self):
        """Método principal para gerar parcelas"""
        try:
            # Coletar dados básicos
            self.tipo_despesa_valor = self.tipo_despesa.get()
            valor_original = float(self.valor_original.get().replace(',', '.'))
            num_parcelas = int(self.num_parcelas.get())
            referencia_base = self.referencia_base.get().strip()
            nf = self.campos_nf.get().strip()
##            print(f"NF capturado do formulário: {nf}")  # Adicione esta linha
            tipo = self.tipo_parcelamento.get()

            # Validar dados
            if not self.validar_dados_entrada(valor_original, num_parcelas, referencia_base, tipo):
                return False

            # Data base é a data da despesa
            data_base = datetime.strptime(self.data_despesa.get(), '%d/%m/%Y')
            
            # Limpar lista de parcelas anterior
            self.parcelas = []

            # Calcular valores das parcelas
            valores_parcelas = self.calcular_valores_parcelas(valor_original, num_parcelas)
            if not valores_parcelas:
                return False

            # Gerar parcelas conforme o tipo
            if tipo == "Prazo Fixo em Dias":
                self.gerar_parcelas_prazo_fixo(data_base, valores_parcelas, referencia_base, num_parcelas, nf)
            elif tipo == "Datas Específicas":
                self.gerar_parcelas_datas_especificas(data_base, valores_parcelas, referencia_base, num_parcelas, nf)
            elif tipo == "Cartão de Crédito":
                self.gerar_parcelas_cartao(data_base, valores_parcelas, referencia_base, num_parcelas, nf)

            if self.parcelas:
                messagebox.showinfo("Sucesso", f"{len(self.parcelas)} parcela(s) gerada(s) com sucesso!")
                self.limpar_campos()
                return True

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar parcelas: {str(e)}")
            return False

    def adicionar_parcela(self, data_rel, dt_vencto, valor_parcela, referencia_base, i, num_parcelas, eh_primeira_parcela, nf):
        """
        Método auxiliar para criar uma parcela com todos os dados necessários
        """
##        print(f"Adicionando parcela com NF: {nf}")  # Adicione esta linha
        parcela = {
            'data_rel': data_rel.strftime('%d/%m/%Y'),
            'dt_vencto': dt_vencto.strftime('%d/%m/%Y'),
            'valor': valor_parcela,
            'referencia': self.gerar_referencia_parcela(referencia_base, i, num_parcelas, eh_primeira_parcela),
            'nf': nf
        }
        self.parcelas.append(parcela)

    def gerar_parcelas_prazo_fixo(self, data_base, valores_parcelas, referencia_base, num_parcelas, nf):
        """Gera parcelas com prazo fixo em dias"""
        prazo_dias = int(self.prazo_dias.get())
        
        for i, valor_parcela in enumerate(valores_parcelas):
            eh_primeira_parcela = (i == 0)
            
            if eh_primeira_parcela and self.tem_entrada.get():
                dt_vencto = data_base
                data_rel = self.calcular_data_rel(data_base, dt_vencto, True)
            else:
                dt_vencto = data_base + relativedelta(days=prazo_dias * (i + (0 if self.tem_entrada.get() else 1)))
                dt_vencto = self.proximo_dia_util(dt_vencto)
                data_rel = self.calcular_data_rel(data_base, dt_vencto, eh_primeira_parcela)
            
            self.adicionar_parcela(
                data_rel,
                dt_vencto,
                valor_parcela,
                referencia_base,
                i,
                num_parcelas,
                eh_primeira_parcela,
                nf
            )

    def gerar_parcelas_datas_especificas(self, data_base, valores_parcelas, referencia_base, num_parcelas, nf):
        """Gera parcelas com datas específicas"""
        datas_texto = self.texto_datas.get("1.0", tk.END).strip().split('\n')
        datas_texto = [d.strip() for d in datas_texto if d.strip()]
        
        num_datas_esperado = num_parcelas
        if len(datas_texto) != num_datas_esperado:
            messagebox.showerror(
                "Erro", 
                f"Para {num_parcelas} {'parcelas após a entrada' if self.tem_entrada.get() else 'parcelas'}, "
                f"é necessário informar {num_datas_esperado} data(s) de vencimento."
            )
            return

        for i, valor_parcela in enumerate(valores_parcelas):
            eh_primeira_parcela = (i == 0)
            
            try:
                if eh_primeira_parcela and self.tem_entrada.get():
                    dt_vencto = data_base
                    data_rel = self.calcular_data_rel(data_base, dt_vencto, True)
                else:
                    idx_data = i - 1 if self.tem_entrada.get() else i
                    if 0 <= idx_data < len(datas_texto):
                        dt_vencto = datetime.strptime(datas_texto[idx_data], '%d/%m/%Y')
                        dt_vencto = self.proximo_dia_util(dt_vencto)
                        data_rel = self.calcular_data_rel(data_base, dt_vencto, eh_primeira_parcela)
                    else:
                        raise ValueError(f"Índice de data inválido: {idx_data}")
                
                self.adicionar_parcela(
                    data_rel,
                    dt_vencto,
                    valor_parcela,
                    referencia_base,
                    i,
                    num_parcelas,
                    eh_primeira_parcela,
                    nf
                )
                
            except ValueError as e:
                messagebox.showerror("Erro", f"Erro ao processar data: {str(e)}")
                return
            except IndexError:
                messagebox.showerror("Erro", "Número insuficiente de datas fornecidas")
                return

    def gerar_parcelas_cartao(self, data_base, valores_parcelas, referencia_base, num_parcelas, nf):
        """Gera parcelas para pagamento com cartão"""
        dia_vencimento = int(self.dia_vencimento.get())
        
        for i, valor_parcela in enumerate(valores_parcelas):
            eh_primeira_parcela = (i == 0)
            
            if eh_primeira_parcela:
                data_atual = data_base + relativedelta(months=1)
            else:
                data_atual = data_base + relativedelta(months=i + 1)
            
            try:
                dt_vencto = data_atual.replace(day=dia_vencimento)
            except ValueError:
                dt_vencto = data_atual + relativedelta(day=31)
            
            dt_vencto = self.proximo_dia_util(dt_vencto)
            
            if eh_primeira_parcela:
                hoje = datetime.now()
                if hoje.day <= 5:
                    data_rel = hoje.replace(day=5)
                elif hoje.day <= 20:
                    data_rel = hoje.replace(day=20)
                else:
                    proximo_mes = hoje + relativedelta(months=1)
                    data_rel = proximo_mes.replace(day=5)
            else:
                data_rel = self.calcular_data_rel(data_base, dt_vencto, False)
                
            self.adicionar_parcela(
                data_rel,
                dt_vencto,
                valor_parcela,
                referencia_base,
                i,
                num_parcelas,
                eh_primeira_parcela,
                nf
            )


    # Métodos de cálculo e utilitários
    def calcular_valores_parcelas(self, valor_original, num_parcelas):
        """Calcula os valores das parcelas considerando entrada se houver"""
        try:
            if self.tem_entrada.get():
                if not self.modalidade_entrada.get():
                    messagebox.showerror("Erro", "Selecione a modalidade de entrada!")
                    return None
                valores_parcelas = self.calcular_parcelas_entrada(valor_original, num_parcelas)
            else:
                valores_parcelas = self.calcular_parcelas_ajustadas(valor_original, num_parcelas)

            # Verificar se a soma está correta
            soma_parcelas = sum(valores_parcelas)
            if abs(soma_parcelas - valor_original) > 0.01:
                messagebox.showerror(
                    "Erro", 
                    f"Erro no cálculo das parcelas: soma ({soma_parcelas:.2f}) " 
                    f"diferente do valor original ({valor_original:.2f})!"
                )
                return None

            return valores_parcelas
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao calcular valores: {str(e)}")
            return None

    def calcular_parcelas_entrada(self, valor_total, num_parcelas):
        """Calcula valores das parcelas considerando a modalidade de entrada"""
        modalidade = self.modalidade_entrada.get()
        valores_parcelas = []
        
        # Se tem entrada, o número de parcelas informado é adicional à entrada
        num_parcelas_real = num_parcelas + 1 if self.tem_entrada.get() else num_parcelas
        
        if modalidade == "Percentual do valor total na primeira parcela":
            try:
                percentual = float(self.valor_entrada.get().replace(',', '.'))
                if not (0 < percentual < 100):
                    raise ValueError("Percentual deve estar entre 0 e 100")
                
                valor_entrada = (percentual / 100) * valor_total
                valor_restante = valor_total - valor_entrada
                
                valores_parcelas = [valor_entrada]  # Primeira parcela (entrada)
                # Distribuir o valor restante no número de parcelas informado
                demais_parcelas = self.calcular_parcelas_ajustadas(valor_restante, num_parcelas)
                valores_parcelas.extend(demais_parcelas)
                
            except ValueError as e:
                raise ValueError(f"Erro no percentual de entrada: {str(e)}")
        
        elif modalidade == "Primeira parcela igual às demais (arredonda no final)":
            # Dividir o valor total pelo número total de parcelas (incluindo entrada)
            valores_parcelas = self.calcular_parcelas_ajustadas(valor_total, num_parcelas_real)
            
        elif modalidade == "Valor específico na primeira parcela":
            try:
                valor_entrada = float(self.valor_entrada.get().replace(',', '.'))
                if valor_entrada >= valor_total:
                    raise ValueError("Valor da entrada não pode ser maior ou igual ao valor total")
                
                valor_restante = valor_total - valor_entrada
                valores_parcelas = [valor_entrada]  # Primeira parcela (entrada)
                # Distribuir o valor restante no número de parcelas informado
                demais_parcelas = self.calcular_parcelas_ajustadas(valor_restante, num_parcelas)
                valores_parcelas.extend(demais_parcelas)
                
            except ValueError as e:
                raise ValueError(f"Erro no valor da entrada: {str(e)}")
        
        return valores_parcelas

    def calcular_parcelas_ajustadas(self, valor_total, num_parcelas):
        """Calcula valores das parcelas garantindo que a soma seja igual ao valor total"""
        valor_parcela_base = valor_total / num_parcelas
        valor_parcela_round = round(valor_parcela_base, 2)
        
        # Calcular diferença total devido aos arredondamentos
        diferenca = valor_total - (valor_parcela_round * num_parcelas)
        
        # Distribuir a diferença na última parcela
        parcelas = [valor_parcela_round] * (num_parcelas - 1)
        ultima_parcela = valor_parcela_round + round(diferenca, 2)
        parcelas.append(ultima_parcela)
        
        return parcelas

    def calcular_data_rel(self, data_base, dt_vencto, eh_primeira_parcela):
        """
        Calcula a data do relatório com base na data de vencimento e tipo de despesa.
        Agora considera a data atual para não retroagir em períodos fechados.
        """
        try:
            hoje = datetime.now()
            
            # Se for entrada, calcula a partir da data atual
            if eh_primeira_parcela and self.tem_entrada.get():
                if hoje.day <= 5:
                    data_rel = hoje.replace(day=5)
                elif hoje.day <= 20:
                    data_rel = hoje.replace(day=20)
                else:
                    proximo_mes = hoje + relativedelta(months=1)
                    data_rel = proximo_mes.replace(day=5)
                return data_rel
                
            # Para as demais parcelas, manter a lógica existente
            tp_desp = self.tipo_despesa_valor
            
            if dt_vencto.day == 5:
                # Se vence dia 5, relatório é dia 20 do mês anterior
                data_rel = (dt_vencto - relativedelta(months=1)).replace(day=20)
            elif dt_vencto.day == 20:
                # Se vence dia 20, relatório é dia 5 do mesmo mês
                data_rel = dt_vencto.replace(day=5)
            elif tp_desp == '5':
                if dt_vencto.day <= 5:
                    data_rel = dt_vencto.replace(day=5)
                elif dt_vencto.day <= 20:
                    data_rel = dt_vencto.replace(day=20)
                else:
                    proximo_mes = dt_vencto + relativedelta(months=1)
                    data_rel = proximo_mes.replace(day=5)
            else:
                if dt_vencto.day <= 5:
                    data_rel = (dt_vencto - relativedelta(months=1)).replace(day=20)
                elif dt_vencto.day <= 20:
                    data_rel = dt_vencto.replace(day=5)
                else:
                    data_rel = dt_vencto.replace(day=20)
                    
            # Garantir que a data do relatório não seja anterior à data atual
            if data_rel < hoje:
                if hoje.day <= 5:
                    data_rel = hoje.replace(day=5)
                elif hoje.day <= 20:
                    data_rel = hoje.replace(day=20)
                else:
                    proximo_mes = hoje + relativedelta(months=1)
                    data_rel = proximo_mes.replace(day=5)
                    
            return data_rel
        except Exception as e:
            print(f"Erro ao calcular data do relatório: {str(e)}")
            return dt_vencto

    def configurar_calendario(self, dateentry):
        """Configura o comportamento do calendário"""
        def on_calendar_click(event):
            # Permite cliques no calendário
            return True
            
        def on_calendar_select(event):
            dateentry._top_cal.withdraw()  # Fecha o calendário
            self.janela_parcelas.after(100, lambda: self.janela_parcelas.focus_set())  # Retorna foco
        
        def on_calendar_focus(event):
            # Mantém o foco quando o calendário está aberto
            if dateentry._top_cal:
                dateentry._top_cal.focus_set()
            return True

        # Configurar bindings
        dateentry.bind('<<DateEntrySelected>>', on_calendar_select)
        dateentry.bind('<FocusIn>', on_calendar_focus)
        
        if hasattr(dateentry, '_top_cal'):
            cal = dateentry._top_cal
            cal.bind('<Button-1>', on_calendar_click)
            for w in cal.winfo_children():
                w.bind('<Button-1>', on_calendar_click)

        
    def proximo_dia_util(self, data):
        """
        Ajusta a data para o próximo dia útil se cair em fim de semana ou feriado
        """
        # Lista de feriados nacionais fixos
        feriados_fixos = [
            (1, 1),   # Ano Novo
            (21, 4),  # Tiradentes
            (1, 5),   # Dia do Trabalho
            (7, 9),   # Independência
            (12, 10), # Nossa Senhora
            (2, 11),  # Finados
            (15, 11), # Proclamação da República
            (25, 12), # Natal
        ]

        while True:
            # Verifica se é fim de semana
            if data.weekday() >= 5:  # 5 = Sábado, 6 = Domingo
                data = data + relativedelta(days=1)
                continue

            # Verifica se é feriado fixo
            if (data.day, data.month) in feriados_fixos:
                data = data + relativedelta(days=1)
                continue

            # Se não é fim de semana nem feriado, é dia útil
            break

        return data

    def gerar_referencia_parcela(self, referencia_base, indice, num_parcelas, eh_primeira_parcela):
        """Gera a referência apropriada para a parcela"""
        if eh_primeira_parcela and self.tem_entrada.get():
            return f"{referencia_base} - ENTRADA"
        else:
            if self.tem_entrada.get():
                # Para as parcelas após a entrada
                return f"{referencia_base} - PARC. {indice}/{num_parcelas}"
            else:
                # Para parcelamento sem entrada
                return f"{referencia_base} - PARC. {indice + 1}/{num_parcelas}"
           
    # Métodos de limpeza e finalização
    def limpar_campos(self):
        """Limpa todos os campos após sucesso"""
        # Limpar referências de widgets
        self.frame_modalidade = None
        self.frame_valor_entrada = None
        self.lbl_entrada = None
        self.valor_entrada = None
        self.modalidade_entrada = None
        
        # Resetar checkbox
        if self._var_tem_entrada:
            self._var_tem_entrada.set(False)
        
        # Fechar janela
        if self.janela_parcelas:
            self.janela_parcelas.destroy()
            self.janela_parcelas = None

    def cancelar_parcelamento(self):
        """Cancela o parcelamento e limpa todos os campos"""
        self.parcelas = []
        
        # Limpar referências de widgets
        self.frame_modalidade = None
        self.frame_valor_entrada = None
        self.lbl_entrada = None
        self.valor_entrada = None
        self.modalidade_entrada = None
        
        # Resetar variável de entrada
        if self._var_tem_entrada:
            self._var_tem_entrada.set(False)
        
        if self.janela_parcelas:
            self.janela_parcelas.destroy()
            self.janela_parcelas = None

            
        


    # Fechando os métodos/classes anteriores
    def run(self):
        """Inicia a execução do sistema"""
        self.root.mainloop()

    def __del__(self):
        """Destrutor da classe"""
        if hasattr(self, 'root'):
            self.root.destroy()


# Aqui termina a última classe
# Agora pode vir o if __name__ == "__main__"
if __name__ == "__main__":
    app = SistemaEntradaDados()
    app.root.mainloop()
