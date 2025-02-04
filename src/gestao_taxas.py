import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import os
import sys
from pathlib import Path

# Ajusta o path para acessar a pasta config
current_dir = Path(__file__).resolve().parent
root_dir = current_dir.parent
sys.path.insert(0, str(root_dir))  # Mudamos append para insert(0)

from config.utils import (
    PASTA_CLIENTES,
    validar_data,
    formatar_moeda
)

class GestaoTaxasAdministracao:
    def __init__(self, parent=None):
        self.parent = parent
        self.controle_pagamentos = None
        self.finalizacao_quinzena = None
        # Não inicializar aqui - será feito sob demanda
    
    def inicializar_controles(self):
        """Inicializa os controladores de taxa"""
        try:
            print("Preparando sistema...")
            
            # Imports locais para evitar circular imports
            from controle_pagamentos import ControlePagamentos
            from finalizacao_quinzena import FinalizacaoQuinzena
            
            print("Carregando módulo de Controle de Pagamentos...")
            self._classe_controle = ControlePagamentos
            
            print("Carregando módulo de Finalização de Quinzena...")
            self._classe_quinzena = FinalizacaoQuinzena
            
            # Verificar se ambos foram carregados
            if not hasattr(self, '_classe_controle') or not hasattr(self, '_classe_quinzena'):
                raise Exception("Falha ao carregar um ou mais módulos")
            
            print("Sistema preparado com sucesso!")
            return True
            
        except Exception as e:
            print(f"Erro ao preparar sistema: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao preparar sistema: {str(e)}")
            return False
    
    def atualizar_status(self):
        """Atualiza o status dos módulos"""
        print("\nAtualizando status do sistema...")
        
        # Se os módulos não estiverem preparados, inicializar
        if not hasattr(self, '_classe_controle') or not hasattr(self, '_classe_quinzena'):
            print("Sistema não está completamente preparado. Iniciando preparação...")
            self.inicializar_controles()
        
        # Verificar novamente após a inicialização
        sistema_preparado = hasattr(self, '_classe_controle') and hasattr(self, '_classe_quinzena')
        
        if hasattr(self, 'status_controle'):
            if sistema_preparado:
                print("Sistema de Controle de Pagamentos está disponível")
                self.status_controle.config(
                    text="Controle de Pagamentos: Disponível",
                    foreground="green"
                )
            else:
                print("Sistema de Controle de Pagamentos não está disponível")
                self.status_controle.config(
                    text="Controle de Pagamentos: Não disponível",
                    foreground="red"
                )
            
        if hasattr(self, 'status_quinzena'):
            if sistema_preparado:
                print("Sistema de Finalização de Quinzena está disponível")
                self.status_quinzena.config(
                    text="Finalização de Quinzena: Disponível",
                    foreground="green"
                )
            else:
                print("Sistema de Finalização de Quinzena não está disponível")
                self.status_quinzena.config(
                    text="Finalização de Quinzena: Não disponível",
                    foreground="red"
                )

    def abrir_controle_pagamentos(self):
        """Abre a interface de controle de pagamentos"""
        if not hasattr(self, '_classe_controle'):
            if not self.inicializar_controles():
                return
        
        try:
            print("\nAbrindo Controle de Pagamentos...")
            # Criar nova instância apenas quando necessário
            self.controle_pagamentos = self._classe_controle(self.parent)
            self.parent.withdraw()
            # Fechar a janela de menu de taxas
            if hasattr(self, 'menu') and self.menu:
                self.menu.destroy()
            self.controle_pagamentos.run()
        except Exception as e:
            print(f"Erro ao abrir controle de pagamentos: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao abrir controle de pagamentos: {str(e)}")
            self.parent.deiconify()
    
    def abrir_finalizacao_quinzena(self):
        """Abre a interface de finalização de quinzena"""
        if not hasattr(self, '_classe_quinzena'):
            if not self.inicializar_controles():
                return
        
        try:
            print("\nAbrindo Finalização de Quinzena...")
            # Criar nova instância apenas quando necessário
            self.finalizacao_quinzena = self._classe_quinzena(self.parent)
            self.parent.withdraw()
            # Fechar a janela de menu de taxas
            if hasattr(self, 'menu') and self.menu:
                self.menu.destroy()
            self.finalizacao_quinzena.run()
        except Exception as e:
            print(f"Erro ao abrir finalização de quinzena: {str(e)}")
            messagebox.showerror("Erro", f"Erro ao abrir finalização de quinzena: {str(e)}")
            self.parent.deiconify()
    
    def abrir_menu_taxas(self):
        """Abre o menu principal de taxas"""
        self.menu = tk.Toplevel(self.parent)
        self.menu.title("Gestão de Taxas de Administração")
        self.menu.geometry("500x400")
        
        # Frame principal
        main_frame = ttk.Frame(self.menu, padding="20")
        main_frame.pack(fill='both', expand=True)
        
        # Título
        ttk.Label(
            main_frame,
            text="Gestão de Taxas de Administração",
            font=('Helvetica', 16, 'bold')
        ).pack(pady=(0, 20))
        
        # Frame para botões
        buttons_frame = ttk.Frame(main_frame)
        buttons_frame.pack(fill='x', pady=20)
        
        # Botões principais
        ttk.Button(
            buttons_frame,
            text="Controle de Pagamentos de Taxa",
            command=self.abrir_controle_pagamentos,
            width=30
        ).pack(pady=10)
        
        ttk.Button(
            buttons_frame,
            text="Finalização de Quinzena",
            command=self.abrir_finalizacao_quinzena,
            width=30
        ).pack(pady=10)
        
        # Frame para status
        status_frame = ttk.LabelFrame(main_frame, text="Status do Sistema", padding="10")
        status_frame.pack(fill='x', pady=20)
        
        # Informações de status
        self.status_controle = ttk.Label(status_frame, text="Controle de Pagamentos: Não inicializado")
        self.status_controle.pack(anchor='w')
        
        self.status_quinzena = ttk.Label(status_frame, text="Finalização de Quinzena: Não inicializado")
        self.status_quinzena.pack(anchor='w')
        
        # Botão de atualização
        ttk.Button(
            main_frame,
            text="Atualizar Status",
            command=self.atualizar_status
        ).pack(pady=10)
        
        # Botão fechar
        ttk.Button(
            main_frame,
            text="Fechar",
            command=self.menu.destroy
        ).pack(pady=20)
        
        # Centralizar janela
        self.menu.update_idletasks()
        width = self.menu.winfo_width()
        height = self.menu.winfo_height()
        x = (self.menu.winfo_screenwidth() // 2) - (width // 2)
        y = (self.menu.winfo_screenheight() // 2) - (height // 2)
        self.menu.geometry(f'{width}x{height}+{x}+{y}')
        
        # Atualizar status inicial
        self.atualizar_status()

        
    
    def verificar_conflitos(self, cliente, data_ref):
        """Verifica se existem conflitos entre os dois tipos de lançamento"""
        try:
            arquivo_cliente = PASTA_CLIENTES / f"{cliente}.xlsx"
            if not os.path.exists(arquivo_cliente):
                return False, "Arquivo do cliente não encontrado"
            
            # Converter data para datetime se necessário
            if isinstance(data_ref, str):
                data_ref = datetime.strptime(data_ref, '%d/%m/%Y')
            
            from openpyxl import load_workbook
            
            wb = load_workbook(arquivo_cliente)
            ws_dados = wb['Dados']
            
            # Verificar lançamentos na mesma data
            lancamentos = []
            for row in ws_dados.iter_rows(min_row=2, values_only=True):
                if row[0] and isinstance(row[0], datetime):
                    if (row[0].day == data_ref.day and 
                        row[0].month == data_ref.month and 
                        row[0].year == data_ref.year):
                        
                        tipo = row[1]  # Tipo do lançamento
                        valor = row[7]  # Valor
                        descricao = row[4]  # Descrição
                        
                        lancamentos.append({
                            'tipo': tipo,
                            'valor': valor,
                            'descricao': descricao
                        })
            
            wb.close()
            
            # Verificar conflitos
            tem_taxa_fixa = any(l['tipo'] == 2 for l in lancamentos)  # Tipo 2 = Taxa fixa
            tem_taxa_percentual = any(l['tipo'] == 7 for l in lancamentos)  # Tipo 7 = Taxa percentual
            
            if tem_taxa_fixa and tem_taxa_percentual:
                return True, "ATENÇÃO: Já existem lançamentos de taxa fixa e percentual na mesma data"
            elif tem_taxa_fixa:
                return True, "ATENÇÃO: Já existe lançamento de taxa fixa nesta data"
            elif tem_taxa_percentual:
                return True, "ATENÇÃO: Já existe lançamento de taxa percentual nesta data"
                
            return False, "Nenhum conflito encontrado"
            
        except Exception as e:
            return True, f"Erro ao verificar conflitos: {str(e)}"

