# utils.py

from datetime import datetime
from dateutil.relativedelta import relativedelta
import pandas as pd
from validate_docbr import CPF, CNPJ
import os
from openpyxl import load_workbook
import re
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from datetime import datetime

from pathlib import Path


from .config import (
    BASE_PATH,
    PASTA_CLIENTES,
    ARQUIVO_CLIENTES,
    ARQUIVO_FORNECEDORES,
    ARQUIVO_MODELO,
    ARQUIVO_CONTROLE
)

# === DATA VALIDATION AND FORMATTING ===
def validar_data(data_str):
    """Valida se uma string está no formato de data correto"""
    try:
        if not data_str:
            return False
        if not re.match(r'^\d{2}/\d{2}/\d{4}$', data_str):
            return False
        datetime.strptime(data_str, '%d/%m/%Y')
        return True
    except ValueError:
        return False

def validar_data_quinzena(data):
    """Valida se a data é dia 5 ou 20 e ajusta se necessário"""
    if not (data.day == 5 or data.day == 20):
        if data.day < 5:
            data_ajustada = data.replace(day=5)
            msg = f"Data ajustada para dia 5: {data_ajustada.strftime('%d/%m/%Y')}"
        elif data.day < 20:
            data_ajustada = data.replace(day=20)
            msg = f"Data ajustada para dia 20: {data_ajustada.strftime('%d/%m/%Y')}"
        else:
            if data.month == 12:
                data_ajustada = data.replace(year=data.year + 1, month=1, day=5)
            else:
                data_ajustada = data.replace(month=data.month + 1, day=5)
            msg = f"Data ajustada para dia 5 do próximo mês: {data_ajustada.strftime('%d/%m/%Y')}"
        return data_ajustada, msg
    return data, None

def calcular_proxima_data_quinzena(data):
    """Calcula a próxima data quinzenal"""
    if data.day == 5:
        return data.replace(day=20)
    else:
        if data.month == 12:
            return data.replace(year=data.year + 1, month=1, day=5)
        else:
            return data.replace(month=data.month + 1, day=5)

# === DOCUMENT VALIDATION ===
def validar_cnpj_cpf(documento):
    """Valida CNPJ ou CPF"""
    if len(documento) <= 11:
        cpf = CPF()
        return cpf.validate(documento)
    else:
        cnpj = CNPJ()
        return cnpj.validate(documento)

def formatar_cnpj_cpf(documento):
    """Formata CNPJ/CPF com zeros à esquerda"""
    if len(documento) <= 11:
        return documento.zfill(11)
    return documento.zfill(14)

# === FILE OPERATIONS ===
def verificar_arquivo_excel(caminho):
    """Verifica se arquivo Excel existe e pode ser aberto"""
    try:
        if not os.path.exists(caminho):
            return False
        wb = load_workbook(caminho)
        wb.close()
        return True
    except Exception:
        return False

# === VALUE FORMATTING ===
def formatar_moeda(valor):
    """Formata valor para moeda brasileira"""
    try:
        return f"R$ {float(valor):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except (ValueError, TypeError):
        return "R$ 0,00"


# === FIND SUPPLIER ===
def buscar_fornecedor(tree_fornecedores, termo_busca):
    """Busca fornecedores na base e atualiza o treeview"""
    for item in tree_fornecedores.get_children():
        tree_fornecedores.delete(item)
        
    termo = termo_busca.lower()
    try:
        wb = load_workbook(ARQUIVO_FORNECEDORES)
        ws = wb['Fornecedores']
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if termo in str(row[3]).lower():  # Busca no nome
                tree_fornecedores.insert('', 'end', values=(row[0], row[3], row[11]))
        
        wb.close()
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao buscar fornecedores: {str(e)}")

def selecionar_fornecedor(tree_fornecedores, campos_fornecedor, campos_despesa=None, notebook=None, buscar_fornecedor_completo=None):
    """Preenche campos com o fornecedor selecionado"""
    selecionado = tree_fornecedores.selection()
    if not selecionado:
        messagebox.showwarning("Aviso", "Selecione um fornecedor")
        return None

    fornecedor = tree_fornecedores.item(selecionado)['values']

    # Para o caso de campos simplificados (apenas cnpj_cpf e nome)
    if 'cnpj_cpf' in campos_fornecedor and 'nome' in campos_fornecedor:
        for campo in ['cnpj_cpf', 'nome']:
            campos_fornecedor[campo].config(state='normal')
            campos_fornecedor[campo].delete(0, tk.END)
            idx = 0 if campo == 'cnpj_cpf' else 1
            campos_fornecedor[campo].insert(0, str(fornecedor[idx]))
            campos_fornecedor[campo].config(state='readonly')
        return fornecedor
    
    # Limpar e preencher campos básicos
    for entry in campos_fornecedor.values():
        entry.config(state='normal')
        entry.delete(0, tk.END)

    campos_fornecedor['cnpj_cpf'].insert(0, str(fornecedor[0]).zfill(14))
    campos_fornecedor['nome'].insert(0, fornecedor[1])
    campos_fornecedor['categoria'].insert(0, fornecedor[2])

    # Tratamento de dados bancários se necessário
    if 'dados_bancarios' in campos_fornecedor and buscar_fornecedor_completo:
        campos_fornecedor['dados_bancarios'].config(state='normal')
        campos_fornecedor['dados_bancarios'].delete(0, tk.END)

        tp_desp = campos_despesa['tp_desp'].get().strip() if campos_despesa else ''
        fornecedor_completo = buscar_fornecedor_completo(fornecedor[0])

        if fornecedor_completo:
            if tp_desp not in ['3', '5']:
                if fornecedor_completo['chave_pix']:
                    dados_bancarios = f"PIX: {fornecedor_completo['chave_pix']}"
                else:
                    dados_bancarios = (f"{fornecedor_completo['banco'] or ''} "
                                     f"{fornecedor_completo['op'] or ''} - "
                                     f"{fornecedor_completo['agencia'] or ''} "
                                     f"{fornecedor_completo['conta'] or ''}").strip()
                if dados_bancarios.strip() in ['', ' - ']:
                    dados_bancarios = 'DADOS BANCÁRIOS NÃO CADASTRADOS'
            else:
                dados_bancarios = ''
            
            campos_fornecedor['dados_bancarios'].insert(0, dados_bancarios)

    # Configurar estados finais
    for campo, entry in campos_fornecedor.items():
        if campo != 'categoria':
            entry.config(state='readonly')

    # Mudar para próxima aba se necessário
    if notebook:
        notebook.select(2)

    return fornecedor

# === CONSTANTS ===
DIAS_QUINZENA = [5, 20]
TIPOS_DESPESA = {
    1: "Despesas com Colaboradores",
    2: "Transferências Programadas",
    3: "Boletos",
    4: "Ressarcimentos",
    5: "Despesas Pagas pelo Cliente",
    6: "Pagamentos Caixa",
    7: "Administração"
}



